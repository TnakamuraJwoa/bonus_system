from django.shortcuts import render
# Create your views here.
from django.contrib.auth.views import LoginView
from allauth.account.forms import LoginForm
import csv
import hashlib
import io
import json
import logging
import time
from decimal import Decimal, InvalidOperation
from django.core.cache import cache
from django.views import generic
from django.contrib import messages
from .forms import InquiryForm
from django.urls import reverse, reverse_lazy
from django.shortcuts import render
from datetime import date
from datetime import datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from django.db import connections, transaction
from django.shortcuts import redirect
import math
import unicodedata
from urllib.parse import urlencode
from django.db import connections, transaction, IntegrityError, OperationalError
import traceback
from django.http import HttpResponse
from openpyxl import Workbook
import openpyxl

from django.db.models import Q, Sum
from django.utils.timezone import make_aware
from .models import TitleMaster, PeriodMaster, UserTitles, Orders, User, PurchaseInfoList, MonthlyPeriod
from .models import Settings
from .business_search_registration import (
    MONTH_PERSONAL_RESULT_TABLE,
    WEEK_PERSONAL_RESULT_TABLE,
    build_kibetu_condition,
    fetch_registration_history_rows,
    join_kibetu_list,
    parse_kibetu_list,
    resolve_default_kibetu_list,
)

from connect.placement_tree_builder import build_member_tree_view, fetch_tree_search_path
from connect.sql.week_bonus_sql import WEEK_BONUS_SQL
from connect.sql.month_bonus_sql import MONTH_BONUS_SQL
from connect.sql.month_title_sql import MONTH_TITLE_SQL

from connect.sql.drive_bonus_sql import DRIVE_BONUS_SQL
from connect.sql.basic_bonus_sql import BASIC_BONUS_SQL
from connect.sql.matching_bonus_sql import MATCHING_BONUS_SQL
from connect.sql.title_bonus_sql import TITLE_BONUS_SQL
from connect.sql.title_diff_bonus_sql import TITLE_DIFF_BONUS_SQL
from connect.sql.repurchase_over_bonus_sql import REPURCHASE_OVER_BONUS_SQL
from connect.sql.three_star_diamond_global_bonus_q_sql import THREE_STAR_DIAMOND_GLOBAL_BONUS_Q_SQL
from connect.sql.crown_diamond_global_bonus_y_sql import CROWN_DIAMOND_GLOBAL_BONUS_Y_SQL

from connect.sql.repurchase_last_month_sql import REPURCHASE_LAST_MONTH
from connect.sql.basic_bv_line_sql import BASIC_BV_LINE_SQL


from connect.sql import register_sql

from accounts.access import get_user_access
from connect.audit import fetch_one_dict, record_change_audit
from connect.bonus_help import list_bonus_help, save_bonus_help
from connect.sql.placement_tree_sql import PLACEMENT_TREE_REBUILD_CACHE_SQL
from connect.templatetags.custom_filters import jst_datetime

logger = logging.getLogger(__name__)


def _excel_jst_datetime(value):
    value = jst_datetime(value)
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def insert_bonus_register_history(bonus_name, kibetu, username, comment_text):
    """結果データが0件の場合も、登録操作の履歴を残す。"""
    with connections["rds"].cursor() as cursor:
        logger.info(
            "ボーナス登録履歴登録: bonus_name=%s kibetu=%s comment=%s",
            bonus_name,
            kibetu,
            comment_text,
        )
        cursor.execute(
            """
                INSERT INTO bonus_db.bonus_register_history (
                    bonus_name,
                    kibetu,
                    registered_at,
                    registered_by,
                    comment_text
                )
                VALUES (
                    %s,
                    %s,
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                    %s,
                    %s
                )
            """,
            [
                bonus_name,
                kibetu,
                username,
                comment_text,
            ],
        )
    record_change_audit(
        None,
        screen_name="ボーナス登録",
        action_type="execute",
        target_table="bonus_register_history",
        target_pk=f"{bonus_name}:{kibetu}",
        summary=comment_text,
        before_values=None,
        after_values={
            "bonus_name": bonus_name,
            "kibetu": kibetu,
            "registered_by": username,
            "comment_text": comment_text,
        },
        changed_by=username,
    )


def get_rds_jst_now():
    with connections["rds"].cursor() as cursor:
        cursor.execute("SELECT CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')")
        row = cursor.fetchone()
    return row[0] if row else None


def delete_auto_register_history_after(kibetu, bonus_names, username, started_at):
    """事前登録失敗時に、今回作成した登録履歴だけ削除する。"""
    if not kibetu or not bonus_names or not started_at:
        return False

    placeholders = ", ".join(["%s"] * len(bonus_names))
    sql = f"""
        DELETE FROM bonus_db.bonus_register_history
        WHERE kibetu = %s
          AND bonus_name IN ({placeholders})
          AND registered_by = %s
          AND registered_at >= %s
          AND comment_text LIKE %s
    """
    params = [
        kibetu,
        *bonus_names,
        username,
        started_at,
        "%ボーナス表示前の自動登録:%",
    ]

    try:
        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.execute(sql, params)
                deleted_count = cursor.rowcount
    except Exception:
        logger.exception(
            "事前登録失敗後の登録履歴削除に失敗しました: kibetu=%s bonus_names=%s",
            kibetu,
            ",".join(bonus_names),
        )
        return False

    logger.info(
        "事前登録失敗により登録履歴を削除しました: kibetu=%s bonus_names=%s deleted=%s",
        kibetu,
        ",".join(bonus_names),
        deleted_count,
    )
    return True


def has_month_title_rows(selected_kibetu):
    with connections["rds"].cursor() as cursor:
        cursor.execute(
            """
                SELECT 1
                FROM bonus_db.month_title
                WHERE kibetu = %s
                LIMIT 1
            """,
            [selected_kibetu],
        )
        return cursor.fetchone() is not None


def warn_month_title_required(request, action_label="計算"):
    messages.warning(
        request,
        f"{action_label}するには、同じ期別の月タイトルを先に計算・登録してください。",
    )


def normalize_bonus_histry_rows(rows):
    for row in rows:
        for key, value in list(row.items()):
            if key.endswith("_is_empty"):
                row[key] = bool(value)
    return rows


def insert_empty_bonus_history_on_display(request, bonus_name, kibetu):
    """表示時に0件だった場合も、同日の重複を避けて履歴を残す。"""
    if not kibetu:
        return

    with transaction.atomic(using="rds"):
        with connections["rds"].cursor() as cursor:
            cursor.execute(
                """
                    SELECT 1
                    FROM bonus_db.bonus_register_history
                    WHERE bonus_name = %s
                      AND kibetu = %s
                      AND DATE(registered_at) = DATE(CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'))
                      AND comment_text LIKE '0件%%'
                    LIMIT 1
                """,
                [
                    bonus_name,
                    kibetu,
                ],
            )
            if cursor.fetchone():
                return

        insert_bonus_register_history(
            bonus_name,
            kibetu,
            request.user.username,
            "0件表示（対象データなし）",
        )


def get_previous_month_from_kibetu(selected_kibetu):
    kibetu_year = int(selected_kibetu[0:4])
    kibetu_month = int(selected_kibetu[5:7])
    current_month_first = datetime(kibetu_year, kibetu_month, 1)
    prev_month_last = current_month_first - timedelta(days=1)
    return prev_month_last.year, prev_month_last.month


def get_user_target_rank_cutoff(year, month):
    return datetime(year, month, 1, 0, 0, 0) + relativedelta(months=1)


def get_user_add_rank_setting():
    with connections["rds"].cursor() as cursor:
        cursor.execute(
            """
                SELECT value
                FROM bonus_db.settings
                WHERE name = 'user_add_rank'
                LIMIT 1
            """
        )
        row = cursor.fetchone()
    return str(row[0]) if row and row[0] is not None else ""


def has_user_target_rank_history(kibetu):
    with connections["rds"].cursor() as cursor:
        cursor.execute(
            """
                SELECT 1
                FROM bonus_db.bonus_register_history
                WHERE bonus_name = 'user_target_rank'
                  AND kibetu = %s
                LIMIT 1
            """,
            [kibetu],
        )
        return cursor.fetchone() is not None


def register_users_target_rank(year, month):
    cutoff_dt = get_user_target_rank_cutoff(year, month)
    target_rank = f"{year}{month:02d}"

    with connections["rds"].cursor() as cursor:
        cursor.execute("TRUNCATE TABLE bonus_db.users_target_rank")
        cursor.execute(register_sql.USERS_TARGET_RANK_INSERT_SQL, [cutoff_dt])
        inserted_count = cursor.rowcount
        cursor.execute(
            """
                UPDATE bonus_db.settings
                SET value = %s
                WHERE name = 'user_add_rank'
            """,
            [target_rank],
        )

    return inserted_count, target_rank


def ensure_user_target_rank_for_kibetu(request, selected_kibetu):
    """ドライブボーナス計算用に、期別の前月ユーザーランクを用意する。"""
    try:
        year, month = get_previous_month_from_kibetu(selected_kibetu)
        required_rank = f"{year}{month:02d}"
        current_rank = get_user_add_rank_setting()
        has_history = has_user_target_rank_history(required_rank)

        if current_rank == required_rank:
            if not has_history:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "user_target_rank",
                        required_rank,
                        request.user.username,
                        (
                            "ドライブボーナス計算前の自動登録: "
                            f"登録済みのため更新スキップ（{required_rank}）"
                        ),
                    )
            return True

        with transaction.atomic(using="rds"):
            inserted_count, target_rank = register_users_target_rank(year, month)
            insert_bonus_register_history(
                "user_target_rank",
                target_rank,
                request.user.username,
                (
                    "ドライブボーナス計算前の自動登録: "
                    f"{year}年{month}月（{target_rank}） {inserted_count}件登録"
                ),
            )

        messages.info(
            request,
            f"ドライブボーナス計算前に、{year}年{month}月（{required_rank}）のユーザーランクを自動登録しました。",
        )
        return True
    except Exception:
        logger.exception(
            "ユーザーランク自動登録エラー: kibetu=%s",
            selected_kibetu,
        )
        messages.error(
            request,
            "ドライブボーナス計算前のユーザーランク自動登録に失敗しました。",
        )
        return False


def parse_input_date(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("/", "-")
    parts = normalized.split("-")
    if len(parts) != 3:
        raise ValueError(
            f'"{text}" は無効な日付形式です。YYYY-MM-DD または YYYY/MM/DD 形式にしてください。'
        )

    try:
        year, month, day = (int(parts[0]), int(parts[1]), int(parts[2]))
        return date(year, month, day)
    except ValueError as exc:
        raise ValueError(
            f'"{text}" は無効な日付形式です。YYYY-MM-DD または YYYY/MM/DD 形式にしてください。'
        ) from exc


def parse_optional_input_date(value):
    text = (value or "").strip()
    if not text:
        return None
    return parse_input_date(text)


def _build_kibetu_filter_redirect(request, url_name):
    params = {}
    filter_kibetu = (
        request.POST.get("filter_kibetu") or request.GET.get("kibetu") or ""
    ).strip()
    filter_q_kibetu = (
        request.POST.get("filter_q_kibetu") or request.GET.get("q_kibetu") or ""
    ).strip()
    if filter_kibetu:
        params["kibetu"] = filter_kibetu
    if filter_q_kibetu:
        params["q_kibetu"] = filter_q_kibetu
    url = reverse(url_name)
    if params:
        url += "?" + urlencode(params)
    return url


class BonusHelpTextView(generic.TemplateView):
    template_name = "help_text.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = list_bonus_help()
        ctx["total_count"] = len(ctx["rows"])
        return ctx

    def post(self, request, *args, **kwargs):
        user_access = get_user_access(request.user)
        if not user_access.can_menu("help_text") or not user_access.can_update:
            return HttpResponse("権限がありません。", status=403)

        help_key = (request.POST.get("help_key") or "").strip()
        title = (request.POST.get("title") or "").strip()
        content = request.POST.get("content") or ""

        try:
            before_rows = list_bonus_help()
            before_row = next((row for row in before_rows if row.get("help_key") == help_key), None)
            save_bonus_help(help_key, title, content)
            record_change_audit(
                request,
                screen_name="ヘルプテキスト",
                action_type="update",
                target_table="bonus_help_text",
                target_pk=help_key,
                summary=f"{help_key} のヘルプテキストを保存",
                before_values=before_row,
                after_values={
                    "help_key": help_key,
                    "title": title,
                    "content": content,
                },
            )
            messages.success(request, "ヘルプテキストを保存しました。")
        except ValueError as exc:
            messages.error(request, str(exc))
        except Exception:
            logger.exception("ヘルプテキストの保存に失敗しました。help_key=%s", help_key)
            messages.error(request, "ヘルプテキストの保存に失敗しました。")

        return redirect("connect:help_text")


def get_week_purchase_check_months(selected_kibetu, period=None):
    if period and getattr(period, "st_date", None) and getattr(period, "end_date", None):
        st_date = period.st_date
        end_date = period.end_date
    else:
        weekly_periods = list(
            PeriodMaster.objects.using("rds")
            .filter(kibetu__startswith=f"{selected_kibetu}W")
            .exclude(st_date__isnull=True)
            .exclude(end_date__isnull=True)
            .order_by("st_date", "end_date")
        )
        if not weekly_periods:
            return []

        st_date = weekly_periods[0].st_date
        end_date = max(p.end_date for p in weekly_periods)

    target_dates = [
        st_date - relativedelta(months=1),
        st_date,
        end_date,
    ]

    months = []
    seen = set()
    for target_date in target_dates:
        year_month = (target_date.year, target_date.month)
        if year_month not in seen:
            seen.add(year_month)
            months.append(year_month)

    return months


def has_purchase_info_for_month(register_year, register_month):
    with connections["rds"].cursor() as cursor:
        cursor.execute(
            """
            SELECT 1
            FROM bonus_db.purchase_info_list
            WHERE register_year = %s
              AND register_month = %s
            LIMIT 1
            """,
            [register_year, register_month],
        )
        return cursor.fetchone() is not None


def fetch_purchase_info_rows_for_month(
    year,
    month,
    register_year=None,
    register_month=None,
):
    register_year = register_year or year
    register_month = register_month or month

    start = datetime(year, month, 1)
    end = start + relativedelta(months=1)
    params = [
        register_year,
        register_month,
        start,
        end,
        register_year,
        register_month,
        year,
        month,
    ]

    with connections["rds"].cursor() as cursor:
        cursor.execute(REPURCHASE_LAST_MONTH, params)
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, r)) for r in cursor.fetchall()]


def insert_purchase_info_rows(rows):
    insert_sql = """
INSERT INTO bonus_db.purchase_info_list
(
    register_year,
    register_month,
    order_year,
    order_month,
    jwoa_code,
    send_bv_name,
    order_code,
    total_bv,
    bv,
    order_type,
    deposit_at,
    order_at,
    bonus_payment_date
)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

    data = [
        (
            r["register_year"],
            r["register_month"],
            r["order_year"],
            r["order_month"],
            r["jwoa_code"],
            r["send_bv_name"],
            r["order_code"],
            r["total_bv"],
            r["bv"],
            r["order_type"],
            r["deposit_at"],
            r["order_at"],
            r["payment_date"],
        )
        for r in rows
    ]

    with connections["rds"].cursor() as cursor:
        cursor.executemany(insert_sql, data)


def auto_register_purchase_info_for_months(request, target_months):
    no_source_months = []
    registered_months = []

    for year, month in target_months:
        if has_purchase_info_for_month(year, month):
            continue

        rows = fetch_purchase_info_rows_for_month(year, month)
        if not rows:
            no_source_months.append((year, month))
            continue

        with transaction.atomic(using="rds"):
            insert_purchase_info_rows(rows)
        registered_months.append((year, month, len(rows)))

    if registered_months:
        registered_text = "、".join(
            f"{year}年{month}月({count}件)"
            for year, month, count in registered_months
        )
        messages.success(request, f"購入情報を自動登録しました: {registered_text}")

    if no_source_months:
        missing_text = "、".join(
            f"{year}年{month}月"
            for year, month in no_source_months
        )
        messages.error(
            request,
            f"購入情報の元データがないため登録できません: {missing_text}",
        )
        return False

    return True


def auto_register_purchase_info_for_kibetu_month(
    request,
    register_year,
    register_month,
):
    if has_purchase_info_for_month(register_year, register_month):
        return True

    rows = fetch_purchase_info_rows_for_month(register_year, register_month)
    if rows:
        with transaction.atomic(using="rds"):
            insert_purchase_info_rows(rows)

        messages.success(
            request,
            (
                "購入情報を自動登録しました: "
                f"{register_year}年{register_month}月({len(rows)}件)"
            ),
        )
        return True

    messages.error(
        request,
        (
            "購入情報の元データがないため登録できません: "
            f"{register_year}年{register_month}月"
        ),
    )
    return False


def get_kibetu_register_year_month(selected_kibetu):
    return int(selected_kibetu[0:4]), int(selected_kibetu[5:7])


def ensure_kibetu_purchase_info(request, selected_kibetu, period=None):
    try:
        register_year, register_month = get_kibetu_register_year_month(
            selected_kibetu
        )
    except (TypeError, ValueError):
        messages.error(request, "期別から対象年月を判定できません。")
        return False

    return auto_register_purchase_info_for_kibetu_month(
        request,
        register_year,
        register_month,
    )


def parse_target_month(value):
    text = (value or "").strip()
    if not text:
        raise ValueError

    if len(text) >= 7 and text[4].upper() == "C":
        year = int(text[0:4])
        month = int(text[5:7])
    else:
        year, month = map(int, text.split("-"))

    if month < 1 or month > 12:
        raise ValueError

    return year, month


def format_target_month(year, month):
    return f"{year}-{month:02d}"


def format_target_kibetu(year, month):
    return f"{year}C{month:02d}"


def get_target_year_month_from_params(params):
    target_year = (params.get("target_year") or "").strip()
    target_month = (params.get("target_month") or "").strip()

    if target_year and target_month and "-" not in target_month:
        year = int(target_year)
        month = int(target_month)
    else:
        selected = (
            params.get("target_month_choice")
            or params.get("target_month")
            or ""
        ).strip()
        year, month = parse_target_month(selected)

    if month < 1 or month > 12:
        raise ValueError

    return year, month


def ensure_week_purchase_info(request, selected_kibetu, period=None):
    target_months = get_week_purchase_check_months(selected_kibetu, period)
    if not target_months:
        messages.error(request, "期別から週次の対象期間を判定できません。")
        return False

    return auto_register_purchase_info_for_months(request, target_months)


class KeysetPaginationMixin:

    DEFAULT_PER_PAGE = 200
    MAX_PER_PAGE = 500

    def get_per_page(self):

        try:
            per_page = int(
                self.request.GET.get(
                    "per_page",
                    self.DEFAULT_PER_PAGE
                )
            )

        except ValueError:
            per_page = self.DEFAULT_PER_PAGE

        return max(
            1,
            min(per_page, self.MAX_PER_PAGE)
        )

    def get_page_number(self, total_pages):

        try:
            page = int(self.request.GET.get("page") or "1")
        except ValueError:
            page = 1

        return max(1, min(page, total_pages))

    def build_base_qs(self, params):

        clean_params = {}

        for key, value in params.items():

            if value not in ["", None]:
                clean_params[key] = value

        return urlencode(clean_params)

    @staticmethod
    def build_pagination_pages(current_page, total_pages, adjacent=3):

        if total_pages <= 1:
            return []

        pages = []

        if current_page > adjacent + 1:
            pages.append(1)
            pages.append(None)

        start = max(1, current_page - adjacent)
        end = min(total_pages, current_page + adjacent)

        for num in range(start, end + 1):
            pages.append(num)

        if current_page < total_pages - adjacent:
            pages.append(None)
            pages.append(total_pages)

        return pages

    def set_page_context(
        self,
        ctx,
        rows,
        per_page,
        total_count,
        total_pages,
        page,
        base_params,
    ):

        ctx["rows"] = rows
        ctx["total_count"] = total_count
        ctx["per_page"] = per_page
        ctx["page"] = page
        ctx["total_pages"] = total_pages
        if total_count > 0:
            ctx["display_from"] = (page - 1) * per_page + 1
            ctx["display_to"] = min(page * per_page, total_count)
        else:
            ctx["display_from"] = 0
            ctx["display_to"] = 0
        ctx["base_qs"] = self.build_base_qs(base_params)
        ctx["has_prev"] = page > 1
        ctx["has_next"] = page < total_pages
        ctx["prev_page"] = page - 1
        ctx["next_page"] = page + 1
        ctx["pagination_pages"] = self.build_pagination_pages(page, total_pages)

        return ctx


def get_bonus_sort_context(request, allowed_sort_columns, default_sort="id", default_direction="asc"):
    sort = request.GET.get("sort", default_sort)
    direction = request.GET.get("direction", default_direction)

    if sort not in allowed_sort_columns:
        sort = default_sort
        direction = default_direction

    if direction not in ("asc", "desc"):
        direction = default_direction

    order_column = allowed_sort_columns[sort]
    order_direction = "DESC" if direction == "desc" else "ASC"
    next_direction = "desc" if direction == "asc" else "asc"

    return {
        "sort": sort,
        "direction": direction,
        "next_direction": next_direction,
        "order_sql": f"{order_column} {order_direction}",
    }


def add_sort_params(base_params, sort_ctx):
    """ページ送りリンクに現在の並び替え条件を引き継ぐ。"""
    base_params["sort"] = sort_ctx["sort"]
    base_params["direction"] = sort_ctx["direction"]
    return base_params


def apply_like_filters(sql, params, request, field_map):
    """GETパラメータの部分一致（LIKE）条件をSQLに追加する。"""
    filter_values = {}
    for param, column in field_map.items():
        value = request.GET.get(param, "").strip()
        filter_values[param] = value
        if value:
            sql += f"\n            AND {column} LIKE %s"
            params.append(f"%{value}%")
    return sql, filter_values


def build_bonus_export_filename(base_name, kibetu=None, kibetu_list=None):
    """Excel出力ファイル名（期別付き）"""
    name = base_name
    if kibetu_list:
        valid = [str(k) for k in kibetu_list if k]
        if valid:
            name += "_" + "_".join(valid)
    elif kibetu:
        name += f"_{kibetu}"
    return f"{name}.xlsx"


def _format_export_cell(value, fmt=None):
    if value is None or value == "":
        return ""
    if fmt == "int":
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return value
    if fmt == "decimal2":
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    return value


def export_search_rows_to_excel(rows, columns, sheet_title, filename):
    """検索画面の表示列定義どおりに Excel を出力する。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    ws.append([col[0] for col in columns])

    for row in rows:
        ws.append([
            _format_export_cell(
                row.get(col[1]),
                col[2] if len(col) > 2 else None,
            )
            for col in columns
        ])

    for col_idx, col in enumerate(columns, start=1):
        fmt = col[2] if len(col) > 2 else None
        if fmt == "int":
            number_format = "#,##0"
        elif fmt == "decimal2":
            number_format = "#,##0.00"
        else:
            continue
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# 検索画面の一覧表示と同じ列（label, field_key, optional_format）
SEARCH_EXPORT_COLUMNS = {
    "drive_bonus": [
        ("期別", "kibetu"),
        ("紹介者タイトル", "title_name"),
        ("紹介者ID", "introducer_code"),
        ("会員ID", "jwoa_code"),
        ("会員名", "jwoa_name"),
        ("BV合計", "sum_bv", "int"),
        ("レート", "rate", "decimal2"),
        ("報酬", "sum_bonus_amount", "decimal2"),
    ],
    "basic_bonus": [
        ("期別", "kibetu"),
        ("上位者コード", "placement_code"),
        ("上位者名", "placement_name"),
        ("上位者ランク", "placement_rank"),
        ("ラインコード", "line_code"),
        ("購入者コード", "purchaser_code"),
        ("購入者名", "purchaser_name"),
        ("BV合計", "sum_bv", "int"),
        ("レート", "bonus_rate", "decimal2"),
        ("ボーナス金額", "bonus_amount", "decimal2"),
        ("ブルーダイヤ判定", "blue_daiya_flg"),
        ("作成日時", "created_at"),
    ],
    "matching_bonus": [
        ("期別", "kibetu"),
        ("紹介者コード", "introducer_code"),
        ("紹介者名", "introducer_name"),
        ("ライン上位コード", "line_code"),
        ("ベーシック取得者コード", "basic_code"),
        ("ベーシック取得者名", "basic_name"),
        ("対象ベーシックボーナス", "basic_bonus", "decimal2"),
        ("マッチングボーナス", "matching_bonus", "decimal2"),
        ("マッチングレベル", "level", "int"),
        ("紹介者ツリー階層", "tree_level", "int"),
        ("直紹介アクティブ人数", "active_count", "int"),
        ("作成日時", "created_at"),
    ],
    "title_bonus": [
        ("kibetu", "kibetu"),
        ("root_jwoa_code", "root_jwoa_code"),
        ("root_name", "root_name"),
        ("up_jwoa_code", "up_jwoa_code"),
        ("down_jwoa_code", "down_jwoa_code"),
        ("down_name", "down_name"),
        ("tree_level", "tree_level", "int"),
        ("match_level", "match_level", "int"),
        ("title_name", "タイトル名"),
        ("sum_bv", "sum_bv", "int"),
        ("rate", "rate", "decimal2"),
        ("bonus_amount", "bonus_amount", "decimal2"),
        ("created_at", "created_at"),
    ],
    "title_diff_bonus": [
        ("期別", "kibetu"),
        ("root_title_name", "root_title_name"),
        ("root_bonus_rate", "root_bonus_rate", "decimal2"),
        ("root_jwoa_code", "root_jwoa_code"),
        ("root_name", "root_name"),
        ("down_title_name", "down_title_name"),
        ("down_bonus_rate", "down_bonus_rate", "decimal2"),
        ("down_jwoa_code", "down_jwoa_code"),
        ("down_name", "down_name"),
        ("pay_bonus_rate", "pay_bonus_rate", "decimal2"),
        ("sum_bv", "sum_bv", "int"),
        ("title_diff_bonus", "title_diff_bonus", "decimal2"),
        ("created_at", "created_at"),
        ("updated_at", "updated_at"),
    ],
    "repurchase_over_bonus": [
        ("kibetu", "kibetu"),
        ("root_code", "root_code"),
        ("root_name", "root_name"),
        ("down_code", "down_code"),
        ("down_name", "down_name"),
        ("tree_level", "tree_level", "int"),
        ("match_count", "match_count", "int"),
        ("rate", "rate", "decimal2"),
        ("sum_bv", "sum_bv", "int"),
        ("over_bonus", "over_bonus", "decimal2"),
        ("created_at", "created_at"),
        ("updated_at", "updated_at"),
    ],
    "three_star_global_bonus": [
        ("kibetu", "kibetu"),
        ("jwoa_code", "jwoa_code"),
        ("jwoa_name", "jwoa_name"),
        ("title_name", "title_name"),
        ("score", "score", "int"),
        ("total_score", "total_score", "int"),
        ("total_over_bv", "total_over_bv", "decimal2"),
        ("bonus_amount", "bonus_amount", "decimal2"),
        ("created_at", "created_at"),
        ("updated_at", "updated_at"),
    ],
    "global_bonus": [
        ("kibetu", "kibetu"),
        ("jwoa_code", "jwoa_code"),
        ("jwoa_name", "jwoa_name"),
        ("title_name", "title_name"),
        ("score", "score", "int"),
        ("total_score", "total_score", "int"),
        ("total_bv", "total_bv", "decimal2"),
        ("bonus_amount", "bonus_amount", "decimal2"),
        ("created_at", "created_at"),
        ("updated_at", "updated_at"),
    ],
    "week_bonus": [
        ("期別", "kibetu"),
        ("会員コード", "jwoa_code"),
        ("会員名", "jwoa_name"),
        ("ドライブボーナス", "drive_bonus", "decimal2"),
        ("ベーシックボーナス", "basic_bonus", "decimal2"),
        ("マッチングボーナス", "matching_bonus", "decimal2"),
        ("週間ボーナス", "week_bonus", "decimal2"),
        ("決済時間", "updated_at"),
    ],
    "month_title": [
        ("期別", "kibetu"),
        ("会員コード", "jwoa_code"),
        ("会員名", "jwoa_name"),
        ("収入ライン系列BV", "income_line_bv", "int"),
        ("基本ライン系列BV", "basic_line_bv", "int"),
        ("タイトル", "title_name"),
        ("決済時間", "updated_at"),
    ],
    "month_bonus": [
        ("期別", "kibetu"),
        ("会員コード", "jwoa_code"),
        ("会員名", "jwoa_name"),
        ("タイトルボーナス", "title_bonus", "decimal2"),
        ("リピート購入オーバーボーナス", "repurchase_over_bonus", "decimal2"),
        ("差額ボーナス", "title_diff_bonus", "decimal2"),
        ("３つ星ダイヤグローバル配当", "three_star_diamond_global_bonus", "decimal2"),
        ("大使ダイヤグローバル配当", "crown_three_star_diamond_global_bonus", "decimal2"),
        ("月間ボーナス", "month_bonus", "decimal2"),
        ("決済時間", "updated_at"),
    ],
}


BONUS_RESULT_DELETE_CONFIG = {
    "drive_bonus": {
        "label": "ドライブボーナス",
        "result_tables": ["bonus_db.B_drive_bonus_result"],
        "history_names": ["drive_bonus"],
        "period_model": PeriodMaster,
        "redirect_name": "connect:drive_bonus",
    },
    "basic_bonus": {
        "label": "ベーシックボーナス",
        "result_tables": ["bonus_db.B_basic_bonus_result", "bonus_db.basic_bv_line"],
        "history_names": ["basic_bonus"],
        "period_model": PeriodMaster,
        "redirect_name": "connect:basic_bonus",
    },
    "matching_bonus": {
        "label": "マッチングボーナス",
        "result_tables": ["bonus_db.B_matching_bonus_result"],
        "history_names": ["matching_bonus"],
        "period_model": PeriodMaster,
        "redirect_name": "connect:matching_bonus",
    },
    "title_bonus": {
        "label": "タイトルボーナス",
        "result_tables": ["bonus_db.B_title_bonus_result"],
        "history_names": ["title_bonus"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:title_bonus",
    },
    "title_diff_bonus": {
        "label": "タイトル差額ボーナス",
        "result_tables": ["bonus_db.B_title_diff_bonus_result"],
        "history_names": ["title_diff_bonus"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:title_diff_bonus",
    },
    "repurchase_over_bonus": {
        "label": "再購入オーバーボーナス",
        "result_tables": ["bonus_db.B_repurchase_over_bonus_result"],
        "history_names": ["repurchase_over_bonus"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:repurchase_over_bonus",
    },
    "three_star_global_bonus": {
        "label": "3スターダイヤグローバル配当",
        "result_tables": ["bonus_db.B_three_star_global_bonus_result"],
        "history_names": ["three_star_global_bonus"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:three_star_global_bonus",
    },
    "global_bonus": {
        "label": "グローバル配当",
        "result_tables": ["bonus_db.B_global_bonus_result"],
        "history_names": ["global_bonus"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:global_bonus",
    },
    "week_bonus": {
        "label": "週間ボーナス",
        "result_tables": ["bonus_db.B_week_bonus_result"],
        "history_names": ["week_bonus"],
        "period_model": PeriodMaster,
        "redirect_name": "connect:week_bonus",
    },
    "month_title": {
        "label": "月タイトル",
        "result_tables": ["bonus_db.month_title"],
        "history_names": ["month_title", "title_registration"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:month_title",
    },
    "month_bonus": {
        "label": "月間ボーナス",
        "result_tables": ["bonus_db.B_month_bonus_result"],
        "history_names": ["month_bonus"],
        "period_model": MonthlyPeriod,
        "redirect_name": "connect:month_bonus",
    },
}


def _bonus_redirect_url(config, kibetu=None):
    base_url = redirect(config["redirect_name"]).url
    if kibetu:
        return f"{base_url}?kibetu={kibetu}"
    return base_url


def delete_bonus_result_for_kibetu(request, bonus_key):
    config = BONUS_RESULT_DELETE_CONFIG[bonus_key]
    selected_kibetu = (request.POST.get("kibetu") or "").strip()

    if not get_user_access(request.user).can_delete:
        messages.error(request, "削除権限がありません。")
        return redirect(config["redirect_name"])

    if not selected_kibetu:
        messages.error(request, "期別を選択してください。")
        return redirect(config["redirect_name"])

    period = (
        config["period_model"]
        .objects.using("rds")
        .filter(kibetu=selected_kibetu)
        .first()
    )
    if not period:
        messages.error(request, "選択された期別が存在しません。")
        return redirect(config["redirect_name"])

    redirect_url = _bonus_redirect_url(config, selected_kibetu)
    reset_redirect_url = _bonus_redirect_url(config)
    result_counts = {}
    deleted_counts = {}
    history_counts = {}
    history_deleted = 0

    try:
        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                for table_name in config["result_tables"]:
                    cursor.execute(
                        f"SELECT COUNT(*) FROM {table_name} WHERE kibetu = %s",
                        [selected_kibetu],
                    )
                    row = cursor.fetchone()
                    result_counts[table_name] = int(row[0]) if row else 0

                placeholders = ", ".join(["%s"] * len(config["history_names"]))
                cursor.execute(
                    f"""
                        SELECT bonus_name, COUNT(*)
                        FROM bonus_db.bonus_register_history
                        WHERE kibetu = %s
                          AND bonus_name IN ({placeholders})
                        GROUP BY bonus_name
                    """,
                    [selected_kibetu, *config["history_names"]],
                )
                history_counts = {row[0]: int(row[1]) for row in cursor.fetchall()}

                for table_name in config["result_tables"]:
                    logger.info(
                        "ボーナス計算結果削除SQLを実行します。table=%s kibetu=%s",
                        table_name,
                        selected_kibetu,
                    )
                    cursor.execute(
                        f"DELETE FROM {table_name} WHERE kibetu = %s",
                        [selected_kibetu],
                    )
                    deleted_counts[table_name] = cursor.rowcount

                cursor.execute(
                    f"""
                        DELETE FROM bonus_db.bonus_register_history
                        WHERE kibetu = %s
                          AND bonus_name IN ({placeholders})
                    """,
                    [selected_kibetu, *config["history_names"]],
                )
                history_deleted = cursor.rowcount

            record_change_audit(
                request,
                screen_name=config["label"],
                action_type="delete",
                target_table=", ".join(config["result_tables"]),
                target_pk=f"{bonus_key}:{selected_kibetu}",
                summary=(
                    f"{selected_kibetu} の{config['label']}登録済みデータを削除: "
                    f"結果{sum(deleted_counts.values())}件、履歴{history_deleted}件"
                ),
                before_values={
                    "bonus_key": bonus_key,
                    "kibetu": selected_kibetu,
                    "result_counts": result_counts,
                    "history_counts": history_counts,
                },
                after_values=None,
            )
    except Exception as e:
        logger.exception("ボーナス計算結果削除エラー。bonus_key=%s kibetu=%s", bonus_key, selected_kibetu)
        messages.error(request, f"削除中にエラーが発生しました: {e}")
        return redirect(redirect_url)

    total_deleted = sum(deleted_counts.values())
    if total_deleted or history_deleted:
        messages.success(
            request,
            f"{selected_kibetu} の{config['label']}登録済みデータを削除しました。"
            f"（結果{total_deleted}件、履歴{history_deleted}件）",
        )
    else:
        messages.warning(request, f"{selected_kibetu} の削除対象データはありませんでした。")
    return redirect(reset_redirect_url)


class IndexView(LoginView):
    template_name = "account/login.html"
    form_class = LoginForm


class DriveBonusView(generic.ListView):
    template_name = "drive_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster

    def get_queryset(self):

        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_drive_bonus_result
                ORDER BY kibetu DESC
            """)

            registered_kibetu_list = [
                row[0]
                for row in cursor.fetchall()
            ]

        if not registered_kibetu_list:
            return PeriodMaster.objects.using("rds").none()

        return (
            PeriodMaster.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-kibetu")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()
        save_basic_bv_line = request.POST.get("save_basic_bv_line", "1") == "1"

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "drive_bonus")

        if action != "register_drive_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:drive_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:drive_bonus")

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:drive_bonus")

        try:
            if not ensure_user_target_rank_for_kibetu(request, selected_kibetu):
                return redirect(f"/drive_bonus/?kibetu={selected_kibetu}")

            rows = self._get_drive_bonus_rows(selected_kibetu, period)

            if not rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "drive_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/drive_bonus/?kibetu={selected_kibetu}")

            insert_sql = """
                INSERT INTO bonus_db.B_drive_bonus_result (
                    kibetu,
                    title_name,
                    introducer_code,
                    jwoa_code,
                    jwoa_name,
                    rate,
                    sum_bv,
                    sum_bonus_amount,
                    created_at
                ) VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    NOW()
                )
                ON DUPLICATE KEY UPDATE
                    title_name = VALUES(title_name),
                    introducer_code = VALUES(introducer_code),
                    jwoa_name = VALUES(jwoa_name),
                    rate = VALUES(rate),
                    sum_bv = VALUES(sum_bv),
                    sum_bonus_amount = VALUES(sum_bonus_amount),
                    created_at = NOW()
            """

            insert_params = []
            for r in rows:
                insert_params.append([
                    selected_kibetu,
                    r.get("title_name") or "",
                    r.get("introducer_code") or "",
                    r.get("jwoa_code") or "",
                    r.get("jwoa_name") or "",
                    r.get("rate") or 0,
                    r.get("sum_bv") or 0,
                    r.get("sum_bonus_amount") or 0,
                ])

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    # ドライブボーナス登録
                    cursor.executemany(insert_sql, insert_params)

                    # 履歴登録
                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "drive_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(rows)}件登録"
                        ]
                    )

            messages.success(request, f"{len(rows)}件をドライブボーナス結果に登録しました。")

        except Exception as e:
            logger.exception("ドライブボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/drive_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_week_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:drive_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            return ctx

        ctx["selected_period"] = period
        if not ensure_week_purchase_info(self.request, selected_kibetu, period):
            insert_empty_bonus_history_on_display(
                self.request,
                "drive_bonus",
                selected_kibetu,
            )
            return ctx

        if not ensure_user_target_rank_for_kibetu(self.request, selected_kibetu):
            insert_empty_bonus_history_on_display(
                self.request,
                "drive_bonus",
                selected_kibetu,
            )
            return ctx

        ctx["rows"] = self._get_drive_bonus_rows(selected_kibetu, period)
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "drive_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_drive_bonus_rows(self, selected_kibetu, period):
        st_date = period.st_date
        end_date = period.end_date

        kibetu_year = int(selected_kibetu[0:4])
        kibetu_month = int(selected_kibetu[5:7])

        start_dt = make_aware(datetime.combine(st_date, time.min))
        end_dt = make_aware(datetime.combine(end_date + timedelta(days=1), time.min))

        current_month_first = datetime(kibetu_year, kibetu_month, 1)
        prev_month_last = current_month_first - timedelta(days=1)

        prev_year = prev_month_last.year
        prev_month = prev_month_last.month

        be_start_dt = make_aware(datetime(prev_year, prev_month, 1, 0, 0, 0))
        be_end_dt = make_aware(datetime(kibetu_year, kibetu_month, 1, 0, 0, 0))


        params = [
            prev_year,
            prev_month,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            be_start_dt,
            be_end_dt,
            be_end_dt,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(DRIVE_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows



class InquiryView(generic.FormView):
    template_name = "inquiry.html"
    form_class = InquiryForm
    success_url = reverse_lazy('connect:inquiry')

    def form_valid(self, form):
        form.send_email()
        messages.info(self.request, f'メッセージを送信しました')
        logger.info('Inquiry sent by {}'.format(form.cleaned_data['name']))
        return super().form_valid(form)



class KibetuView(generic.ListView):
    template_name = "kibetu.html"
    context_object_name = "rows"
    model = PeriodMaster

    def get_queryset(self):
        qs = PeriodMaster.objects.using("rds").all()

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        q_kibetu = (self.request.GET.get("q_kibetu") or "").strip()

        if selected_kibetu:
            qs = qs.filter(kibetu=selected_kibetu)

        if q_kibetu:
            qs = qs.filter(kibetu__icontains=q_kibetu)

        return qs.order_by("-st_date", "-kibetu")

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")

        if action == "bulk_update":
            return self._bulk_update(request)

        kibetu = (request.POST.get("kibetu") or "").strip()
        st_date = request.POST.get("st_date") or None
        end_date = request.POST.get("end_date") or None
        payment_date = request.POST.get("payment_date") or None

        if not kibetu:
            messages.error(request, "期別を入力してください。")
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu"))

        try:
            completion_date = parse_optional_input_date(
                request.POST.get("completion_date")
            )
        except ValueError as e:
            messages.error(request, str(e))
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu"))

        try:
            with transaction.atomic(using="rds"):

                if action == "create":
                    PeriodMaster.objects.using("rds").create(
                        kibetu=kibetu,
                        st_date=st_date,
                        end_date=end_date,
                        payment_date=payment_date,
                        completion_date=completion_date,
                    )
                    record_change_audit(
                        request,
                        screen_name="期別(週)",
                        action_type="create",
                        target_table="period_master",
                        target_pk=kibetu,
                        summary=f"{kibetu} を追加",
                        before_values=None,
                        after_values={
                            "kibetu": kibetu,
                            "st_date": st_date,
                            "end_date": end_date,
                            "payment_date": payment_date,
                            "completion_date": completion_date,
                        },
                    )
                    messages.success(request, f"{kibetu} を追加しました。")

                elif action == "update":
                    obj = PeriodMaster.objects.using("rds").get(kibetu=kibetu)
                    before_values = {
                        "kibetu": obj.kibetu,
                        "st_date": obj.st_date,
                        "end_date": obj.end_date,
                        "payment_date": obj.payment_date,
                        "completion_date": obj.completion_date,
                    }
                    obj.st_date = st_date
                    obj.end_date = end_date
                    obj.payment_date = payment_date
                    obj.completion_date = completion_date
                    obj.save(using="rds")
                    record_change_audit(
                        request,
                        screen_name="期別(週)",
                        action_type="update",
                        target_table="period_master",
                        target_pk=kibetu,
                        summary=f"{kibetu} を変更",
                        before_values=before_values,
                        after_values={
                            "kibetu": kibetu,
                            "st_date": st_date,
                            "end_date": end_date,
                            "payment_date": payment_date,
                            "completion_date": completion_date,
                        },
                    )
                    messages.success(request, f"{kibetu} を変更しました。")

                elif action == "delete":
                    before_obj = PeriodMaster.objects.using("rds").filter(kibetu=kibetu).first()
                    before_values = None
                    if before_obj:
                        before_values = {
                            "kibetu": before_obj.kibetu,
                            "st_date": before_obj.st_date,
                            "end_date": before_obj.end_date,
                            "payment_date": before_obj.payment_date,
                            "completion_date": before_obj.completion_date,
                        }
                    deleted_count, _ = PeriodMaster.objects.using("rds").filter(kibetu=kibetu).delete()
                    if deleted_count:
                        record_change_audit(
                            request,
                            screen_name="期別(週)",
                            action_type="delete",
                            target_table="period_master",
                            target_pk=kibetu,
                            summary=f"{kibetu} を削除",
                            before_values=before_values,
                            after_values=None,
                        )
                    messages.success(request, f"{kibetu} を削除しました。")

                else:
                    messages.error(request, "不正な操作です。")

        except PeriodMaster.DoesNotExist:
            messages.error(request, f"{kibetu} は存在しません。")
        except IntegrityError:
            messages.error(request, f"{kibetu} はすでに存在します。")
        except Exception as e:
            messages.error(request, f"エラーが発生しました: {e}")

        return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu"))

    def _bulk_update(self, request):
        kibetus = request.POST.getlist("bulk_kibetu")
        st_dates = request.POST.getlist("bulk_st_date")
        end_dates = request.POST.getlist("bulk_end_date")
        payment_dates = request.POST.getlist("bulk_payment_date")
        completion_dates = request.POST.getlist("bulk_completion_date")
        row_lengths = {
            len(kibetus),
            len(st_dates),
            len(end_dates),
            len(payment_dates),
            len(completion_dates),
        }

        if not kibetus or len(row_lengths) != 1:
            messages.error(request, "一括変更するデータがありません。")
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu"))

        updated_count = 0
        errors = []

        try:
            with transaction.atomic(using="rds"):
                for idx, kibetu in enumerate(kibetus):
                    kibetu = (kibetu or "").strip()
                    if not kibetu:
                        continue

                    try:
                        completion_date = parse_optional_input_date(
                            completion_dates[idx]
                        )
                    except ValueError as e:
                        errors.append(f"{kibetu}: {e}")
                        continue

                    try:
                        obj = PeriodMaster.objects.using("rds").get(kibetu=kibetu)
                    except PeriodMaster.DoesNotExist:
                        errors.append(f"{kibetu}: 存在しません。")
                        continue

                    obj.st_date = (st_dates[idx] or "").strip() or None
                    obj.end_date = (end_dates[idx] or "").strip() or None
                    obj.payment_date = (payment_dates[idx] or "").strip() or None
                    obj.completion_date = completion_date
                    obj.save(using="rds")
                    updated_count += 1

        except Exception as e:
            messages.error(request, f"エラーが発生しました: {e}")
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu"))

        if errors:
            messages.error(request, " / ".join(errors))
        if updated_count:
            record_change_audit(
                request,
                screen_name="期別(週)",
                action_type="bulk_update",
                target_table="period_master",
                target_pk=None,
                summary=f"期別(週)を {updated_count}件 一括変更",
                before_values=None,
                after_values={"count": updated_count},
            )
            messages.success(request, f"{updated_count} 件を変更しました。")
        elif not errors:
            messages.info(request, "変更対象がありませんでした。")

        return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu"))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["selected_kibetu"] = (self.request.GET.get("kibetu") or "").strip()
        ctx["q_kibetu"] = (self.request.GET.get("q_kibetu") or "").strip()

        ctx["kibetu_choices"] = list(
            PeriodMaster.objects.using("rds")
            .order_by("-st_date")
            .values_list("kibetu", flat=True)
            .distinct()
        )

        return ctx



class TitleListView(generic.ListView):
    template_name = "title_list.html"
    context_object_name = "rows"
    model = TitleMaster

    def get_queryset(self):
        # 並び順はお好みで（title_id順など）
        return TitleMaster.objects.using("rds").order_by("title_id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # 件数表示用（テンプレで rows|length を使わない）
        ctx["total_count"] = ctx["rows"].count()
        return ctx


class KibetuMonthView(generic.ListView):
    template_name = "kibetu_month.html"
    context_object_name = "rows"
    model = MonthlyPeriod

    def get_queryset(self):
        qs = MonthlyPeriod.objects.using("rds").all()

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        q_kibetu = (self.request.GET.get("q_kibetu") or "").strip()

        if selected_kibetu:
            qs = qs.filter(kibetu=selected_kibetu)

        if q_kibetu:
            qs = qs.filter(kibetu__icontains=q_kibetu)

        return qs.order_by("-year", "-month", "-kibetu")

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")

        if action == "bulk_update":
            return self._bulk_update(request)

        kibetu = (request.POST.get("kibetu") or "").strip()
        year = request.POST.get("year") or ""
        month = request.POST.get("month") or ""
        payment_date = request.POST.get("payment_date") or ""

        create_form = {
            "kibetu": kibetu,
            "year": year,
            "month": month,
            "payment_date": payment_date,
        }

        def render_create_form_on_error():
            if action != "create":
                return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu_month"))
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context["create_form"] = create_form
            return self.render_to_response(context)

        if not kibetu:
            messages.error(request, "期別を入力してください。")
            return render_create_form_on_error()

        try:
            parsed_payment_date = parse_input_date(payment_date)
        except ValueError as e:
            messages.error(request, str(e))
            if action == "create":
                return render_create_form_on_error()
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu_month"))

        try:
            with transaction.atomic(using="rds"):

                if action == "create":
                    MonthlyPeriod.objects.using("rds").create(
                        kibetu=kibetu,
                        year=year or None,
                        month=month or None,
                        payment_date=parsed_payment_date,
                    )
                    record_change_audit(
                        request,
                        screen_name="期別(月)",
                        action_type="create",
                        target_table="monthly_period",
                        target_pk=kibetu,
                        summary=f"{kibetu} を追加",
                        before_values=None,
                        after_values={
                            "kibetu": kibetu,
                            "year": year or None,
                            "month": month or None,
                            "payment_date": parsed_payment_date,
                        },
                    )
                    messages.success(request, f"{kibetu} を追加しました。")

                elif action == "update":
                    obj = MonthlyPeriod.objects.using("rds").get(kibetu=kibetu)
                    before_values = {
                        "kibetu": obj.kibetu,
                        "year": obj.year,
                        "month": obj.month,
                        "payment_date": obj.payment_date,
                    }
                    obj.year = year or None
                    obj.month = month or None
                    obj.payment_date = parsed_payment_date
                    obj.save(using="rds")
                    record_change_audit(
                        request,
                        screen_name="期別(月)",
                        action_type="update",
                        target_table="monthly_period",
                        target_pk=kibetu,
                        summary=f"{kibetu} を変更",
                        before_values=before_values,
                        after_values={
                            "kibetu": kibetu,
                            "year": year or None,
                            "month": month or None,
                            "payment_date": parsed_payment_date,
                        },
                    )
                    messages.success(request, f"{kibetu} を変更しました。")

                elif action == "delete":
                    before_obj = MonthlyPeriod.objects.using("rds").filter(kibetu=kibetu).first()
                    before_values = None
                    if before_obj:
                        before_values = {
                            "kibetu": before_obj.kibetu,
                            "year": before_obj.year,
                            "month": before_obj.month,
                            "payment_date": before_obj.payment_date,
                        }
                    deleted_count, _ = MonthlyPeriod.objects.using("rds").filter(kibetu=kibetu).delete()
                    if deleted_count:
                        record_change_audit(
                            request,
                            screen_name="期別(月)",
                            action_type="delete",
                            target_table="monthly_period",
                            target_pk=kibetu,
                            summary=f"{kibetu} を削除",
                            before_values=before_values,
                            after_values=None,
                        )
                    messages.success(request, f"{kibetu} を削除しました。")

                else:
                    messages.error(request, "不正な操作です。")

        except MonthlyPeriod.DoesNotExist:
            messages.error(request, f"{kibetu} は存在しません。")
            return render_create_form_on_error()
        except IntegrityError:
            messages.error(request, f"{kibetu} はすでに存在します。")
            return render_create_form_on_error()
        except Exception as e:
            messages.error(request, f"エラーが発生しました: {e}")
            return render_create_form_on_error()

        return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu_month"))

    def _bulk_update(self, request):
        kibetus = request.POST.getlist("bulk_kibetu")
        years = request.POST.getlist("bulk_year")
        months = request.POST.getlist("bulk_month")
        payment_dates = request.POST.getlist("bulk_payment_date")
        row_lengths = {
            len(kibetus),
            len(years),
            len(months),
            len(payment_dates),
        }

        if not kibetus or len(row_lengths) != 1:
            messages.error(request, "一括変更するデータがありません。")
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu_month"))

        updated_count = 0
        errors = []

        try:
            with transaction.atomic(using="rds"):
                for idx, kibetu in enumerate(kibetus):
                    kibetu = (kibetu or "").strip()
                    if not kibetu:
                        continue

                    try:
                        parsed_payment_date = parse_input_date(payment_dates[idx])
                    except ValueError as e:
                        errors.append(f"{kibetu}: {e}")
                        continue

                    try:
                        obj = MonthlyPeriod.objects.using("rds").get(kibetu=kibetu)
                    except MonthlyPeriod.DoesNotExist:
                        errors.append(f"{kibetu}: 存在しません。")
                        continue

                    year = (years[idx] or "").strip()
                    month = (months[idx] or "").strip()
                    obj.year = year or None
                    obj.month = month or None
                    obj.payment_date = parsed_payment_date
                    obj.save(using="rds")
                    updated_count += 1

        except Exception as e:
            messages.error(request, f"エラーが発生しました: {e}")
            return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu_month"))

        if errors:
            messages.error(request, " / ".join(errors))
        if updated_count:
            record_change_audit(
                request,
                screen_name="期別(月)",
                action_type="bulk_update",
                target_table="monthly_period",
                target_pk=None,
                summary=f"期別(月)を {updated_count}件 一括変更",
                before_values=None,
                after_values={"count": updated_count},
            )
            messages.success(request, f"{updated_count} 件を変更しました。")
        elif not errors:
            messages.info(request, "変更対象がありませんでした。")

        return redirect(_build_kibetu_filter_redirect(request, "connect:kibetu_month"))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["selected_kibetu"] = (self.request.GET.get("kibetu") or "").strip()
        ctx["q_kibetu"] = (self.request.GET.get("q_kibetu") or "").strip()
        ctx.setdefault("create_form", {})

        ctx["kibetu_choices"] = list(
            MonthlyPeriod.objects.using("rds")
            .order_by("-year", "-month")
            .values_list("kibetu", flat=True)
            .distinct()
        )

        return ctx


class RepurchaseListView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "repurchase_list.html"

    DEFAULT_PER_PAGE = 100
    MAX_PER_PAGE = 500
    BV_ACTIVED_FLG_LABELS = {
        0: "未反映",
        1: "反映済",
        3: "反映無効",
    }
    SORT_COLUMNS = {
        "register_year": "register_year",
        "register_month": "register_month",
        "order_year": "order_year",
        "order_month": "order_month",
        "order_code": "order_code",
        "order_type": "order_type",
        "jwoa_code": "jwoa_code",
        "send_bv_name": "send_bv_name",
        "bonus_payment_date": "bonus_payment_date",
        "bv": "bv",
        "bv_actived_flg": "bv_actived_flg",
        "deposit_at": "deposit_at",
        "order_at": "order_at",
        "created_at": "created_at",
    }

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="bonus_payment_date",
            default_direction="desc",
        )

    def _build_where(
        self,
        year=None,
        month=None,
        q_code: str = "",
        q_name: str = "",
        q_order_code: str = "",
        q_order_types=None,
        q_bonus_date_from: str = "",
        q_bonus_date_to: str = "",
    ):
        where = ["1=1"]
        params = []

        if q_order_types is None:
            q_order_types = []

        if year is not None and month is not None:
            where.append("register_year = %s")
            where.append("register_month = %s")
            params.extend([year, month])

        if q_code:
            where.append("jwoa_code LIKE %s")
            params.append(f"%{q_code}%")

        if q_name:
            where.append("send_bv_name LIKE %s")
            params.append(f"%{q_name}%")

        if q_order_code:
            where.append("order_code LIKE %s")
            params.append(f"%{q_order_code}%")

        if q_order_types:
            placeholders = ", ".join(["%s"] * len(q_order_types))
            where.append(f"order_type IN ({placeholders})")
            params.extend(q_order_types)

        if q_bonus_date_from:
            where.append("bonus_payment_date >= %s")
            params.append(q_bonus_date_from)

        if q_bonus_date_to:
            where.append("bonus_payment_date < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_bonus_date_to)

        return "WHERE " + " AND ".join(where), params

    def _get_registered_months(self):
        sql = """
            SELECT DISTINCT CONCAT(register_year, '-', LPAD(register_month, 2, '0')) AS ym
            FROM bonus_db.purchase_info_list
            ORDER BY ym DESC
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            return [row[0] for row in cursor.fetchall()]

    def _fetch_last_created_at(self):
        sql = "SELECT MAX(created_at) FROM bonus_db.purchase_info_list"

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            row = cursor.fetchone()

        last_created_at = row[0] if row else None
        return jst_datetime(last_created_at)

    def _fetch_rows(
        self,
        year=None,
        month=None,
        q_code: str = "",
        q_name: str = "",
        q_order_code: str = "",
        q_order_types=None,
        q_bonus_date_from: str = "",
        q_bonus_date_to: str = "",
        order_sql: str = "",
        limit: int = 100,
        offset: int = 0,
    ):
        where_sql, params = self._build_where(
            year=year,
            month=month,
            q_code=q_code,
            q_name=q_name,
            q_order_code=q_order_code,
            q_order_types=q_order_types,
            q_bonus_date_from=q_bonus_date_from,
            q_bonus_date_to=q_bonus_date_to,
        )

        # purchase_info_list.order_code は utf8mb4、orders.order_code は utf8mb3。
        # CONVERT を挟まないと照合順序の変換で orders 側のインデックスが使えず、
        # 1ページ表示に10秒かかる。
        sql = f"""
            SELECT
                id,
                order_code,
                order_type,
                jwoa_code,
                send_bv_name,
                total_bv,
                bv,
                deposit_at,
                order_at,
                bonus_payment_date,
                created_at,
                register_year,
                register_month,
                order_year,
                order_month,
                (
                    SELECT o.bv_actived_flg
                    FROM nexus_production.orders AS o
                    WHERE o.order_code = CONVERT(
                        bonus_db.purchase_info_list.order_code USING utf8mb3
                    )
                    LIMIT 1
                ) AS bv_actived_flg
            FROM bonus_db.purchase_info_list
            {where_sql}
            ORDER BY {order_sql or "bonus_payment_date DESC"}, id DESC
            LIMIT %s OFFSET %s
        """

        params.extend([limit, offset])

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    def _count_rows(
        self,
        year=None,
        month=None,
        q_code: str = "",
        q_name: str = "",
        q_order_code: str = "",
        q_order_types=None,
        q_bonus_date_from: str = "",
        q_bonus_date_to: str = "",
    ) -> int:
        where_sql, params = self._build_where(
            year=year,
            month=month,
            q_code=q_code,
            q_name=q_name,
            q_order_code=q_order_code,
            q_order_types=q_order_types,
            q_bonus_date_from=q_bonus_date_from,
            q_bonus_date_to=q_bonus_date_to,
        )

        sql = f"""
            SELECT COUNT(*) AS cnt
            FROM bonus_db.purchase_info_list
            {where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
            return int(row[0]) if row else 0

    def _get_month_choices(self):
        sql = """
            SELECT DISTINCT
                register_year,
                register_month
            FROM bonus_db.purchase_info_list
            ORDER BY register_year DESC, register_month DESC
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            return [
                {
                    "value": format_target_month(row[0], row[1]),
                    "year": row[0],
                    "month": row[1],
                }
                for row in cursor.fetchall()
            ]

    def _get_period_choices(self):
        return list(
            PeriodMaster.objects.using("rds")
            .exclude(st_date__isnull=True)
            .exclude(end_date__isnull=True)
            .order_by("-st_date", "-kibetu")
        )

    def _resolve_kibetu_period(self, q_kibetu):
        if not q_kibetu:
            return None

        return (
            PeriodMaster.objects.using("rds")
            .filter(kibetu=q_kibetu)
            .exclude(st_date__isnull=True)
            .exclude(end_date__isnull=True)
            .first()
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_month = (self.request.GET.get("target_month") or "").strip()
        q_kibetu = (self.request.GET.get("q_kibetu") or "").strip()

        q_code = (self.request.GET.get("q_code") or "").strip()
        q_name = (self.request.GET.get("q_name") or "").strip()
        q_order_code = (self.request.GET.get("q_order_code") or "").strip()

        q_order_types = self.request.GET.getlist("q_order_type")
        q_order_types = [x for x in q_order_types if x]

        q_bonus_date_from = (self.request.GET.get("q_bonus_date_from") or "").strip()
        q_bonus_date_to = (self.request.GET.get("q_bonus_date_to") or "").strip()

        try:
            per_page = int(self.request.GET.get("per_page") or str(self.DEFAULT_PER_PAGE))
        except ValueError:
            per_page = self.DEFAULT_PER_PAGE

        per_page = max(1, min(per_page, self.MAX_PER_PAGE))

        try:
            page = int(self.request.GET.get("page") or "1")
        except ValueError:
            page = 1

        page = max(1, page)

        ctx["month_choices"] = self._get_month_choices()
        ctx["period_choices"] = self._get_period_choices()
        ctx["last_created_at"] = self._fetch_last_created_at()

        ctx["selected_month"] = selected_month
        ctx["q_kibetu"] = q_kibetu
        ctx["selected_kibetu"] = q_kibetu
        ctx["history_rows"] = get_week_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:repurchase_list"
        ctx["q_code"] = q_code
        ctx["q_name"] = q_name
        ctx["q_order_code"] = q_order_code
        ctx["q_order_types"] = q_order_types
        ctx["per_page"] = per_page

        history_select_base_params = {}
        if selected_month:
            history_select_base_params["target_month"] = selected_month
        if q_code:
            history_select_base_params["q_code"] = q_code
        if q_name:
            history_select_base_params["q_name"] = q_name
        if q_order_code:
            history_select_base_params["q_order_code"] = q_order_code
        if q_bonus_date_from:
            history_select_base_params["q_bonus_date_from"] = q_bonus_date_from
        if q_bonus_date_to:
            history_select_base_params["q_bonus_date_to"] = q_bonus_date_to
        if per_page != self.DEFAULT_PER_PAGE:
            history_select_base_params["per_page"] = per_page
        ctx["history_select_base_qs"] = urlencode(history_select_base_params)
        for order_type in q_order_types:
            if ctx["history_select_base_qs"]:
                ctx["history_select_base_qs"] += "&"
            ctx["history_select_base_qs"] += urlencode({"q_order_type": order_type})

        year = None
        month = None
        ctx["selected_period"] = None

        if selected_month:
            try:
                year, month = map(int, selected_month.split("-"))
                ctx["selected_period"] = {
                    "year": year,
                    "month": month,
                }
            except ValueError:
                year = None
                month = None

        selected_kibetu_period = self._resolve_kibetu_period(q_kibetu)
        ctx["selected_kibetu_period"] = selected_kibetu_period
        if q_kibetu and not selected_kibetu_period:
            ctx["kibetu_period_error"] = "選択された期別の期間が見つかりません。"

        if selected_kibetu_period:
            effective_bonus_date_from = selected_kibetu_period.st_date.strftime("%Y-%m-%d")
            effective_bonus_date_to = selected_kibetu_period.end_date.strftime("%Y-%m-%d")
        else:
            effective_bonus_date_from = q_bonus_date_from
            effective_bonus_date_to = q_bonus_date_to

        ctx["input_bonus_date_from"] = q_bonus_date_from
        ctx["input_bonus_date_to"] = q_bonus_date_to
        ctx["q_bonus_date_from"] = effective_bonus_date_from
        ctx["q_bonus_date_to"] = effective_bonus_date_to

        total_count = self._count_rows(
            year=year,
            month=month,
            q_code=q_code,
            q_name=q_name,
            q_order_code=q_order_code,
            q_order_types=q_order_types,
            q_bonus_date_from=effective_bonus_date_from,
            q_bonus_date_to=effective_bonus_date_to,
        )

        total_pages = max(1, math.ceil(total_count / per_page))

        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page

        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        rows = self._fetch_rows(
            year=year,
            month=month,
            q_code=q_code,
            q_name=q_name,
            q_order_code=q_order_code,
            q_order_types=q_order_types,
            q_bonus_date_from=effective_bonus_date_from,
            q_bonus_date_to=effective_bonus_date_to,
            order_sql=sort_ctx["order_sql"],
            limit=per_page,
            offset=offset,
        )

        base_params = {}

        if selected_month:
            base_params["target_month"] = selected_month
        if q_kibetu:
            base_params["q_kibetu"] = q_kibetu
        if q_code:
            base_params["q_code"] = q_code
        if q_name:
            base_params["q_name"] = q_name
        if q_order_code:
            base_params["q_order_code"] = q_order_code
        if q_bonus_date_from and not selected_kibetu_period:
            base_params["q_bonus_date_from"] = q_bonus_date_from
        if q_bonus_date_to and not selected_kibetu_period:
            base_params["q_bonus_date_to"] = q_bonus_date_to
        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        add_sort_params(base_params, sort_ctx)

        ctx = self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )

        base_qs = ctx["base_qs"]
        for order_type in q_order_types:
            if base_qs:
                base_qs += "&"
            base_qs += urlencode({"q_order_type": order_type})
        ctx["base_qs"] = base_qs

        return ctx



class SettingsView(generic.ListView):
    template_name = "settings.html"
    context_object_name = "rows"
    model = Settings

    SCREEN_NAME = "設定"
    TARGET_TABLE = "settings"
    PERMISSION_BY_ACTION = {
        "create": "can_create",
        "update": "can_update",
        "delete": "can_delete",
    }

    def get_queryset(self):
        qs = Settings.objects.using("rds").all()

        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(value__icontains=q)
                | Q(comment__icontains=q)
            )

        # 並び順はお好みで（title_id順など）
        return qs.order_by("id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # 件数表示用（テンプレで rows|length を使わない）
        ctx["total_count"] = ctx["rows"].count()
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        return ctx

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip()

        if action not in self.PERMISSION_BY_ACTION:
            messages.error(request, "不正な操作です。")
            return redirect(self._redirect_url(request))

        user_access = get_user_access(request.user)
        if not user_access.can_menu("settings") or not user_access.has(
            self.PERMISSION_BY_ACTION[action]
        ):
            return HttpResponse("権限がありません。", status=403)

        try:
            with transaction.atomic(using="rds"):
                if action == "create":
                    self._create(request)
                elif action == "update":
                    self._update(request)
                else:
                    self._delete(request)

        except Settings.DoesNotExist:
            messages.error(request, "対象の設定が存在しません。")
        except IntegrityError:
            messages.error(request, "同じ設定キーがすでに存在します。")
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"エラーが発生しました: {e}")

        return redirect(self._redirect_url(request))

    def _create(self, request):
        name = (request.POST.get("name") or "").strip()
        value = (request.POST.get("value") or "").strip()
        comment = (request.POST.get("comment") or "").strip()

        if not name:
            raise ValueError("設定キーを入力してください。")

        if Settings.objects.using("rds").filter(name=name).exists():
            raise ValueError(f"設定キー {name} はすでに存在します。")

        obj = Settings.objects.using("rds").create(
            name=name,
            value=value,
            comment=comment or None,
        )
        record_change_audit(
            request,
            screen_name=self.SCREEN_NAME,
            action_type="create",
            target_table=self.TARGET_TABLE,
            target_pk=obj.pk,
            summary=f"{name} を追加",
            before_values=None,
            after_values={"name": name, "value": value, "comment": comment or None},
        )
        messages.success(request, f"{name} を追加しました。")

    def _update(self, request):
        obj = self._get_target(request, "更新対象が指定されていません。")

        value = (request.POST.get("value") or "").strip()
        comment = (request.POST.get("comment") or "").strip()

        if obj.value == value and (obj.comment or "") == comment:
            messages.info(request, f"{obj.name} は変更がありませんでした。")
            return

        before_values = {
            "name": obj.name,
            "value": obj.value,
            "comment": obj.comment,
        }
        obj.value = value
        obj.comment = comment or None
        obj.save(using="rds", update_fields=["value", "comment"])

        record_change_audit(
            request,
            screen_name=self.SCREEN_NAME,
            action_type="update",
            target_table=self.TARGET_TABLE,
            target_pk=obj.pk,
            summary=f"{obj.name} を変更",
            before_values=before_values,
            after_values={
                "name": obj.name,
                "value": obj.value,
                "comment": obj.comment,
            },
        )
        messages.success(request, f"{obj.name} を変更しました。")

    def _delete(self, request):
        obj = self._get_target(request, "削除対象が指定されていません。")

        before_values = {
            "name": obj.name,
            "value": obj.value,
            "comment": obj.comment,
        }
        target_pk = obj.pk
        obj.delete(using="rds")

        record_change_audit(
            request,
            screen_name=self.SCREEN_NAME,
            action_type="delete",
            target_table=self.TARGET_TABLE,
            target_pk=target_pk,
            summary=f"{before_values['name']} を削除",
            before_values=before_values,
            after_values=None,
        )
        messages.success(request, f"{before_values['name']} を削除しました。")

    def _get_target(self, request, missing_message):
        row_id = (request.POST.get("row_id") or "").strip()
        if not row_id:
            raise ValueError(missing_message)
        return Settings.objects.using("rds").get(pk=row_id)

    def _redirect_url(self, request):
        q = (request.POST.get("filter_q") or request.GET.get("q") or "").strip()
        url = reverse("connect:settings")
        if q:
            url += "?" + urlencode({"q": q})
        return url


class UserTargetRankView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "user_target_rank.html"

    DISPLAY_COLUMNS = [
        "id",
        "jmoa_code",
        "introducer_code",
        "placement_code",
        "group_code",
        "send_bv_name",
        "status_code",
        "rank",
        "salon_administrator",
        "salon_name",
        "interim_at",
        "activated_at",
        "created_at",
        "target_rank",
        "max_up_at",
        "new_rank",
    ]

    DEFAULT_PER_PAGE = 100
    MAX_PER_PAGE = 500

    # ----------------------------
    # UI: 月リスト
    # ----------------------------
    def get_month_list(self):
        today = date.today().replace(day=1)
        months = []
        for i in range(0, 13):
            d = today - relativedelta(months=i)
            months.append({
                "value": f"{d.year}-{d.month:02d}",
                "year": d.year,
                "month": d.month,
            })
        return months

    def _month_end_exclusive(self, year: int, month: int):
        base = datetime(year, month, 1, 0, 0, 0)
        return base + relativedelta(months=1)

    # ----------------------------
    # WHERE句
    # ----------------------------
    def _build_where(self, q_code: str = "", q_name: str = "", q_new_rank: str = ""):
        where = ["1=1"]
        params = []

        if q_code:
            where.append("t.jmoa_code LIKE %s")
            params.append(f"%{q_code}%")

        if q_name:
            where.append("t.send_bv_name LIKE %s")
            params.append(f"%{q_name}%")

        if q_new_rank:
            where.append("""
CASE
  WHEN t.status_code <> 1 THEN 9
  WHEN x.fluctuation_name REGEXP '^[0-9]+$' THEN CAST(x.fluctuation_name AS UNSIGNED)
  ELSE t.`rank`
END = %s
""")
            params.append(q_new_rank)

        where_sql = "WHERE " + " AND ".join(where)
        return where_sql, params

    # ----------------------------
    # 総件数
    # ----------------------------
    def _fetch_total_count(self, cutoff_dt: datetime, q_code: str = "", q_name: str = "", q_new_rank: str = "") -> int:
        where_sql, params = self._build_where(q_code=q_code, q_name=q_name, q_new_rank=q_new_rank)

        sql = f"""
SELECT COUNT(*)
FROM nexus_production.users t
LEFT JOIN (
  SELECT user_id, fluctuation_name, created_at
  FROM (
    SELECT
      user_id,
      fluctuation_name,
      created_at,
      id,
      ROW_NUMBER() OVER (
        PARTITION BY user_id
        ORDER BY created_at DESC, id DESC
      ) AS rn
    FROM bonus_db.users_rank_up_history
    WHERE created_at <= %s
  ) r
  WHERE rn = 1
) x
  ON t.jmoa_code = x.user_id
{where_sql}
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [cutoff_dt] + params)
            return int(cursor.fetchone()[0])

    # ----------------------------
    # 表示用データ
    # ----------------------------
    def _fetch_users(
        self,
        cutoff_dt: datetime,
        q_code: str = "",
        q_name: str = "",
        q_new_rank: str = "",
        limit: int = 100,
        offset: int = 0,
    ):
        where_sql, params = self._build_where(q_code=q_code, q_name=q_name, q_new_rank=q_new_rank)

        sql = f"""
SELECT
  t.id,
  t.jmoa_code,
  t.introducer_code,
  t.placement_code,
  t.group_code,
  t.send_bv_name,
  t.status_code,
  t.`rank`,
  t.salon_administrator,
  t.salon_name,
  t.interim_at,
  t.activated_at,
  t.created_at,

  CASE
    WHEN x.fluctuation_name REGEXP '^[0-9]+$' THEN CAST(x.fluctuation_name AS UNSIGNED)
    ELSE NULL
  END AS target_rank,

  x.created_at AS max_up_at,

  CASE
    WHEN t.status_code <> 1 THEN 9
    WHEN x.fluctuation_name REGEXP '^[0-9]+$' THEN CAST(x.fluctuation_name AS UNSIGNED)
    ELSE t.`rank`
  END AS new_rank

FROM nexus_production.users t
LEFT JOIN (
  SELECT user_id, fluctuation_name, created_at
  FROM (
    SELECT
      user_id,
      fluctuation_name,
      created_at,
      id,
      ROW_NUMBER() OVER (
        PARTITION BY user_id
        ORDER BY created_at DESC, id DESC
      ) AS rn
    FROM bonus_db.users_rank_up_history
    WHERE created_at <= %s
  ) r
  WHERE rn = 1
) x
  ON t.jmoa_code = x.user_id
{where_sql}
ORDER BY t.jmoa_code
LIMIT %s OFFSET %s
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [cutoff_dt] + params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    # ----------------------------
    # settings 取得
    # ----------------------------
    def _get_select_month_setting(self):
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT value
                FROM bonus_db.settings
                WHERE name = 'user_add_rank'
                LIMIT 1
            """)
            row = cursor.fetchone()
        return row[0] if row else ""

    # ----------------------------
    # GET
    # ----------------------------
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_prev_month = (self.request.GET.get("prev_month") or "").strip()
        q_code = (self.request.GET.get("q_code") or "").strip()
        q_name = (self.request.GET.get("q_name") or "").strip()
        q_new_rank = (self.request.GET.get("q_new_rank") or "").strip()

        try:
            per_page = int(self.request.GET.get("per_page") or str(self.DEFAULT_PER_PAGE))
        except ValueError:
            per_page = self.DEFAULT_PER_PAGE
        per_page = max(1, min(per_page, self.MAX_PER_PAGE))

        try:
            page = int(self.request.GET.get("page") or "1")
        except ValueError:
            page = 1
        page = max(1, page)

        ctx["month_list"] = self.get_month_list()
        ctx["selected_prev_month"] = selected_prev_month
        ctx["selected_period"] = None
        ctx["columns"] = self.DISPLAY_COLUMNS
        ctx["rows"] = []
        ctx["total_count"] = 0
        select_month = self._get_select_month_setting()
        ctx["select_month"] = select_month
        ctx["select_month_value"] = ""
        if select_month and len(select_month) == 6 and select_month.isdigit():
            ctx["select_month_value"] = f"{select_month[0:4]}-{select_month[4:6]}"

        ctx["q_code"] = q_code
        ctx["q_name"] = q_name
        ctx["q_new_rank"] = q_new_rank
        ctx["per_page"] = per_page
        ctx["page"] = 1
        ctx["total_pages"] = 1
        ctx["has_prev"] = False
        ctx["has_next"] = False
        ctx["prev_page"] = 1
        ctx["next_page"] = 1
        ctx["base_qs"] = ""

        if not selected_prev_month:
            return ctx

        try:
            y, m = map(int, selected_prev_month.split("-"))
        except ValueError:
            return ctx

        ctx["selected_period"] = {"year": y, "month": m}
        cutoff_dt = self._month_end_exclusive(y, m)

        total_count = self._fetch_total_count(
            cutoff_dt=cutoff_dt,
            q_code=q_code,
            q_name=q_name,
            q_new_rank=q_new_rank,
        )

        total_pages = max(1, math.ceil(total_count / per_page))
        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page

        rows = self._fetch_users(
            cutoff_dt=cutoff_dt,
            q_code=q_code,
            q_name=q_name,
            q_new_rank=q_new_rank,
            limit=per_page,
            offset=offset,
        )

        base_params = {
            "prev_month": selected_prev_month,
        }
        if q_code:
            base_params["q_code"] = q_code
        if q_name:
            base_params["q_name"] = q_name
        if q_new_rank:
            base_params["q_new_rank"] = q_new_rank
        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        ctx["q_code"] = q_code
        ctx["q_name"] = q_name
        ctx["q_new_rank"] = q_new_rank

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )

    # ----------------------------
    # POST: 登録（全件）
    # ----------------------------
    def post(self, request, *args, **kwargs):
        selected_prev_month = (request.POST.get("prev_month") or "").strip()
        if not selected_prev_month:
            messages.error(request, "対象年月が未選択です。")
            return redirect("connect:user_target_rank")

        year, month = map(int, selected_prev_month.split("-"))
        target_rank = f"{year}{month:02d}"

        register_users_target_rank(year, month)

        messages.success(request, f"{year}年{month}月（{target_rank}）で全件登録しました。")
        return redirect(f"{redirect('connect:user_target_rank').url}?prev_month={selected_prev_month}")


class TitleUserView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "title_user.html"

    DEFAULT_PER_PAGE = 200
    MAX_PER_PAGE = 500
    EXPORT_FETCH_SIZE = 5000
    SORT_COLUMNS = {
        "jmoa_code": "ut.jmoa_code",
        "jwoa_name": "u.send_bv_name",
        "title_name": "ut.title_id",
        "update_date": "ut.update_date",
    }

    # user_titles.jmoa_code は utf8mb4_bin、users.jmoa_code は utf8mb3_bin。
    # そのまま結合すると照合順序の変換で users 側のインデックスが使えず、
    # 3万件×7万件の総当たりになって画面が返ってこない。
    # COLLATE まで揃えるとバイナリ比較のまま eq_ref で引ける。
    MEMBER_JOIN_SQL = """
LEFT JOIN nexus_production.users u
  ON CONVERT(ut.jmoa_code USING utf8mb3) COLLATE utf8mb3_bin = u.jmoa_code
"""

    def _get_sort_context(self):
        title_id = (self.request.GET.get("title_id") or "").strip()
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="jmoa_code" if title_id else "title_name",
            default_direction="asc",
        )

    def _build_where(self, title_id: str, q_jpid: str):
        where = []
        params = []

        if title_id:
            where.append("ut.title_id = %s")
            params.append(title_id)

        if q_jpid:
            where.append("ut.jmoa_code LIKE %s")
            params.append(f"{q_jpid}%")

        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        return where_sql, params

    def _fetch_total_count(self, title_id: str, q_jpid: str) -> int:
        where_sql, params = self._build_where(title_id, q_jpid)

        sql = f"""
SELECT COUNT(*)
FROM bonus_db.user_titles ut
{where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            return int(cursor.fetchone()[0])

    def _fetch_rows(
        self,
        title_id: str,
        q_jpid: str,
        limit: int,
        offset: int = 0,
        order_sql: str = "",
    ):
        where_sql, params = self._build_where(title_id, q_jpid)

        sql = f"""
SELECT
  ut.jmoa_code AS jmoa_code,
  u.send_bv_name AS jwoa_name,
  ut.title_id AS title_id,
  COALESCE(tm.title_name, 'タイトルなし') AS title_name,
  ut.update_date AS update_date
FROM bonus_db.user_titles ut
{self.MEMBER_JOIN_SQL}
LEFT JOIN bonus_db.title_master tm
  ON ut.title_id = tm.title_id
{where_sql}
ORDER BY {order_sql or "ut.title_id ASC"}, ut.jmoa_code ASC
LIMIT %s OFFSET %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    def _fetch_title_choices(self):
        sql = """
SELECT
  tm.title_id,
  tm.title_name
FROM bonus_db.title_master tm
ORDER BY tm.title_id
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            choices = [{"title_id": r[0], "title_name": r[1]} for r in cursor.fetchall()]
        if not any(str(choice["title_id"]) == "0" for choice in choices):
            choices.insert(0, {"title_id": 0, "title_name": "タイトルなし"})
        return choices

    def _title_exists(self, title_id: int) -> bool:
        if title_id == 0:
            return True

        with connections["rds"].cursor() as cursor:
            cursor.execute(
                """
SELECT 1
FROM bonus_db.title_master
WHERE title_id = %s
LIMIT 1
                """,
                [title_id],
            )
            return cursor.fetchone() is not None

    def _update_user_title(self, jmoa_code: str, title_id: int) -> int:
        with connections["rds"].cursor() as cursor:
            logger.info(
                "最高ピンタイトル一覧からタイトルIDを更新します。jmoa_code=%s title_id=%s",
                jmoa_code,
                title_id,
            )
            cursor.execute(
                """
UPDATE bonus_db.user_titles
SET
  title_id = %s,
  update_date = CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
WHERE jmoa_code = %s
                """,
                [title_id, jmoa_code],
            )
            return cursor.rowcount

    def _can_update_title_user(self, user) -> bool:
        profile = getattr(user, "access_profile", None)
        if profile is not None:
            return bool(profile.can_update)
        return get_user_access(user).can_update

    def post(self, request, *args, **kwargs):
        user_access = get_user_access(request.user)
        if not user_access.can_menu("title_user") or not self._can_update_title_user(request.user):
            return HttpResponse("権限がありません。", status=403)

        next_query = (request.POST.get("next_query") or "").strip()
        list_url = redirect("connect:title_user").url
        next_url = f"{list_url}?{next_query}" if next_query else list_url

        action = (request.POST.get("action") or "").strip()
        if action != "update":
            messages.error(request, "不正な操作です。")
            return redirect(next_url)

        jmoa_code = (request.POST.get("jmoa_code") or "").strip()
        title_id_text = (request.POST.get("title_id") or "").strip()
        if not jmoa_code:
            messages.error(request, "更新対象の会員IDが不正です。")
            return redirect(next_url)

        try:
            title_id = int(title_id_text)
        except (TypeError, ValueError):
            messages.error(request, "タイトルIDが不正です。")
            return redirect(next_url)

        if title_id < 0 or not self._title_exists(title_id):
            messages.error(request, "存在しないタイトルIDです。")
            return redirect(next_url)

        before_row = fetch_one_dict(
            "rds",
            """
SELECT jmoa_code, title_id, update_date
FROM bonus_db.user_titles
WHERE jmoa_code = %s
            """,
            [jmoa_code],
        )

        try:
            with transaction.atomic(using="rds"):
                updated_count = self._update_user_title(jmoa_code, title_id)
        except Exception as e:
            logger.exception("最高ピンタイトル一覧のタイトルID更新エラー")
            messages.error(request, f"タイトルID更新中にエラーが発生しました: {e}")
            return redirect(next_url)

        if updated_count:
            after_row = fetch_one_dict(
                "rds",
                """
SELECT jmoa_code, title_id, update_date
FROM bonus_db.user_titles
WHERE jmoa_code = %s
                """,
                [jmoa_code],
            )
            record_change_audit(
                request,
                screen_name="最高ピンタイトル一覧",
                action_type="update",
                target_table="user_titles",
                target_pk=jmoa_code,
                summary=f"{jmoa_code} のタイトルIDを {title_id} に変更",
                before_values=before_row,
                after_values=after_row,
            )
            messages.success(request, f"{jmoa_code} のタイトルIDを {title_id} に変更しました。")
        else:
            messages.warning(request, "更新対象データがありません。")

        return redirect(next_url)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        title_id = (self.request.GET.get("title_id") or "").strip()
        q_jpid = (self.request.GET.get("q_jpid") or "").strip()

        try:
            per_page = int(self.request.GET.get("per_page") or str(self.DEFAULT_PER_PAGE))
        except ValueError:
            per_page = self.DEFAULT_PER_PAGE
        per_page = max(1, min(per_page, self.MAX_PER_PAGE))

        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        total_count = self._fetch_total_count(title_id, q_jpid)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            title_id=title_id,
            q_jpid=q_jpid,
            limit=per_page,
            offset=offset,
            order_sql=sort_ctx["order_sql"],
        )

        base_params = {}
        if title_id:
            base_params["title_id"] = title_id
        if q_jpid:
            base_params["q_jpid"] = q_jpid
        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        add_sort_params(base_params, sort_ctx)

        ctx["title_choices"] = self._fetch_title_choices()
        ctx["selected_title_id"] = title_id
        ctx["q_jpid"] = q_jpid
        ctx["title_user_can_update"] = self._can_update_title_user(self.request.user)

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )


class TitleUserExportView(TitleUserView):
    def get(self, request, *args, **kwargs):
        title_id = (request.GET.get("title_id") or "").strip()
        q_jpid = (request.GET.get("q_jpid") or "").strip()
        where_sql, params = self._build_where(title_id, q_jpid)
        order_sql = self._get_sort_context()["order_sql"]

        sql = f"""
SELECT
  ut.jmoa_code AS jmoa_code,
  u.send_bv_name AS jwoa_name,
  ut.title_id AS title_id,
  COALESCE(tm.title_name, 'タイトルなし') AS title_name,
  ut.update_date AS update_date
FROM bonus_db.user_titles ut
{self.MEMBER_JOIN_SQL}
LEFT JOIN bonus_db.title_master tm
  ON ut.title_id = tm.title_id
{where_sql}
ORDER BY {order_sql}, ut.jmoa_code ASC
        """

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("最高ピンタイトル一覧")
        ws.append(["jmoa_code", "jwoa_name", "title_id", "title_name", "update_date"])

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            while True:
                rows = cursor.fetchmany(self.EXPORT_FETCH_SIZE)
                if not rows:
                    break
                for row in rows:
                    ws.append(list(row))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="title_user.xlsx"'
        )
        wb.save(response)
        return response



class BasicBonusView(generic.ListView):
    template_name = "basic_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster

    def get_queryset(self):
        return PeriodMaster.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()
        save_basic_bv_line = request.POST.get("save_basic_bv_line", "1") == "1"

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "basic_bonus")

        if action != "register_basic_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:basic_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:basic_bonus")

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:basic_bonus")

        try:
            basic_bonus_rows = self._get_basic_bonus_rows(selected_kibetu, period)
            basic_bv_line_rows = (
                self._get_basic_bv_line_rows(selected_kibetu, period)
                if save_basic_bv_line
                else []
            )
            bv_line_comment = (
                "繰り越しBV保存あり" if save_basic_bv_line else "繰り越しBV保存なし"
            )

            if not basic_bonus_rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "basic_bonus",
                        selected_kibetu,
                        request.user.username,
                        f"0件登録（対象データなし、{bv_line_comment}）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/basic_bonus/?kibetu={selected_kibetu}")

            (
                delete_bonus_result_sql,
                delete_bv_line_sql,
                delete_params,
                insert_sql,
                insert_params,
                basic_bv_line_insert_sql,
                basic_bv_line_insert_params,
            ) = register_sql.get_basic_bonus_delete_insert_data(
                selected_kibetu,
                basic_bonus_rows,
                basic_bv_line_rows,
            )

            #登録
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    logger.info(
                        "ベーシックボーナス登録前削除SQLを実行します。kibetu=%s",
                        selected_kibetu,
                    )
                    cursor.execute(delete_bonus_result_sql, delete_params)
                    if save_basic_bv_line:
                        cursor.execute(delete_bv_line_sql, delete_params)

                    logger.info(
                        "ベーシックボーナス登録INSERT SQLを実行します。kibetu=%s",
                        selected_kibetu,
                    )
                    cursor.executemany(insert_sql, insert_params)

                    if save_basic_bv_line and basic_bv_line_insert_params:
                        cursor.executemany(
                            basic_bv_line_insert_sql, basic_bv_line_insert_params
                        )

                    # 履歴登録
                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "basic_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(basic_bonus_rows)}件登録（{bv_line_comment}）"
                        ]
                    )

            messages.success(
                request,
                f"{len(basic_bonus_rows)}件をベーシックボーナス結果に登録しました。"
                f"（{bv_line_comment}）"
            )

        except Exception as e:
            logger.exception("ベーシックボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/basic_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_week_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:basic_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            return ctx

        ctx["selected_period"] = period
        ctx["rows"] = self._get_basic_bonus_rows(selected_kibetu, period)
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "basic_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_basic_bonus_rows(self, selected_kibetu, period):
        st_date = period.st_date
        end_date = period.end_date

        kibetu_year = int(selected_kibetu[0:4])
        kibetu_month = int(selected_kibetu[5:7])

        start_dt = make_aware(datetime.combine(st_date, time.min))
        end_dt = make_aware(datetime.combine(end_date + timedelta(days=1), time.min))

        current_month_first = datetime(kibetu_year, kibetu_month, 1)
        prev_month_last = current_month_first - timedelta(days=1)

        prev_year = prev_month_last.year
        prev_month = prev_month_last.month

        be_start_dt = make_aware(datetime(prev_year, prev_month, 1, 0, 0, 0))
        be_end_dt = make_aware(datetime(kibetu_year, kibetu_month, 1, 0, 0, 0))


        params = [
            selected_kibetu,
            prev_year,
            prev_month,
            be_start_dt,
            be_end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(BASIC_BONUS_SQL, params)
#             logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows


    def _get_basic_bv_line_rows(self, selected_kibetu, period):
        st_date = period.st_date
        end_date = period.end_date

        kibetu_year = int(selected_kibetu[0:4])
        kibetu_month = int(selected_kibetu[5:7])

        start_dt = make_aware(datetime.combine(st_date, time.min))
        end_dt = make_aware(datetime.combine(end_date + timedelta(days=1), time.min))

        current_month_first = datetime(kibetu_year, kibetu_month, 1)
        prev_month_last = current_month_first - timedelta(days=1)

        prev_year = prev_month_last.year
        prev_month = prev_month_last.month

        be_start_dt = make_aware(datetime(prev_year, prev_month, 1, 0, 0, 0))
        be_end_dt = make_aware(datetime(kibetu_year, kibetu_month, 1, 0, 0, 0))


        params = [
            selected_kibetu,
            prev_year,
            prev_month,
            be_start_dt,
            be_end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(BASIC_BV_LINE_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows




class TitleRegistrationView(generic.TemplateView):
    template_name = "title_registration.html"

    def _get_month_choices(self):
        today = date.today().replace(day=1)
        return [
            {
                "value": (today - relativedelta(months=i)).strftime("%Y-%m"),
                "year": (today - relativedelta(months=i)).year,
                "month": (today - relativedelta(months=i)).month,
            }
            for i in range(12)
        ]

    def _resolve_kibetu(self, year, month):
        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(year=year, month=month)
            .order_by("-kibetu")
            .first()
        )
        return period.kibetu if period else ""

    def _resolve_period(self, kibetu="", target_month=""):
        if kibetu:
            return MonthlyPeriod.objects.using("rds").filter(kibetu=kibetu).first()

        if target_month:
            try:
                year, month = map(int, target_month.split("-"))
            except (ValueError, TypeError):
                return None

            return (
                MonthlyPeriod.objects.using("rds")
                .filter(year=year, month=month)
                .order_by("-kibetu")
                .first()
            )

        return None

    def _redirect_url(self, kibetu=""):
        if kibetu:
            return f"/title_registration/?kibetu={kibetu}"
        return "/title_registration/"

    def _fetch_rows(self, year, month, kibetu=None):
        kibetu = kibetu or self._resolve_kibetu(year, month)
        if not kibetu:
            return []

        sql = """
            SELECT
                mt.jwoa_code,
                mt.basic_line_bv AS basic_bv,
                mt.income_line_bv AS income_bv,
                mt.title_id,
                COALESCE(tm.title_name, 'タイトルなし') AS title_name,
                %s AS year,
                %s AS month
            FROM bonus_db.month_title mt
            JOIN bonus_db.user_titles u
              ON u.jmoa_code = mt.jwoa_code
            LEFT JOIN bonus_db.title_master tm
              ON mt.title_id = tm.title_id
            WHERE mt.kibetu = %s
              AND mt.title_id > 0
              AND mt.title_id >= IFNULL(u.title_id, 0)
            ORDER BY mt.title_id DESC, mt.jwoa_code
        """
        params = [year, month, kibetu]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    def _exists_data(self, year, month):
        sql = """
SELECT 1
FROM bonus_db.title_update_history
WHERE year = %s
  AND month = %s
LIMIT 1
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [year, month])
            logger.info(f"Executed SQL: {cursor._executed}")
            return cursor.fetchone() is not None

    def _insert_rows(self, year, month, kibetu=None):
        kibetu = kibetu or self._resolve_kibetu(year, month)
        if not kibetu:
            return

        sql = """
            INSERT INTO bonus_db.title_update_history (
                jwoa_code,
                basic_bv,
                income_bv,
                title_id,
                year,
                month,
                created_at
            )
            SELECT
                mt.jwoa_code,
                mt.basic_line_bv,
                mt.income_line_bv,
                mt.title_id,
                %s AS year,
                %s AS month,
                CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
            FROM bonus_db.month_title mt
            JOIN bonus_db.user_titles u
              ON u.jmoa_code = mt.jwoa_code
            WHERE mt.kibetu = %s
              AND mt.title_id > 0
              AND mt.title_id >= IFNULL(u.title_id, 0)
        """

        params = [year, month, kibetu]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")

    def _update_title(self, year, month, kibetu=None):
        kibetu = kibetu or self._resolve_kibetu(year, month)
        if not kibetu:
            return

        sql = """
            UPDATE bonus_db.user_titles u
            JOIN bonus_db.month_title mt
              ON u.jmoa_code = mt.jwoa_code
            SET
              u.title_id = mt.title_id,
              u.update_date = CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
            WHERE mt.kibetu = %s
              AND mt.title_id > 0
              AND IFNULL(u.title_id, 0) < mt.title_id
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [kibetu])
            logger.info(f"Executed SQL: {cursor._executed}")

    def _update_setting(self, year, month):
        value = f"{year}{month:02d}"

        sql = """
    UPDATE bonus_db.settings
    SET value = %s
    WHERE name = 'set_title'
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [value])
            logger.info(f"Executed SQL: {cursor._executed}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        month_choices = self._get_month_choices()
        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        selected = self.request.GET.get("target_month")
        selected_period = self._resolve_period(selected_kibetu, selected)

        ctx["month_choices"] = month_choices
        ctx["selected_kibetu"] = selected_period.kibetu if selected_period else selected_kibetu
        ctx["selected_period"] = selected_period
        ctx["selected_month"] = (
            f"{selected_period.year}-{selected_period.month:02d}"
            if selected_period
            else selected
        )
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:title_registration"
        ctx["rows"] = []

        if selected_period:
            ctx["rows"] = self._fetch_rows(
                selected_period.year,
                selected_period.month,
                selected_period.kibetu,
            )
            if not ctx["rows"]:
                insert_empty_bonus_history_on_display(
                    self.request,
                    "title_registration",
                    selected_period.kibetu,
                )

        return ctx

    def post(self, request, *args, **kwargs):
        selected_kibetu = (request.POST.get("kibetu") or "").strip()
        selected = request.POST.get("target_month")
        period = self._resolve_period(selected_kibetu, selected)

        if not period:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:title_registration")

        y = period.year
        m = period.month
        kibetu = period.kibetu
        selected = f"{y}-{m:02d}"

        rows = self._fetch_rows(y, m, kibetu)
        if not rows:
            with transaction.atomic(using="rds"):
                insert_bonus_register_history(
                    "title_registration",
                    kibetu,
                    request.user.username,
                    "0件登録（対象データなし）",
                )
            messages.warning(request, "同じ期別の月タイトル登録データがありません。")
            return redirect(self._redirect_url(kibetu))

        # 事前チェック
        if self._exists_data(y, m):
            with transaction.atomic(using="rds"):
                insert_bonus_register_history(
                    "title_registration",
                    kibetu,
                    request.user.username,
                    "登録済みのため更新スキップ",
                )
            messages.warning(request, "すでに登録されています。")
            return redirect(self._redirect_url(kibetu))

        try:
            with transaction.atomic(using="rds"):
                # タイトル更新履歴に登録
                self._insert_rows(y, m, kibetu)

                # タイトルユーザーを更新
                self._update_title(y, m, kibetu)

                # 設定を更新
                self._update_setting(y, m)

                insert_bonus_register_history(
                    "title_registration",
                    kibetu,
                    request.user.username,
                    f"{len(rows)}件登録",
                )

        except IntegrityError as e:
            print("IntegrityError:", e)
            traceback.print_exc()
            messages.warning(request, "すでに登録されています。")
            return redirect(self._redirect_url(kibetu))

        except Exception as e:
            print("Exception:", e)
            traceback.print_exc()
            messages.error(request, f"登録中にエラーが発生しました: {e}")
            return redirect(self._redirect_url(kibetu))

        messages.success(request, "登録完了")
        return redirect(self._redirect_url(kibetu))



class RepurchaseLastMonthView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "repurchase_last_month.html"
    DEFAULT_PER_PAGE = 1000
    MAX_PER_PAGE = 2000

    def _get_month_choices(self):
        today = date.today().replace(day=1)
        return [
            {
                "value": format_target_month(
                    (today - relativedelta(months=i)).year,
                    (today - relativedelta(months=i)).month,
                ),
                "kibetu": format_target_kibetu(
                    (today - relativedelta(months=i)).year,
                    (today - relativedelta(months=i)).month,
                ),
                "year": (today - relativedelta(months=i)).year,
                "month": (today - relativedelta(months=i)).month,
            }
            for i in range(12)
        ]

    def _get_registered_months(self):
        sql = """
        SELECT DISTINCT CONCAT(register_year, '-', LPAD(register_month, 2, '0')) AS ym
        FROM bonus_db.purchase_info_list
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            return [row[0] for row in cursor.fetchall()]

    def _get_registered_month_options(self):
        sql = """
            SELECT
                register_year,
                register_month,
                COUNT(*) AS row_count
            FROM bonus_db.purchase_info_list
            GROUP BY register_year, register_month
            ORDER BY register_year DESC, register_month DESC
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            return [
                {
                    "value": format_target_month(row[0], row[1]),
                    "year": row[0],
                    "month": row[1],
                    "count": int(row[2]),
                }
                for row in cursor.fetchall()
            ]

    def _build_query_params(self, year, month):
        start = datetime(year, month, 1)
        end = start + relativedelta(months=1)
        return [year, month, start, end, year, month, year, month]

    def _count_rows(self, year, month):
        sql = f"""
SELECT COUNT(*)
FROM (
{REPURCHASE_LAST_MONTH}
) AS source_rows
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, self._build_query_params(year, month))
            row = cursor.fetchone()
            return int(row[0]) if row else 0

    def _fetch_rows(self, year, month, limit=None, offset=0):
        params = self._build_query_params(year, month)
        sql = REPURCHASE_LAST_MONTH

        with connections["rds"].cursor() as cursor:
            if limit is not None:
                sql = f"""
{sql}
ORDER BY payment_date ASC, order_code ASC, jwoa_code ASC
LIMIT %s OFFSET %s
"""
                params.extend([limit, offset])

            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    def _delete_rows(self, year, month):
        sql = """
DELETE FROM bonus_db.purchase_info_list
WHERE register_year = %s
  AND register_month = %s
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [year, month])
            return cursor.rowcount

    def _count_registered_rows(self, year, month):
        sql = """
            SELECT COUNT(*)
            FROM bonus_db.purchase_info_list
            WHERE register_year = %s
              AND register_month = %s
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [year, month])
            row = cursor.fetchone()
            return int(row[0]) if row else 0

    def _redirect_url(self, year=None, month=None):
        base_url = redirect("connect:repurchase_last_month").url
        if year is not None and month is not None:
            return f"{base_url}?target_year={year}&target_month={month}"
        return base_url

    def _enrich_month_context(self, ctx, year, month):
        selected_month = format_target_month(year, month)
        ctx["is_registered_month"] = selected_month in ctx.get("registered_months", [])
        ctx["registered_count"] = self._count_registered_rows(year, month)
        return ctx

    def _insert_rows(self, rows):
        insert_sql = """
INSERT INTO bonus_db.purchase_info_list
(
    register_year,
    register_month,
    order_year,
    order_month,
    jwoa_code,
    send_bv_name,
    order_code,
    total_bv,
    bv,
    order_type,
    deposit_at,
    order_at,
    bonus_payment_date
)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

        data = [
            (
                r["register_year"],
                r["register_month"],
                r["order_year"],
                r["order_month"],
                r["jwoa_code"],
                r["send_bv_name"],
                r["order_code"],
                r["total_bv"],
                r["bv"],
                r["order_type"],
                r["deposit_at"],
                r["order_at"],
                r["payment_date"],
            )
            for r in rows
        ]

        with connections["rds"].cursor() as cursor:
            cursor.executemany(insert_sql, data)

    def _parse_manual_int(self, request, field_name, label, min_value=None, max_value=None):
        raw_value = (request.POST.get(field_name) or "").strip()
        if raw_value == "":
            raise ValueError(f"{label}を入力してください。")

        try:
            value = int(raw_value)
        except ValueError as exc:
            raise ValueError(f"{label}は数値で入力してください。") from exc

        if min_value is not None and value < min_value:
            raise ValueError(f"{label}は{min_value}以上で入力してください。")
        if max_value is not None and value > max_value:
            raise ValueError(f"{label}は{max_value}以下で入力してください。")
        return value

    def _parse_manual_datetime(self, request, field_name, label):
        raw_value = (request.POST.get(field_name) or "").strip()
        if not raw_value:
            return None

        normalized = raw_value.replace("/", "-")
        if "T" in normalized:
            normalized = normalized.replace("T", " ")
        if len(normalized) == 16:
            normalized = f"{normalized}:00"

        try:
            return datetime.fromisoformat(normalized)
        except ValueError as exc:
            raise ValueError(
                f"{label}は YYYY-MM-DD HH:MM または YYYY/MM/DD HH:MM 形式で入力してください。"
            ) from exc

    def _build_manual_purchase_row(self, request):
        row = {
            "register_year": self._parse_manual_int(request, "manual_register_year", "登録年", 1900, 2100),
            "register_month": self._parse_manual_int(request, "manual_register_month", "登録月", 1, 12),
            "order_year": self._parse_manual_int(request, "manual_order_year", "注文年", 1900, 2100),
            "order_month": self._parse_manual_int(request, "manual_order_month", "注文月", 1, 12),
            "order_code": (request.POST.get("manual_order_code") or "").strip(),
            "order_type": self._parse_manual_int(request, "manual_order_type", "注文区分", 1),
            "jwoa_code": (request.POST.get("manual_jwoa_code") or "").strip(),
            "send_bv_name": (request.POST.get("manual_send_bv_name") or "").strip(),
            "total_bv": self._parse_manual_int(request, "manual_total_bv", "total_bv", 0),
            "bv": self._parse_manual_int(request, "manual_bv", "bv", 0),
            "deposit_at": self._parse_manual_datetime(request, "manual_deposit_at", "BV反映日時"),
            "order_at": self._parse_manual_datetime(request, "manual_order_at", "注文日時"),
            "payment_date": parse_input_date(request.POST.get("manual_bonus_payment_date")),
        }

        if not row["order_code"]:
            raise ValueError("注文番号を入力してください。")
        if not row["jwoa_code"]:
            raise ValueError("会員番号を入力してください。")
        if not row["send_bv_name"]:
            raise ValueError("会員名を入力してください。")
        if row["payment_date"] is None:
            raise ValueError("ボーナス支払日を入力してください。")

        return row

    def _insert_manual_purchase_row(self, row):
        insert_sql = """
INSERT INTO bonus_db.purchase_info_list
(
    register_year,
    register_month,
    order_year,
    order_month,
    jwoa_code,
    send_bv_name,
    order_code,
    total_bv,
    bv,
    order_type,
    deposit_at,
    order_at,
    bonus_payment_date
)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(
                insert_sql,
                [
                    row["register_year"],
                    row["register_month"],
                    row["order_year"],
                    row["order_month"],
                    row["jwoa_code"],
                    row["send_bv_name"],
                    row["order_code"],
                    row["total_bv"],
                    row["bv"],
                    row["order_type"],
                    row["deposit_at"],
                    row["order_at"],
                    row["payment_date"],
                ],
            )
            return cursor.lastrowid

    def _build_registered_search_context(self, ctx):
        selected_year = str(ctx.get("target_year") or "")
        selected_month = str(ctx.get("target_month") or "")
        filters = {
            "edit_register_year": (self.request.GET.get("edit_register_year") or selected_year or "").strip(),
            "edit_register_month": (self.request.GET.get("edit_register_month") or selected_month or "").strip(),
            "edit_order_year": (self.request.GET.get("edit_order_year") or "").strip(),
            "edit_order_month": (self.request.GET.get("edit_order_month") or "").strip(),
            "edit_order_code": (self.request.GET.get("edit_order_code") or "").strip(),
            "edit_jwoa_code": (self.request.GET.get("edit_jwoa_code") or "").strip(),
            "edit_name": (self.request.GET.get("edit_name") or "").strip(),
            "edit_bonus_payment_date_from": (
                self.request.GET.get("edit_bonus_payment_date_from") or ""
            ).strip(),
            "edit_bonus_payment_date_to": (
                self.request.GET.get("edit_bonus_payment_date_to") or ""
            ).strip(),
        }
        panel_opened = self.request.GET.get("edit_panel") == "1"
        submitted = any(key in self.request.GET for key in filters)

        ctx.update(filters)
        ctx["registered_edit_panel_open"] = panel_opened or submitted
        ctx["registered_edit_searched"] = submitted
        try:
            ctx["registered_edit_rows"] = self._fetch_registered_edit_rows(filters) if submitted else []
        except ValueError as e:
            messages.error(self.request, str(e))
            ctx["registered_edit_rows"] = []
        keep_params = {}
        for key in ("target_year", "target_month", "per_page"):
            value = ctx.get(key)
            if value not in ("", None):
                keep_params[key] = value
        if ctx["registered_edit_panel_open"]:
            keep_params["edit_panel"] = "1"
        for key, value in filters.items():
            if value:
                keep_params[key] = value
        ctx["registered_edit_query"] = urlencode(keep_params)
        return ctx

    def _fetch_registered_edit_rows(self, filters):
        where = []
        params = []

        if filters["edit_register_year"]:
            where.append("register_year = %s")
            params.append(filters["edit_register_year"])
        if filters["edit_register_month"]:
            where.append("register_month = %s")
            params.append(filters["edit_register_month"])
        if filters["edit_order_year"]:
            where.append("order_year = %s")
            params.append(filters["edit_order_year"])
        if filters["edit_order_month"]:
            where.append("order_month = %s")
            params.append(filters["edit_order_month"])
        if filters["edit_order_code"]:
            where.append("order_code LIKE %s")
            params.append(f"%{filters['edit_order_code']}%")
        if filters["edit_jwoa_code"]:
            where.append("jwoa_code LIKE %s")
            params.append(f"%{filters['edit_jwoa_code']}%")
        if filters["edit_name"]:
            where.append("send_bv_name LIKE %s")
            params.append(f"%{filters['edit_name']}%")
        payment_date_from = None
        payment_date_to = None
        if filters["edit_bonus_payment_date_from"]:
            payment_date_from = parse_input_date(filters["edit_bonus_payment_date_from"])
            where.append("bonus_payment_date >= %s")
            params.append(payment_date_from)
        if filters["edit_bonus_payment_date_to"]:
            payment_date_to = parse_input_date(filters["edit_bonus_payment_date_to"])
            where.append("bonus_payment_date <= %s")
            params.append(payment_date_to)
        if payment_date_from and payment_date_to and payment_date_from > payment_date_to:
            raise ValueError("ボーナス支払日のFromはTo以前の日付を指定してください。")

        if not where:
            return []

        sql = f"""
            SELECT
                id,
                register_year,
                register_month,
                order_year,
                order_month,
                order_code,
                order_type,
                jwoa_code,
                send_bv_name,
                total_bv,
                bv,
                deposit_at,
                order_at,
                bonus_payment_date,
                created_at,
                updated_at
            FROM bonus_db.purchase_info_list
            WHERE {" AND ".join(where)}
            ORDER BY register_year DESC, register_month DESC, id DESC
            LIMIT 100
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    def _fetch_manual_purchase_row_for_update(self, cursor, row_id):
        sql = """
            SELECT
                id,
                register_year,
                register_month,
                order_year,
                order_month,
                order_code,
                order_type,
                jwoa_code,
                send_bv_name,
                total_bv,
                bv,
                deposit_at,
                order_at,
                bonus_payment_date,
                created_at,
                updated_at
            FROM bonus_db.purchase_info_list
            WHERE id = %s
            FOR UPDATE
        """
        cursor.execute(sql, [row_id])
        row = cursor.fetchone()
        if not row:
            return None
        columns = [col[0] for col in cursor.description]
        return dict(zip(columns, row))

    def _update_manual_purchase_row(self, cursor, row_id, row):
        update_sql = """
            UPDATE bonus_db.purchase_info_list
            SET
                register_year = %s,
                register_month = %s,
                order_year = %s,
                order_month = %s,
                jwoa_code = %s,
                send_bv_name = %s,
                order_code = %s,
                total_bv = %s,
                bv = %s,
                order_type = %s,
                deposit_at = %s,
                order_at = %s,
                bonus_payment_date = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """
        cursor.execute(
            update_sql,
            [
                row["register_year"],
                row["register_month"],
                row["order_year"],
                row["order_month"],
                row["jwoa_code"],
                row["send_bv_name"],
                row["order_code"],
                row["total_bv"],
                row["bv"],
                row["order_type"],
                row["deposit_at"],
                row["order_at"],
                row["payment_date"],
                row_id,
            ],
        )
        return cursor.rowcount

    def _parse_manual_date_copy_target_ids(self, request, current_row_id):
        if request.POST.get("manual_apply_date_to_selected") != "1":
            return []

        raw_ids = (request.POST.get("manual_date_copy_target_ids") or "").strip()
        if not raw_ids:
            return []

        target_ids = []
        for raw_id in raw_ids.split(","):
            raw_id = raw_id.strip()
            if not raw_id:
                continue
            try:
                target_id = int(raw_id)
            except ValueError as exc:
                raise ValueError("日付データ反映対象IDが不正です。") from exc
            if target_id != current_row_id and target_id not in target_ids:
                target_ids.append(target_id)
        return target_ids

    def _fetch_manual_purchase_rows_for_update(self, cursor, row_ids):
        if not row_ids:
            return []

        placeholders = ", ".join(["%s"] * len(row_ids))
        sql = f"""
            SELECT
                id,
                register_year,
                register_month,
                order_year,
                order_month,
                order_code,
                order_type,
                jwoa_code,
                send_bv_name,
                total_bv,
                bv,
                deposit_at,
                order_at,
                bonus_payment_date,
                created_at,
                updated_at
            FROM bonus_db.purchase_info_list
            WHERE id IN ({placeholders})
            FOR UPDATE
        """
        cursor.execute(sql, row_ids)
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _update_manual_purchase_date_fields(self, cursor, row_ids, row):
        if not row_ids:
            return 0

        placeholders = ", ".join(["%s"] * len(row_ids))
        update_sql = f"""
            UPDATE bonus_db.purchase_info_list
            SET
                register_year = %s,
                register_month = %s,
                bonus_payment_date = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id IN ({placeholders})
        """
        cursor.execute(
            update_sql,
            [
                row["register_year"],
                row["register_month"],
                row["payment_date"],
                *row_ids,
            ],
        )
        return cursor.rowcount

    def _update_setting(self, year, month):
        value = f"{year}{month:02d}"

        sql = """
UPDATE bonus_db.settings
SET value = %s
WHERE name = 'set_title'
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [value])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_choice = (self.request.GET.get("target_month_choice") or "").strip()
        target_year = (self.request.GET.get("target_year") or "").strip()
        target_month = (self.request.GET.get("target_month") or "").strip()
        ctx["month_choices"] = self._get_month_choices()
        ctx["registered_months"] = self._get_registered_months()
        ctx["registered_month_options"] = self._get_registered_month_options()
        ctx["selected_month"] = ""
        ctx["selected_month_choice"] = selected_choice
        ctx["target_year"] = target_year
        ctx["target_month"] = target_month
        ctx["rows"] = []
        ctx["per_page"] = self.get_per_page()
        ctx["registered_count"] = 0
        ctx["is_registered_month"] = False

        if selected_choice or target_year or target_month:
            try:
                y, m = get_target_year_month_from_params(self.request.GET)
                ctx["selected_month"] = format_target_month(y, m)
                ctx["selected_month_choice"] = format_target_month(y, m)
                ctx["target_year"] = y
                ctx["target_month"] = m
                self._enrich_month_context(ctx, y, m)
                per_page = self.get_per_page()
                total_count = self._count_rows(y, m)
                total_pages = max(1, math.ceil(total_count / per_page))
                page = self.get_page_number(total_pages)
                offset = (page - 1) * per_page
                rows = self._fetch_rows(y, m, limit=per_page, offset=offset)
                base_params = {
                    "target_year": y,
                    "target_month": m,
                    "per_page": per_page,
                }
                ctx = self.set_page_context(
                    ctx=ctx,
                    rows=rows,
                    per_page=per_page,
                    total_count=total_count,
                    total_pages=total_pages,
                    page=page,
                    base_params=base_params,
                )
                return self._build_registered_search_context(ctx)
            except (ValueError, TypeError):
                messages.error(self.request, "年月の形式が不正です。年と月を正しく指定してください。")
                ctx["rows"] = []

        return self._build_registered_search_context(ctx)

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "register").strip()
        user_access = get_user_access(request.user)

        if action == "manual_create":
            if not user_access.can_create:
                messages.error(request, "登録権限がありません。")
                return redirect(self._redirect_url())

            try:
                row = self._build_manual_purchase_row(request)
            except ValueError as e:
                messages.error(request, str(e))
                return redirect(self._redirect_url())

            redirect_url = self._redirect_url(row["register_year"], row["register_month"])

            try:
                with transaction.atomic(using="rds"):
                    created_id = self._insert_manual_purchase_row(row)
                    after_values = dict(row)
                    after_values["id"] = created_id
                    record_change_audit(
                        request,
                        screen_name="ボーナス購入情報(登録/削除)",
                        action_type="create",
                        target_table="purchase_info_list",
                        target_pk=created_id,
                        summary=(
                            f"注文番号 {row['order_code']} / "
                            f"会員番号 {row['jwoa_code']} の購入情報を手入力登録"
                        ),
                        before_values=None,
                        after_values=after_values,
                    )
            except IntegrityError:
                logger.exception("ボーナス購入情報の手入力登録重複エラー")
                messages.error(request, "同じ登録年月・会員番号・注文番号のデータは既に登録されています。")
                return redirect(redirect_url)
            except Exception as e:
                logger.exception("ボーナス購入情報の手入力登録エラー")
                messages.error(request, f"登録中にエラーが発生しました: {e}")
                return redirect(redirect_url)

            messages.success(request, "購入情報を1件登録しました。")
            return redirect(redirect_url)

        if action == "manual_update":
            if not user_access.can_update:
                messages.error(request, "更新権限がありません。")
                return redirect(self._redirect_url())

            next_query = (request.POST.get("next_query") or "").strip()
            row_id_text = (request.POST.get("manual_row_id") or "").strip()
            try:
                row_id = int(row_id_text)
            except ValueError:
                messages.error(request, "更新対象IDが不正です。")
                return redirect(self._redirect_url())

            try:
                row = self._build_manual_purchase_row(request)
                date_copy_target_ids = self._parse_manual_date_copy_target_ids(request, row_id)
            except ValueError as e:
                messages.error(request, str(e))
                return redirect(self._redirect_url())

            if next_query:
                redirect_url = f"{self._redirect_url()}?{next_query}"
            else:
                redirect_url = self._redirect_url(row["register_year"], row["register_month"])

            date_copy_count = 0
            try:
                with transaction.atomic(using="rds"):
                    with connections["rds"].cursor() as cursor:
                        before_row = self._fetch_manual_purchase_row_for_update(cursor, row_id)
                        if not before_row:
                            messages.error(request, "更新対象データが見つかりませんでした。")
                            return redirect(redirect_url)

                        before_copy_rows = []
                        if date_copy_target_ids:
                            before_copy_rows = self._fetch_manual_purchase_rows_for_update(
                                cursor,
                                date_copy_target_ids,
                            )
                            found_ids = {int(r["id"]) for r in before_copy_rows}
                            if len(found_ids) != len(date_copy_target_ids):
                                messages.error(request, "日付データ反映対象に存在しないデータが含まれています。")
                                return redirect(redirect_url)

                        updated_count = self._update_manual_purchase_row(cursor, row_id, row)
                        if updated_count:
                            after_values = dict(row)
                            after_values["id"] = row_id
                            record_change_audit(
                                request,
                                screen_name="ボーナス購入情報(登録/削除)",
                                action_type="update",
                                target_table="purchase_info_list",
                                target_pk=row_id,
                                summary=(
                                    f"注文番号 {row['order_code']} / "
                                    f"会員番号 {row['jwoa_code']} の購入情報を編集"
                                ),
                                before_values=before_row,
                                after_values=after_values,
                            )

                        if date_copy_target_ids:
                            date_copy_count = self._update_manual_purchase_date_fields(
                                cursor,
                                date_copy_target_ids,
                                row,
                            )
                            if date_copy_count:
                                record_change_audit(
                                    request,
                                    screen_name="ボーナス購入情報(登録/削除)",
                                    action_type="bulk_update",
                                    target_table="purchase_info_list",
                                    target_pk=",".join(str(i) for i in date_copy_target_ids),
                                    summary=(
                                        "選択行へ日付データを反映: "
                                        f"登録年月 {row['register_year']}年{row['register_month']}月 / "
                                        f"ボーナス支払日 {row['payment_date']} / {date_copy_count}件"
                                    ),
                                    before_values={"rows": before_copy_rows},
                                    after_values={
                                        "ids": date_copy_target_ids,
                                        "register_year": row["register_year"],
                                        "register_month": row["register_month"],
                                        "bonus_payment_date": row["payment_date"],
                                    },
                                )
            except IntegrityError:
                logger.exception("ボーナス購入情報の手入力編集重複エラー")
                messages.error(request, "同じ登録年月・会員番号・注文番号のデータは既に登録されています。")
                return redirect(redirect_url)
            except Exception as e:
                logger.exception("ボーナス購入情報の手入力編集エラー")
                messages.error(request, f"更新中にエラーが発生しました: {e}")
                return redirect(redirect_url)

            if date_copy_count:
                messages.success(request, f"購入情報を更新しました。選択行 {date_copy_count}件へ日付データを反映しました。")
            else:
                messages.success(request, "購入情報を更新しました。")
            return redirect(redirect_url)

        if action == "manual_delete":
            if not user_access.can_delete:
                messages.error(request, "削除権限がありません。")
                return redirect(self._redirect_url())

            next_query = (request.POST.get("next_query") or "").strip()
            row_id_text = (request.POST.get("manual_row_id") or "").strip()
            try:
                row_id = int(row_id_text)
            except ValueError:
                messages.error(request, "削除対象IDが不正です。")
                return redirect(self._redirect_url())

            redirect_url = f"{self._redirect_url()}?{next_query}" if next_query else self._redirect_url()

            try:
                with transaction.atomic(using="rds"):
                    with connections["rds"].cursor() as cursor:
                        before_row = self._fetch_manual_purchase_row_for_update(cursor, row_id)
                        if not before_row:
                            messages.error(request, "削除対象データが見つかりませんでした。")
                            return redirect(redirect_url)

                        cursor.execute(
                            "DELETE FROM bonus_db.purchase_info_list WHERE id = %s",
                            [row_id],
                        )
                        deleted_count = cursor.rowcount

                    if deleted_count:
                        record_change_audit(
                            request,
                            screen_name="ボーナス購入情報(登録/削除)",
                            action_type="delete",
                            target_table="purchase_info_list",
                            target_pk=row_id,
                            summary=(
                                f"注文番号 {before_row.get('order_code')} / "
                                f"会員番号 {before_row.get('jwoa_code')} の購入情報を削除"
                            ),
                            before_values=before_row,
                            after_values=None,
                        )
            except Exception as e:
                logger.exception("ボーナス購入情報の手入力削除エラー")
                messages.error(request, f"削除中にエラーが発生しました: {e}")
                return redirect(redirect_url)

            messages.success(request, "購入情報を1件削除しました。")
            return redirect(redirect_url)

        if action == "delete_registered_month":
            if not user_access.can_delete:
                messages.error(request, "削除権限がありません。")
                return redirect(self._redirect_url())

            if not (
                request.POST.get("target_month_choice")
                or request.POST.get("target_year")
                or request.POST.get("target_month")
            ):
                messages.error(request, "削除する年月を指定してください。")
                return redirect(self._redirect_url())

            try:
                y, m = get_target_year_month_from_params(request.POST)
            except (ValueError, TypeError):
                messages.error(request, "年月の形式が不正です。年と月を正しく指定してください。")
                return redirect(self._redirect_url())

            redirect_url = self._redirect_url(y, m)

            try:
                with transaction.atomic(using="rds"):
                    before_count = self._count_registered_rows(y, m)
                    deleted_count = self._delete_rows(y, m)
                    if deleted_count:
                        record_change_audit(
                            request,
                            screen_name="ボーナス購入情報(登録/削除)",
                            action_type="bulk_delete",
                            target_table="purchase_info_list",
                            target_pk=f"{y}-{m:02d}",
                            summary=f"{y}年{m}月 の購入情報を削除: {deleted_count}件",
                            before_values={"year": y, "month": m, "count": before_count},
                            after_values=None,
                        )
            except Exception as e:
                logger.exception("ボーナス購入情報登録の年月削除エラー")
                messages.error(request, f"削除中にエラーが発生しました: {e}")
                return redirect(redirect_url)

            if deleted_count:
                messages.success(
                    request,
                    f"{y}年{m}月 の購入情報を {deleted_count}件削除しました。",
                )
            else:
                messages.warning(request, f"{y}年{m}月 の削除対象データはありません。")
            return redirect(redirect_url)

        if not (
            request.POST.get("target_month_choice")
            or request.POST.get("target_year")
            or request.POST.get("target_month")
        ):
            messages.error(request, "年月未選択です。")
            return redirect("connect:repurchase_last_month")

        try:
            y, m = get_target_year_month_from_params(request.POST)
        except (ValueError, TypeError):
            messages.error(request, "年月の形式が不正です。年と月を正しく指定してください。")
            return redirect("connect:repurchase_last_month")

        redirect_url = self._redirect_url(y, m)

        if action != "register":
            messages.error(request, "不正な操作です。")
            return redirect(redirect_url)

        if not user_access.can_execute:
            messages.error(request, "実行権限がありません。")
            return redirect(redirect_url)

        rows = self._fetch_rows(y, m)

        if not rows:
            messages.info(request, "対象データなし")
            return redirect(redirect_url)

        try:
            with transaction.atomic(using="rds"):
                self._delete_rows(y, m)
                self._insert_rows(rows)
                self._update_setting(y, m)
                record_change_audit(
                    request,
                    screen_name="ボーナス購入情報(登録/削除)",
                    action_type="bulk_create",
                    target_table="purchase_info_list",
                    target_pk=f"{y}-{m:02d}",
                    summary=f"{y}年{m}月 の購入情報を登録: {len(rows)}件",
                    before_values={"year": y, "month": m},
                    after_values={"year": y, "month": m, "count": len(rows)},
                )

        except Exception as e:
            logger.exception("ボーナス購入情報登録エラー")
            messages.error(request, f"エラー発生: {e}")
            return redirect(redirect_url)

        messages.success(request, f"{len(rows)}件登録完了")
        return redirect(redirect_url)


class RepurchaseExportView(RepurchaseListView):

    def get(self, request):
        selected_month = (request.GET.get("target_month") or "").strip()
        q_kibetu = (request.GET.get("q_kibetu") or "").strip()

        q_code = (request.GET.get("q_code") or "").strip()
        q_name = (request.GET.get("q_name") or "").strip()
        q_order_code = (request.GET.get("q_order_code") or "").strip()

        # ←修正
        q_order_types = request.GET.getlist("q_order_type")
        q_order_types = [x for x in q_order_types if x]

        q_bonus_date_from = (request.GET.get("q_bonus_date_from") or "").strip()
        q_bonus_date_to = (request.GET.get("q_bonus_date_to") or "").strip()
        selected_kibetu_period = self._resolve_kibetu_period(q_kibetu)
        if selected_kibetu_period:
            q_bonus_date_from = selected_kibetu_period.st_date.strftime("%Y-%m-%d")
            q_bonus_date_to = selected_kibetu_period.end_date.strftime("%Y-%m-%d")

        year = None
        month = None

        if selected_month:
            try:
                year, month = map(int, selected_month.split("-"))
            except ValueError:
                pass

        rows = self._fetch_rows(
            year=year,
            month=month,
            q_code=q_code,
            q_name=q_name,
            q_order_code=q_order_code,

            # ←修正
            q_order_types=q_order_types,

            q_bonus_date_from=q_bonus_date_from,
            q_bonus_date_to=q_bonus_date_to,
            limit=1000000,
            offset=0,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "購入情報一覧"

        ws.append([
            "登録年",
            "登録月",
            "注文年",
            "注文月",
            "注文番号",
            "注文区分",
            "会員番号",
            "会員名",
            "total_bv",
            "bv",
            "BV反映FLG",
            "BV反映日時",
            "注文日時",
            "ボーナス支払日",
            "作成日時",
        ])

        for r in rows:

            if r["order_type"] == 101:
                order_type_name = "再購入品"
            elif r["order_type"] == 102:
                order_type_name = "初回購入品"
            elif r["order_type"] == 103:
                order_type_name = "ランクアップ購入品"
            elif r["order_type"] == 105:
                order_type_name = "特別対応購入品"
            elif r["order_type"] == 200:
                order_type_name = "クーリングオフ"
            else:
                order_type_name = r["order_type"]

            ws.append([
                r["register_year"],
                r["register_month"],
                r["order_year"],
                r["order_month"],
                r["order_code"],
                order_type_name,
                r["jwoa_code"],
                r["send_bv_name"],
                r["total_bv"],
                r["bv"],
                self.BV_ACTIVED_FLG_LABELS.get(
                    r["bv_actived_flg"],
                    r["bv_actived_flg"],
                ),
                _excel_jst_datetime(r["deposit_at"]),
                _excel_jst_datetime(r["order_at"]),
                r["bonus_payment_date"],
                _excel_jst_datetime(r["created_at"]),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="repurchase.xlsx"'
        )

        wb.save(response)
        return response



class BonusPaymentDateView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "bonus_payment_date.html"

    DEFAULT_PER_PAGE = 500
    MAX_PER_PAGE = 500
    SORT_COLUMNS = {
        "order_code": "order_code",
        "bonus_payment_date": "bonus_payment_date",
        "created_at": "created_at",
    }

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="created_at",
            default_direction="desc",
        )

    def _build_where(self, q_order_code: str = ""):
        where = ["1=1"]
        params = []

        if q_order_code:
            where.append("order_code LIKE %s")
            params.append(f"%{q_order_code}%")

        return "WHERE " + " AND ".join(where), params

    def _count_rows(self, q_order_code: str = ""):
        where_sql, params = self._build_where(q_order_code=q_order_code)
        sql = f"""
SELECT COUNT(*)
FROM bonus_db.bonus_payment_date
{where_sql}
"""
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
            return int(row[0]) if row else 0

    def _fetch_rows(self, q_order_code: str = "", limit=500, offset=0, order_sql=""):
        where_sql, params = self._build_where(q_order_code=q_order_code)
        sql = """
SELECT
    order_code,
    bonus_payment_date,
    created_at
FROM bonus_db.bonus_payment_date
{where_sql}
ORDER BY {order_sql}, order_code ASC
LIMIT %s OFFSET %s
""".format(where_sql=where_sql, order_sql=order_sql or "created_at DESC")
        params.extend([limit, offset])

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, r)) for r in cursor.fetchall()]

    def _parse_uploaded_rows(self, uploaded_file):
        filename = (uploaded_file.name or "").lower()

        if filename.endswith((".xlsx", ".xlsm")):
            workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = workbook.active
            source_rows = sheet.iter_rows(values_only=True)
        else:
            raw = uploaded_file.read()
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("cp932")
            source_rows = csv.reader(io.StringIO(text))

        rows = []
        errors = []

        for row_no, row in enumerate(source_rows, start=1):
            values = list(row or [])
            if not any(str(value or "").strip() for value in values):
                continue

            if len(values) < 2:
                errors.append(f"{row_no}行目: 注文番号と支払日を指定してください。")
                continue

            order_code = str(values[0] or "").strip()
            payment_value = values[1]

            if row_no == 1 and (
                "注文" in order_code or "order" in order_code.lower()
            ):
                continue

            if not order_code:
                errors.append(f"{row_no}行目: 注文番号が空です。")
                continue

            if isinstance(payment_value, datetime):
                parsed_payment_date = payment_value.date()
            elif isinstance(payment_value, date):
                parsed_payment_date = payment_value
            else:
                try:
                    parsed_payment_date = parse_input_date(payment_value)
                except ValueError as exc:
                    errors.append(f"{row_no}行目: {exc}")
                    continue

            if not parsed_payment_date:
                errors.append(f"{row_no}行目: 支払日を指定してください。")
                continue

            rows.append(
                {
                    "order_code": order_code,
                    "bonus_payment_date": parsed_payment_date,
                    "payment_year": parsed_payment_date.year,
                    "payment_month": parsed_payment_date.month,
                }
            )

        return rows, errors

    def _bulk_upsert_rows(self, rows):
        upsert_payment_sql = """
            INSERT INTO bonus_db.bonus_payment_date (
                order_code,
                bonus_payment_date
            ) VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE
                bonus_payment_date = VALUES(bonus_payment_date)
        """
        update_purchase_sql = """
            UPDATE bonus_db.purchase_info_list
            SET
                bonus_payment_date = %s,
                register_year = %s,
                register_month = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE order_code = %s
        """

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                for row in rows:
                    cursor.execute(
                        upsert_payment_sql,
                        [row["order_code"], row["bonus_payment_date"]],
                    )
                    cursor.execute(
                        update_purchase_sql,
                        [
                            row["bonus_payment_date"],
                            row["payment_year"],
                            row["payment_month"],
                            row["order_code"],
                        ],
                    )

    def _fetch_source_payment_date(self, order_code):
        """支払日の上書きを消したときに戻す元の日付を返す。

        購入情報の元データは COALESCE(この画面の支払日, 入金日) で支払日を決めて
        いるので、上書きが無い状態の値は入金日になる。注文区分105（特別対応）は
        orders に無く api_users_bv の支払日が元の値になるため、そちらも見る。
        どちらも無いときは None を返す。
        """
        lookups = [
            """
            SELECT DATE(deposit_at)
            FROM nexus_production.orders
            WHERE order_code = %s
              AND deposit_at IS NOT NULL
            LIMIT 1
            """,
            """
            SELECT DATE(payment_date)
            FROM nexus_production.api_users_bv
            WHERE doc_no = %s
              AND payment_date IS NOT NULL
            LIMIT 1
            """,
        ]

        with connections["rds"].cursor() as cursor:
            for sql in lookups:
                cursor.execute(sql, [order_code])
                row = cursor.fetchone()
                if row and row[0]:
                    return row[0]

        return None

    def _revert_purchase_info_payment_date(self, cursor, order_code, source_date):
        """購入情報一覧の支払日と登録年月を、上書き前の日付に戻す。"""
        sql = """
        UPDATE bonus_db.purchase_info_list
        SET
            bonus_payment_date = %s,
            register_year = %s,
            register_month = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE order_code = %s
        """

        cursor.execute(
            sql,
            [
                source_date,
                source_date.year,
                source_date.month,
                order_code,
            ],
        )
        return cursor.rowcount

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        q_order_code = (self.request.GET.get("q_order_code") or "").strip()

        ctx["q_order_code"] = q_order_code

        per_page = self.get_per_page()
        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        total_count = self._count_rows(q_order_code=q_order_code)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            q_order_code=q_order_code,
            limit=per_page,
            offset=offset,
            order_sql=sort_ctx["order_sql"],
        )

        base_params = {}
        if q_order_code:
            base_params["q_order_code"] = q_order_code

        add_sort_params(base_params, sort_ctx)

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip()
        order_code = (request.POST.get("order_code") or "").strip()
        bonus_payment_date = (request.POST.get("bonus_payment_date") or "").strip()

        q_order_code = (request.POST.get("q_order_code") or "").strip()
        redirect_path = f"/bonus_payment_date/?q_order_code={q_order_code}"

        try:
            parsed_bonus_payment_date = parse_input_date(bonus_payment_date)
        except ValueError as e:
            messages.error(request, str(e))
            return redirect(redirect_path)
        payment_year = parsed_bonus_payment_date.year if parsed_bonus_payment_date else None
        payment_month = parsed_bonus_payment_date.month if parsed_bonus_payment_date else None

        if action == "create":
            if not order_code:
                messages.error(request, "注文番号を入力してください。")
                return redirect("connect:bonus_payment_date")

            insert_sql = """
            INSERT INTO bonus_db.bonus_payment_date (
                order_code,
                bonus_payment_date
            ) VALUES (%s, %s)
            """

            update_sql = """
            UPDATE bonus_db.purchase_info_list
            SET
                bonus_payment_date = %s,
                register_year = COALESCE(%s, register_year),
                register_month = COALESCE(%s, register_month),
                updated_at = CURRENT_TIMESTAMP
            WHERE order_code = %s
            """

            try:
                with transaction.atomic(using="rds"):
                    with connections["rds"].cursor() as cursor:
                        cursor.execute(insert_sql, [order_code, parsed_bonus_payment_date])
                        cursor.execute(
                            update_sql,
                            [
                                parsed_bonus_payment_date,
                                payment_year,
                                payment_month,
                                order_code,
                            ],
                        )
                    record_change_audit(
                        request,
                        screen_name="注文別ボーナス支払日",
                        action_type="create",
                        target_table="bonus_payment_date",
                        target_pk=order_code,
                        summary=f"注文番号 {order_code} のボーナス支払日を登録",
                        before_values=None,
                        after_values={
                            "order_code": order_code,
                            "bonus_payment_date": parsed_bonus_payment_date,
                        },
                    )

                messages.success(request, "登録しました。")
            except Exception as e:
                messages.error(request, f"登録に失敗しました: {e}")

            return redirect(redirect_path)

        elif action == "bulk_create":
            uploaded_file = request.FILES.get("payment_date_file")
            if not uploaded_file:
                messages.error(request, "一括登録ファイルを選択してください。")
                return redirect(redirect_path)

            try:
                rows, errors = self._parse_uploaded_rows(uploaded_file)
            except Exception as e:
                messages.error(request, f"ファイルの読み込みに失敗しました: {e}")
                return redirect(redirect_path)

            if errors:
                messages.error(request, " / ".join(errors[:5]))
                if len(errors) > 5:
                    messages.error(request, f"他 {len(errors) - 5} 件のエラーがあります。")
                return redirect(redirect_path)

            if not rows:
                messages.error(request, "登録できるデータがありません。")
                return redirect(redirect_path)

            try:
                self._bulk_upsert_rows(rows)
                record_change_audit(
                    request,
                    screen_name="注文別ボーナス支払日",
                    action_type="bulk_create",
                    target_table="bonus_payment_date",
                    target_pk=None,
                    summary=f"ボーナス支払日を {len(rows)}件 一括登録/更新",
                    before_values=None,
                    after_values={"count": len(rows), "sample": rows[:5]},
                )
                messages.success(request, f"{len(rows)}件を一括登録しました。")
            except Exception as e:
                messages.error(request, f"一括登録に失敗しました: {e}")

            return redirect(redirect_path)

        elif action == "update":
            if not order_code:
                messages.error(request, "注文番号が不正です。")
                return redirect(redirect_path)

            sql1 = """
            UPDATE bonus_db.bonus_payment_date
            SET bonus_payment_date = %s
            WHERE order_code = %s
            """

            sql2 = """
            UPDATE bonus_db.purchase_info_list
            SET
                bonus_payment_date = %s,
                register_year = COALESCE(%s, register_year),
                register_month = COALESCE(%s, register_month),
                updated_at = CURRENT_TIMESTAMP
            WHERE order_code = %s
            """

            try:
                before_row = fetch_one_dict(
                    "rds",
                    """
                        SELECT order_code, bonus_payment_date
                        FROM bonus_db.bonus_payment_date
                        WHERE order_code = %s
                    """,
                    [order_code],
                )
                with transaction.atomic(using="rds"):
                    with connections["rds"].cursor() as cursor:
                        cursor.execute(sql1, [parsed_bonus_payment_date, order_code])
                        updated_count = cursor.rowcount
                        cursor.execute(
                            sql2,
                            [
                                parsed_bonus_payment_date,
                                payment_year,
                                payment_month,
                                order_code,
                            ],
                        )
                    if updated_count:
                        record_change_audit(
                            request,
                            screen_name="注文別ボーナス支払日",
                            action_type="update",
                            target_table="bonus_payment_date",
                            target_pk=order_code,
                            summary=f"注文番号 {order_code} のボーナス支払日を更新",
                            before_values=before_row,
                            after_values={
                                "order_code": order_code,
                                "bonus_payment_date": parsed_bonus_payment_date,
                            },
                        )

                messages.success(request, "更新しました。")
            except Exception as e:
                messages.error(request, f"更新に失敗しました: {e}")

            return redirect(redirect_path)

        elif action == "delete":
            if not order_code:
                messages.error(request, "注文番号が不正です。")
                return redirect(redirect_path)

            sql = """
            DELETE FROM bonus_db.bonus_payment_date
            WHERE order_code = %s
            """

            try:
                before_row = fetch_one_dict(
                    "rds",
                    """
                        SELECT order_code, bonus_payment_date
                        FROM bonus_db.bonus_payment_date
                        WHERE order_code = %s
                    """,
                    [order_code],
                )
                source_date = self._fetch_source_payment_date(order_code)

                with transaction.atomic(using="rds"):
                    with connections["rds"].cursor() as cursor:
                        cursor.execute(sql, [order_code])
                        deleted_count = cursor.rowcount

                        if source_date:
                            reverted_count = self._revert_purchase_info_payment_date(
                                cursor,
                                order_code,
                                source_date,
                            )
                        else:
                            reverted_count = 0

                if deleted_count:
                    record_change_audit(
                        request,
                        screen_name="注文別ボーナス支払日",
                        action_type="delete",
                        target_table="bonus_payment_date",
                        target_pk=order_code,
                        summary=(
                            f"注文番号 {order_code} のボーナス支払日を削除"
                            + (
                                f"（購入情報の支払日を {source_date} に戻す）"
                                if source_date
                                else ""
                            )
                        ),
                        before_values=before_row,
                        after_values={
                            "purchase_info_list.bonus_payment_date": source_date,
                            "reverted_count": reverted_count,
                        },
                    )

                if source_date:
                    messages.success(
                        request,
                        f"削除しました。ボーナス購入情報一覧の支払日を {source_date} に戻しました。",
                    )
                else:
                    messages.success(request, "削除しました。")
                    messages.warning(
                        request,
                        "入金日が見つからないため、ボーナス購入情報一覧の支払日は変更していません。",
                    )
            except Exception as e:
                messages.error(request, f"削除に失敗しました: {e}")

            return redirect(redirect_path)

        messages.error(request, "不正な操作です。")
        return redirect(redirect_path)


class BonusPaymentDateTemplateView(generic.View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "注文別ボーナス支払日"
        ws.append(["注文番号", "支払日"])
        ws.append(["MF000000000000", "2026-01-05"])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="bonus_payment_date_template.xlsx"'
        )
        wb.save(response)
        return response


class ActiveUsersTemplateView(generic.View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "アクティブ会員"
        ws.append(["会員コード", "年", "月", "ステータス"])
        ws.append(["JP05215357", 2025, 7, 1])
        ws.append(["JP05215358", 2025, 7, 0])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="active_users_template.xlsx"'
        )
        wb.save(response)
        return response


class ActiveUsersView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "active_users.html"
    DEFAULT_PER_PAGE = 500
    MAX_PER_PAGE = 500
    BULK_UPSERT_CHUNK_SIZE = 1000
    BULK_UPSERT_MAX_RETRIES = 3
    BULK_UPSERT_RETRY_WAIT_SECONDS = 2

    ACTIVE_STATUS_CHOICES = (
        ("1", "アクティブ"),
        ("0", "非アクティブ"),
    )

    def _build_where(self, q_jwoa_code="", q_year="", q_month="", q_active_status=""):
        where_clauses = []
        params = []

        if q_jwoa_code:
            where_clauses.append("a.jwoa_code LIKE %s")
            params.append(f"%{q_jwoa_code}%")

        if q_year:
            where_clauses.append("a.year = %s")
            params.append(int(q_year))

        if q_month:
            where_clauses.append("a.month = %s")
            params.append(int(q_month))

        if q_active_status != "":
            where_clauses.append("a.active_status = %s")
            params.append(int(q_active_status))

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
        return where_sql, params

    def _count_rows(self, q_jwoa_code="", q_year="", q_month="", q_active_status=""):
        where_sql, params = self._build_where(
            q_jwoa_code=q_jwoa_code,
            q_year=q_year,
            q_month=q_month,
            q_active_status=q_active_status,
        )
        sql = f"""
            SELECT COUNT(*)
            FROM active_users a
            {where_sql}
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
            return int(row[0]) if row else 0

    def _fetch_rows(
        self,
        q_jwoa_code="",
        q_year="",
        q_month="",
        q_active_status="",
        limit=500,
        offset=0,
    ):
        where_sql, params = self._build_where(
            q_jwoa_code=q_jwoa_code,
            q_year=q_year,
            q_month=q_month,
            q_active_status=q_active_status,
        )

        sql = f"""
            SELECT
                a.id,
                a.jwoa_code,
                a.year,
                a.month,
                a.active_status,
                a.created_at,
                u.send_bv_name
            FROM active_users a
            LEFT JOIN nexus_production.users u
                ON a.jwoa_code = u.jmoa_code
            {where_sql}
            ORDER BY a.jwoa_code, a.year DESC, a.month DESC
            LIMIT %s OFFSET %s
        """
        params.extend([limit, offset])

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        q_jwoa_code = self.request.GET.get("q_jwoa_code", "").strip()
        q_year = self.request.GET.get("q_year", "").strip()
        q_month = self.request.GET.get("q_month", "").strip()
        q_active_status = self.request.GET.get("q_active_status", "").strip()

        ctx["q_jwoa_code"] = q_jwoa_code
        ctx["q_year"] = q_year
        ctx["q_month"] = q_month
        ctx["q_active_status"] = q_active_status
        ctx["active_status_choices"] = self.ACTIVE_STATUS_CHOICES

        try:
            total_count = self._count_rows(
                q_jwoa_code,
                q_year,
                q_month,
                q_active_status,
            )
        except ValueError:
            messages.error(self.request, "年・月・ステータスは数値で入力してください。")
            total_count = 0

        per_page = self.get_per_page()
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = []
        if total_count:
            rows = self._fetch_rows(
                q_jwoa_code=q_jwoa_code,
                q_year=q_year,
                q_month=q_month,
                q_active_status=q_active_status,
                limit=per_page,
                offset=offset,
            )

        base_params = {}
        if q_jwoa_code:
            base_params["q_jwoa_code"] = q_jwoa_code
        if q_year:
            base_params["q_year"] = q_year
        if q_month:
            base_params["q_month"] = q_month
        if q_active_status != "":
            base_params["q_active_status"] = q_active_status

        ctx = self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )

        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")

        q_jwoa_code = request.POST.get("q_jwoa_code", "").strip()
        q_year = request.POST.get("q_year", "").strip()
        q_month = request.POST.get("q_month", "").strip()
        q_active_status = request.POST.get("q_active_status", "").strip()

        if action == "create":
            return self._create(request, q_jwoa_code, q_year, q_month, q_active_status)

        if action == "bulk_create":
            return self._bulk_create(request, q_jwoa_code, q_year, q_month, q_active_status)

        if action == "update":
            return self._update(request, q_jwoa_code, q_year, q_month, q_active_status)

        if action == "delete":
            return self._delete(request, q_jwoa_code, q_year, q_month, q_active_status)

        messages.error(request, "不正な操作です。")
        return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

    def _create(self, request, q_jwoa_code, q_year, q_month, q_active_status):
        jwoa_code = request.POST.get("jwoa_code", "").strip()
        year = request.POST.get("year", "").strip()
        month = request.POST.get("month", "").strip()
        active_status = request.POST.get("active_status", "1").strip()

        error_message = self._validate_input(jwoa_code, year, month, active_status)
        if error_message:
            messages.error(request, error_message)
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        sql = """
            INSERT INTO active_users (
                jwoa_code,
                year,
                month,
                active_status,
                created_at
            ) VALUES (
                %s,
                %s,
                %s,
                %s,
                NOW()
            )
        """

        try:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.execute(
                        sql,
                        [jwoa_code, int(year), int(month), int(active_status)],
                    )
                    created_id = cursor.lastrowid

                record_change_audit(
                    request,
                    screen_name="アクティブ会員管理",
                    action_type="create",
                    target_table="active_users",
                    target_pk=created_id,
                    summary=f"{jwoa_code} {year}年{month}月 を登録",
                    before_values=None,
                    after_values={
                        "id": created_id,
                        "jwoa_code": jwoa_code,
                        "year": int(year),
                        "month": int(month),
                        "active_status": int(active_status),
                    },
                )

            messages.success(request, "登録しました。")

        except IntegrityError as e:
            error_text = str(e)

            if "uq_active_users_jwoa_year_month" in error_text or "Duplicate entry" in error_text:
                messages.error(request, "この会員コード・年・月のデータはすでに登録されています。")
            else:
                messages.error(request, "登録に失敗しました。会員コードが存在しない可能性があります。")

        except Exception as e:
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

    def _bulk_create(self, request, q_jwoa_code, q_year, q_month, q_active_status):
        uploaded_file = request.FILES.get("active_users_file")
        if not uploaded_file:
            messages.error(request, "一括登録ファイルを選択してください。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        try:
            rows, errors = self._parse_uploaded_rows(uploaded_file)
        except Exception as e:
            messages.error(request, f"ファイルの読み込みに失敗しました: {e}")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        if errors:
            messages.error(request, " / ".join(errors[:5]))
            if len(errors) > 5:
                messages.error(request, f"他 {len(errors) - 5} 件のエラーがあります。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        if not rows:
            messages.error(request, "登録できるデータがありません。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        try:
            self._bulk_upsert_rows(rows)
            record_change_audit(
                request,
                screen_name="アクティブ会員管理",
                action_type="bulk_create",
                target_table="active_users",
                target_pk=None,
                summary=f"アクティブ会員を {len(rows)}件 一括登録/更新",
                before_values=None,
                after_values={"count": len(rows), "sample": rows[:5]},
            )
            messages.success(request, f"{len(rows)}件を一括登録しました。")
        except IntegrityError:
            messages.error(request, "一括登録に失敗しました。会員コードが存在しない可能性があります。")
        except Exception as e:
            messages.error(request, f"一括登録中にエラーが発生しました: {e}")

        return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

    def _parse_uploaded_rows(self, uploaded_file):
        filename = (uploaded_file.name or "").lower()

        if filename.endswith((".xlsx", ".xlsm")):
            workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = workbook.active
            source_rows = sheet.iter_rows(values_only=True)
        else:
            raw = uploaded_file.read()
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("cp932")
            source_rows = csv.reader(io.StringIO(text))

        rows = []
        errors = []

        for row_no, row in enumerate(source_rows, start=1):
            values = list(row or [])
            if not any(str(value or "").strip() for value in values):
                continue

            if len(values) < 3:
                errors.append(f"{row_no}行目: 会員コード・年・月を指定してください。")
                continue

            jwoa_code = str(values[0] or "").strip()
            year = self._normalize_bulk_number(values[1])
            month = self._normalize_bulk_number(values[2])
            active_status = self._normalize_bulk_status(values[3]) if len(values) >= 4 else "1"

            if row_no == 1 and (
                "会員" in jwoa_code
                or "jwoa" in jwoa_code.lower()
                or "code" in jwoa_code.lower()
            ):
                continue

            error_message = self._validate_input(jwoa_code, year, month, active_status)
            if error_message:
                errors.append(f"{row_no}行目: {error_message}")
                continue

            rows.append(
                {
                    "jwoa_code": jwoa_code,
                    "year": int(year),
                    "month": int(month),
                    "active_status": int(active_status),
                }
            )

        return rows, errors

    def _normalize_bulk_number(self, value):
        if value in (None, ""):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_bulk_status(self, value):
        raw_value = self._normalize_bulk_number(value)
        normalized = raw_value.lower()
        if normalized in ("", "1", "active", "アクティブ"):
            return "1"
        if normalized in ("0", "inactive", "非アクティブ"):
            return "0"
        return raw_value

    def _bulk_upsert_rows(self, rows):
        sql = """
            INSERT INTO active_users (
                jwoa_code,
                year,
                month,
                active_status,
                created_at
            ) VALUES (
                %s,
                %s,
                %s,
                %s,
                NOW()
            )
            ON DUPLICATE KEY UPDATE
                active_status = VALUES(active_status)
        """

        for start in range(0, len(rows), self.BULK_UPSERT_CHUNK_SIZE):
            chunk = rows[start:start + self.BULK_UPSERT_CHUNK_SIZE]
            data = [
                (
                    row["jwoa_code"],
                    row["year"],
                    row["month"],
                    row["active_status"],
                )
                for row in chunk
            ]
            self._execute_bulk_upsert_chunk(sql, data)

    def _execute_bulk_upsert_chunk(self, sql, data):
        for attempt in range(1, self.BULK_UPSERT_MAX_RETRIES + 1):
            try:
                with transaction.atomic(using="rds"):
                    with connections["rds"].cursor() as cursor:
                        cursor.executemany(sql, data)
                return
            except OperationalError as e:
                if not self._is_lock_wait_timeout(e) or attempt >= self.BULK_UPSERT_MAX_RETRIES:
                    raise
                time.sleep(self.BULK_UPSERT_RETRY_WAIT_SECONDS)

    def _is_lock_wait_timeout(self, error):
        return any("1205" in str(arg) for arg in getattr(error, "args", ()))

    def _update(self, request, q_jwoa_code, q_year, q_month, q_active_status):
        row_id = request.POST.get("id", "").strip()
        jwoa_code = request.POST.get("jwoa_code", "").strip()
        year = request.POST.get("year", "").strip()
        month = request.POST.get("month", "").strip()
        active_status = request.POST.get("active_status", "1").strip()

        if not row_id:
            messages.error(request, "更新対象IDがありません。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        try:
            row_id_int = int(row_id)
        except ValueError:
            messages.error(request, "更新対象IDが不正です。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        error_message = self._validate_input(jwoa_code, year, month, active_status)
        if error_message:
            messages.error(request, error_message)
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        before_row = fetch_one_dict(
            "rds",
            """
                SELECT id, jwoa_code, year, month, active_status, created_at
                FROM active_users
                WHERE id = %s
            """,
            [row_id_int],
        )

        sql = """
            UPDATE active_users
            SET
                jwoa_code = %s,
                year = %s,
                month = %s,
                active_status = %s
            WHERE id = %s
        """

        try:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.execute(
                        sql,
                        [jwoa_code, int(year), int(month), int(active_status), row_id_int],
                    )
                    updated_count = cursor.rowcount

                if updated_count:
                    after_row = {
                        "id": row_id_int,
                        "jwoa_code": jwoa_code,
                        "year": int(year),
                        "month": int(month),
                        "active_status": int(active_status),
                    }
                    record_change_audit(
                        request,
                        screen_name="アクティブ会員管理",
                        action_type="update",
                        target_table="active_users",
                        target_pk=row_id_int,
                        summary=f"アクティブ会員 {row_id_int} を更新",
                        before_values=before_row,
                        after_values=after_row,
                    )

            messages.success(request, "更新しました。")

        except IntegrityError as e:
            error_text = str(e)

            if "uq_active_users_jwoa_year_month" in error_text or "Duplicate entry" in error_text:
                messages.error(request, "この会員コード・年・月のデータはすでに登録されています。")
            else:
                messages.error(request, "更新に失敗しました。会員コードが存在しない可能性があります。")

        except Exception as e:
            messages.error(request, f"更新中にエラーが発生しました: {e}")

        return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

    def _delete(self, request, q_jwoa_code, q_year, q_month, q_active_status):
        row_id = request.POST.get("id", "").strip()

        if not row_id:
            messages.error(request, "削除対象IDがありません。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        try:
            row_id_int = int(row_id)
        except ValueError:
            messages.error(request, "削除対象IDが不正です。")
            return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

        sql = """
            DELETE FROM active_users
            WHERE id = %s
        """

        try:
            with transaction.atomic(using="rds"):
                before_row = fetch_one_dict(
                    "rds",
                    """
                        SELECT id, jwoa_code, year, month, active_status, created_at
                        FROM active_users
                        WHERE id = %s
                    """,
                    [row_id_int],
                )
                with connections["rds"].cursor() as cursor:
                    cursor.execute(sql, [row_id_int])
                    deleted_count = cursor.rowcount

                if deleted_count:
                    record_change_audit(
                        request,
                        screen_name="アクティブ会員管理",
                        action_type="delete",
                        target_table="active_users",
                        target_pk=row_id_int,
                        summary=f"アクティブ会員 {row_id_int} を削除",
                        before_values=before_row,
                        after_values=None,
                    )

            messages.success(request, "削除しました。")

        except Exception as e:
            messages.error(request, f"削除中にエラーが発生しました: {e}")

        return redirect(self._get_redirect_url(q_jwoa_code, q_year, q_month, q_active_status))

    def _validate_input(self, jwoa_code, year, month, active_status="1"):
        if not jwoa_code:
            return "会員コードを入力してください。"

        if not year:
            return "年を入力してください。"

        if not month:
            return "月を入力してください。"

        try:
            year_int = int(year)
        except ValueError:
            return "年は数値で入力してください。"

        try:
            month_int = int(month)
        except ValueError:
            return "月は数値で入力してください。"

        if year_int < 1900 or year_int > 2100:
            return "年は 1900〜2100 の範囲で入力してください。"

        if month_int < 1 or month_int > 12:
            return "月は 1〜12 の範囲で入力してください。"

        try:
            active_status_int = int(active_status)
        except ValueError:
            return "ステータスは 0 または 1 を指定してください。"

        if active_status_int not in (0, 1):
            return "ステータスは 0 または 1 を指定してください。"

        return None

    def _get_redirect_url(self, q_jwoa_code, q_year, q_month, q_active_status=""):
        base_url = "/active_users/"

        query_params = {}

        if q_jwoa_code:
            query_params["q_jwoa_code"] = q_jwoa_code

        if q_year:
            query_params["q_year"] = q_year

        if q_month:
            query_params["q_month"] = q_month

        if q_active_status != "":
            query_params["q_active_status"] = q_active_status

        if query_params:
            return base_url + "?" + urlencode(query_params)

        return base_url



class PlacementTreeView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "placement_tree.html"

    DEFAULT_PER_PAGE = 200
    MAX_PER_PAGE = 500
    LIST_SORT_COLUMNS = {
        "id": "c.id",
        "placement_code": "c.placement_code",
        "placement_name": "c.placement_name",
        "placement_rank": "c.placement_rank",
        "jwoa_code": "c.jwoa_code",
        "send_bv_name": "c.send_bv_name",
        "rank": "c.`rank`",
        "tree_level": "c.tree_level",
        "created_at": "c.created_at",
    }
    DOWNLINE_SORT_COLUMNS = {
        "rel_level": "d.rel_level",
        "placement_code": "d.placement_code",
        "placement_name": "d.placement_name",
        "placement_rank": "d.placement_rank",
        "jwoa_code": "d.jwoa_code",
        "send_bv_name": "d.send_bv_name",
        "rank": "d.`rank`",
        "created_at": "d.created_at",
    }

    def _get_sort_context(self, view_mode="list"):
        if view_mode == "downline":
            return get_bonus_sort_context(
                self.request,
                self.DOWNLINE_SORT_COLUMNS,
                default_sort="rel_level",
                default_direction="asc",
            )
        return get_bonus_sort_context(
            self.request,
            self.LIST_SORT_COLUMNS,
            default_sort="id",
            default_direction="asc",
        )

    def _build_where(
        self,
        q_jwoa_code: str,
        q_name: str,
        q_placement_code: str,
        q_placement_rank: str,
        q_rank: str,
    ):
        where = []
        params = []

        if q_jwoa_code:
            where.append("c.jwoa_code LIKE %s")
            params.append(f"{q_jwoa_code}%")

        if q_name:
            where.append("c.send_bv_name LIKE %s")
            params.append(f"%{q_name}%")

        if q_placement_code:
            where.append("c.placement_code LIKE %s")
            params.append(f"{q_placement_code}%")

        if q_placement_rank:
            where.append("c.placement_rank = %s")
            params.append(q_placement_rank)

        if q_rank:
            where.append("c.`rank` = %s")
            params.append(q_rank)

        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        return where_sql, params

    def _fetch_total_count(
        self,
        q_jwoa_code: str,
        q_name: str,
        q_placement_code: str,
        q_placement_rank: str,
        q_rank: str,
    ) -> int:
        where_sql, params = self._build_where(
            q_jwoa_code,
            q_name,
            q_placement_code,
            q_placement_rank,
            q_rank,
        )

        sql = f"""
SELECT COUNT(*)
FROM bonus_db.C_users_placement_tree_cache c
{where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            return int(cursor.fetchone()[0])

    def _fetch_rows(
        self,
        q_jwoa_code: str,
        q_name: str,
        q_placement_code: str,
        q_placement_rank: str,
        q_rank: str,
        limit: int,
        offset: int = 0,
        order_sql: str = "",
    ):
        where_sql, params = self._build_where(
            q_jwoa_code,
            q_name,
            q_placement_code,
            q_placement_rank,
            q_rank,
        )

        sql = f"""
SELECT
    c.id,
    c.placement_code,
    c.placement_name,
    c.placement_rank,
    c.jwoa_code,
    c.send_bv_name,
    c.`rank`,
    c.tree_level,
    c.created_at
FROM bonus_db.C_users_placement_tree_cache c
{where_sql}
ORDER BY {order_sql or "c.id ASC"}, c.id ASC
LIMIT %s OFFSET %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [col[0] for col in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]

    def _build_downline_where(
        self,
        q_name: str,
        q_placement_code: str,
        q_placement_rank: str,
        q_rank: str,
    ):
        where = []
        params = []

        if q_name:
            where.append("d.send_bv_name LIKE %s")
            params.append(f"%{q_name}%")

        if q_placement_code:
            where.append("d.placement_code LIKE %s")
            params.append(f"{q_placement_code}%")

        if q_placement_rank:
            where.append("d.placement_rank = %s")
            params.append(q_placement_rank)

        if q_rank:
            where.append("d.`rank` = %s")
            params.append(q_rank)

        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        return where_sql, params

    def _fetch_downline_total_count(
        self,
        q_jwoa_code: str,
        q_name: str,
        q_placement_code: str,
        q_placement_rank: str,
        q_rank: str,
    ) -> int:
        if not q_jwoa_code:
            return 0

        where_sql, params = self._build_downline_where(
            q_name,
            q_placement_code,
            q_placement_rank,
            q_rank,
        )
        sql = f"""
WITH RECURSIVE downline AS (
    SELECT
        c.id,
        c.placement_code,
        c.placement_name,
        c.placement_rank,
        c.jwoa_code,
        c.send_bv_name,
        c.`rank`,
        c.tree_level,
        c.created_at,
        1 AS rel_level
    FROM bonus_db.C_users_placement_tree_cache c
    WHERE c.placement_code = %s

    UNION ALL

    SELECT
        c.id,
        c.placement_code,
        c.placement_name,
        c.placement_rank,
        c.jwoa_code,
        c.send_bv_name,
        c.`rank`,
        c.tree_level,
        c.created_at,
        d.rel_level + 1 AS rel_level
    FROM bonus_db.C_users_placement_tree_cache c
    INNER JOIN downline d
        ON c.placement_code = d.jwoa_code
)
SELECT COUNT(*)
FROM downline d
{where_sql}
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [q_jwoa_code] + params)
            row = cursor.fetchone()
        return int(row[0]) if row else 0

    def _fetch_downline_rows(
        self,
        q_jwoa_code: str,
        q_name: str,
        q_placement_code: str,
        q_placement_rank: str,
        q_rank: str,
        limit: int,
        offset: int = 0,
        order_sql: str = "",
    ):
        if not q_jwoa_code:
            return []

        where_sql, params = self._build_downline_where(
            q_name,
            q_placement_code,
            q_placement_rank,
            q_rank,
        )
        sql = f"""
WITH RECURSIVE downline AS (
    SELECT
        c.id,
        c.placement_code,
        c.placement_name,
        c.placement_rank,
        c.jwoa_code,
        c.send_bv_name,
        c.`rank`,
        c.tree_level,
        c.created_at,
        1 AS rel_level
    FROM bonus_db.C_users_placement_tree_cache c
    WHERE c.placement_code = %s

    UNION ALL

    SELECT
        c.id,
        c.placement_code,
        c.placement_name,
        c.placement_rank,
        c.jwoa_code,
        c.send_bv_name,
        c.`rank`,
        c.tree_level,
        c.created_at,
        d.rel_level + 1 AS rel_level
    FROM bonus_db.C_users_placement_tree_cache c
    INNER JOIN downline d
        ON c.placement_code = d.jwoa_code
)
SELECT
    d.id,
    d.placement_code,
    d.placement_name,
    d.placement_rank,
    d.jwoa_code,
    d.send_bv_name,
    d.`rank`,
    d.tree_level,
    d.rel_level,
    d.created_at
FROM downline d
{where_sql}
ORDER BY {order_sql or "d.rel_level ASC"}, d.placement_code ASC, d.jwoa_code ASC
LIMIT %s OFFSET %s
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [q_jwoa_code] + params + [limit, offset])
            cols = [col[0] for col in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]

    def _fetch_tree_search_path(self, q_jwoa_code: str, tree_search: str):
        return fetch_tree_search_path(q_jwoa_code, tree_search)

    def _fetch_last_recreated_at(self):
        # 再作成は全削除＋一括INSERTのため、先頭行の created_at が最終再作成日時になる。
        sql = "SELECT created_at FROM bonus_db.C_users_placement_tree_cache LIMIT 1"

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql)
            row = cursor.fetchone()

        return row[0] if row else None

    def _rebuild_cache(self) -> int:
        delete_sql = "DELETE FROM bonus_db.C_users_placement_tree_cache"

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                logger.info("上位者Treeキャッシュ再作成前削除SQLを実行します。")
                cursor.execute(delete_sql)
                logger.info("上位者Treeキャッシュ再作成INSERT SQLを実行します。")
                cursor.execute(PLACEMENT_TREE_REBUILD_CACHE_SQL)
                inserted_count = cursor.rowcount

        return inserted_count

    def _delete_all_cache(self) -> int:
        sql = "DELETE FROM bonus_db.C_users_placement_tree_cache"

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.execute(sql)
                deleted_count = cursor.rowcount

        return deleted_count

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip()

        try:
            if action == "copy":
                inserted_count = self._rebuild_cache()
                record_change_audit(
                    request,
                    screen_name="上位者 Tree",
                    action_type="bulk_create",
                    target_table="C_users_placement_tree_cache",
                    target_pk=None,
                    summary=f"上位者 Treeテーブルを再作成: {inserted_count}件",
                    before_values=None,
                    after_values={"count": inserted_count},
                )
                messages.success(
                    request,
                    f"上位者 Treeテーブルを {inserted_count} 件で再作成しました。"
                )
            elif action == "delete":
                deleted_count = self._delete_all_cache()
                record_change_audit(
                    request,
                    screen_name="上位者 Tree",
                    action_type="bulk_delete",
                    target_table="C_users_placement_tree_cache",
                    target_pk=None,
                    summary=f"上位者 Treeテーブルを全件削除: {deleted_count}件",
                    before_values={"count": deleted_count},
                    after_values=None,
                )
                messages.success(
                    request,
                    "上位者 Treeテーブルを全件削除しました。"
                )
            else:
                messages.warning(request, "不正な操作です。")
        except Exception as e:
            messages.error(request, f"処理中にエラーが発生しました: {e}")

        return redirect("connect:placement_tree")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        q_jwoa_code = (self.request.GET.get("q_jwoa_code") or "").strip()
        q_name = (self.request.GET.get("q_name") or "").strip()
        q_placement_code = (self.request.GET.get("q_placement_code") or "").strip()
        q_placement_rank = (self.request.GET.get("q_placement_rank") or "").strip()
        q_rank = (self.request.GET.get("q_rank") or "").strip()
        tree_search = (self.request.GET.get("tree_search") or "").strip()
        view_mode = (self.request.GET.get("view") or "list").strip()
        if view_mode not in ("list", "tree", "downline"):
            view_mode = "list"
        downline_fullscreen = (
            view_mode == "downline"
            and self.request.GET.get("downline_fullscreen") == "1"
        )

        try:
            per_page = int(self.request.GET.get("per_page") or str(self.DEFAULT_PER_PAGE))
        except ValueError:
            per_page = self.DEFAULT_PER_PAGE
        if downline_fullscreen:
            per_page = 5000
        else:
            per_page = max(1, min(per_page, self.MAX_PER_PAGE))

        sort_ctx = self._get_sort_context(view_mode)
        ctx.update(sort_ctx)

        if view_mode == "tree":
            # 一覧件数・行は Tree タブでは使わないので取得しない。
            total_count = 0
            rows = []
            page = 1
        elif view_mode == "downline":
            total_count = self._fetch_downline_total_count(
                q_jwoa_code,
                q_name,
                q_placement_code,
                q_placement_rank,
                q_rank,
            )
            total_pages = max(1, math.ceil(total_count / per_page)) if total_count > 0 else 1
            page = self.get_page_number(total_pages)
            rows = self._fetch_downline_rows(
                q_jwoa_code=q_jwoa_code,
                q_name=q_name,
                q_placement_code=q_placement_code,
                q_placement_rank=q_placement_rank,
                q_rank=q_rank,
                limit=per_page,
                offset=(page - 1) * per_page,
                order_sql=sort_ctx["order_sql"],
            )
        else:
            total_count = self._fetch_total_count(
                q_jwoa_code,
                q_name,
                q_placement_code,
                q_placement_rank,
                q_rank,
            )
            total_pages = max(1, math.ceil(total_count / per_page)) if total_count > 0 else 1
            page = self.get_page_number(total_pages)
            rows = self._fetch_rows(
                q_jwoa_code=q_jwoa_code,
                q_name=q_name,
                q_placement_code=q_placement_code,
                q_placement_rank=q_placement_rank,
                q_rank=q_rank,
                limit=per_page,
                offset=(page - 1) * per_page,
                order_sql=sort_ctx["order_sql"],
            )

        total_pages = max(1, math.ceil(total_count / per_page)) if total_count > 0 else 1

        base_params = {}
        if q_jwoa_code:
            base_params["q_jwoa_code"] = q_jwoa_code
        if q_name:
            base_params["q_name"] = q_name
        if q_placement_code:
            base_params["q_placement_code"] = q_placement_code
        if q_placement_rank:
            base_params["q_placement_rank"] = q_placement_rank
        if q_rank:
            base_params["q_rank"] = q_rank
        if tree_search:
            base_params["tree_search"] = tree_search
        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page
        if downline_fullscreen:
            base_params["downline_fullscreen"] = "1"

        ctx["q_jwoa_code"] = q_jwoa_code
        ctx["q_name"] = q_name
        ctx["q_placement_code"] = q_placement_code
        ctx["q_placement_rank"] = q_placement_rank
        ctx["q_rank"] = q_rank
        ctx["tree_search"] = tree_search
        ctx["last_recreated_at"] = self._fetch_last_recreated_at()

        ctx["view_mode"] = view_mode
        ctx["downline_fullscreen"] = downline_fullscreen

        tab_params = dict(base_params)
        tab_params.pop("downline_fullscreen", None)
        if downline_fullscreen:
            tab_params.pop("per_page", None)
        ctx["list_tab_query"] = urlencode(tab_params) if tab_params else ""
        tree_tab_params = dict(tab_params)
        tree_tab_params["view"] = "tree"
        ctx["tree_tab_query"] = urlencode(tree_tab_params)
        downline_tab_params = dict(tab_params)
        downline_tab_params["view"] = "downline"
        ctx["downline_tab_query"] = urlencode(downline_tab_params)
        downline_fullscreen_params = dict(tab_params)
        downline_fullscreen_params["view"] = "downline"
        downline_fullscreen_params["downline_fullscreen"] = "1"
        downline_fullscreen_params["per_page"] = 5000
        ctx["downline_fullscreen_query"] = urlencode(downline_fullscreen_params)
        ctx["downline_exit_fullscreen_query"] = urlencode(downline_tab_params)

        if view_mode == "tree":
            tree_context = build_member_tree_view(q_jwoa_code)
        else:
            tree_context = {
                "tree_ancestors": [],
                "tree_focus": None,
                "tree_children": [],
                "tree_truncated": False,
                "tree_unavailable_reason": None,
                "tree_node_count": 0,
            }
        ctx.update(tree_context)
        tree_search_path_rows = (
            self._fetch_tree_search_path(q_jwoa_code, tree_search)
            if view_mode == "tree" and q_jwoa_code and tree_search
            else []
        )
        ctx["tree_search_path_rows"] = tree_search_path_rows
        ctx["tree_search_target"] = (
            next((row for row in tree_search_path_rows if row.get("is_target")), None)
            if tree_search_path_rows
            else None
        )
        ctx["tree_search_not_found"] = bool(
            view_mode == "tree" and q_jwoa_code and tree_search and not tree_search_path_rows
        )
        ctx["downline_unavailable_reason"] = (
            "会員コードを入力して検索してください。"
            if view_mode == "downline" and not q_jwoa_code
            else None
        )

        # 表示モードごとに並び替え可能な列が違うので、タブ切替リンクには引き継がない。
        add_sort_params(base_params, sort_ctx)

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )


class PlacementTreeExportView(PlacementTreeView):
    EXPORT_FETCH_SIZE = 5000
    RANK_LABELS = {
        1: "シルバー",
        2: "ゴールド",
        3: "プラチナ",
        4: "ダイヤ",
        9: "一般会員",
    }

    def _rank_label(self, value):
        return self.RANK_LABELS.get(value, "-")

    def get(self, request, *args, **kwargs):
        q_jwoa_code = (request.GET.get("q_jwoa_code") or "").strip()
        q_name = (request.GET.get("q_name") or "").strip()
        q_placement_code = (request.GET.get("q_placement_code") or "").strip()
        q_placement_rank = (request.GET.get("q_placement_rank") or "").strip()
        q_rank = (request.GET.get("q_rank") or "").strip()

        where_sql, params = self._build_where(
            q_jwoa_code=q_jwoa_code,
            q_name=q_name,
            q_placement_code=q_placement_code,
            q_placement_rank=q_placement_rank,
            q_rank=q_rank,
        )

        sql = f"""
SELECT
    c.id,
    c.placement_code,
    c.placement_name,
    c.placement_rank,
    c.jwoa_code,
    c.send_bv_name,
    c.`rank`,
    c.tree_level,
    c.created_at
FROM bonus_db.C_users_placement_tree_cache c
{where_sql}
ORDER BY {self._get_sort_context()["order_sql"]}, c.id ASC
        """

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("上位者Tree")
        ws.append([
            "ID",
            "上位者コード",
            "上位者名",
            "上位者ランク",
            "会員コード",
            "会員名",
            "ランク",
            "階層",
            "作成日時",
        ])

        with connections["rds"].cursor() as cursor:
            logger.info("上位者Tree Excel出力SQLを実行します。")
            cursor.execute(sql, params)
            while True:
                rows = cursor.fetchmany(self.EXPORT_FETCH_SIZE)
                if not rows:
                    break
                for row in rows:
                    ws.append([
                        row[0],
                        row[1],
                        row[2],
                        self._rank_label(row[3]),
                        row[4],
                        row[5],
                        self._rank_label(row[6]),
                        row[7],
                        row[8],
                    ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="placement_tree.xlsx"'
        )
        wb.save(response)
        return response


class MatchingBonusView(generic.ListView):
    template_name = "matching_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster


    def get_queryset(self):
        return PeriodMaster.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)


    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "matching_bonus")

        if action != "register_matching_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:matching_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:matching_bonus")

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:matching_bonus")

        try:
            rows = self._get_basic_bonus_rows(selected_kibetu, period)

            if not rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "matching_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/matching_bonus/?kibetu={selected_kibetu}")

            insert_sql = """
                INSERT INTO bonus_db.B_matching_bonus_result (
                    kibetu,
                    introducer_code,
                    introducer_name,
                    line_code,
                    basic_code,
                    basic_name,
                    level,
                    tree_level,
                    active_count,
                    basic_bonus,
                    matching_bonus,
                    created_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
                )
                ON DUPLICATE KEY UPDATE
                    introducer_name = VALUES(introducer_name),
                    basic_name      = VALUES(basic_name),
                    level           = VALUES(level),
                    tree_level      = VALUES(tree_level),
                    active_count    = VALUES(active_count),
                    basic_bonus     = VALUES(basic_bonus),
                    matching_bonus  = VALUES(matching_bonus),
                    created_at      = NOW()
            """

            insert_params = []
            for r in rows:
                insert_params.append([
                    selected_kibetu,
                    r.get("introducer_code") or "",
                    r.get("introducer_name") or "",
                    r.get("line_code") or "",
                    r.get("basic_code") or "",
                    r.get("basic_name") or "",
                    r.get("level") or 0,
                    r.get("tree_level") or 0,
                    r.get("active_count") or 0,
                    r.get("basic_bonus") or 0,
                    r.get("matching_bonus") or 0,
                ])

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    # マッチングボーナス登録
                    cursor.executemany(insert_sql, insert_params)

                    # 履歴登録
                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "matching_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(rows)}件登録"
                        ]
                    )
            messages.success(request, f"{len(rows)}件をマッチングボーナス結果に登録しました。")

        except Exception as e:
            logger.exception("マッチングボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/basic_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_week_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:matching_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            return ctx

        ctx["selected_period"] = period
        ctx["rows"] = self._get_basic_bonus_rows(selected_kibetu, period)
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "matching_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_basic_bonus_rows(self, selected_kibetu, period):
        st_date = period.st_date
        end_date = period.end_date

        kibetu_year = int(selected_kibetu[0:4])
        kibetu_month = int(selected_kibetu[5:7])

        start_dt = make_aware(datetime.combine(st_date, time.min))
        end_dt = make_aware(datetime.combine(end_date + timedelta(days=1), time.min))

        current_month_first = datetime(kibetu_year, kibetu_month, 1)
        prev_month_last = current_month_first - timedelta(days=1)

        prev_year = prev_month_last.year
        prev_month = prev_month_last.month

        be_start_dt = make_aware(datetime(prev_year, prev_month, 1, 0, 0, 0))
        be_end_dt = make_aware(datetime(kibetu_year, kibetu_month, 1, 0, 0, 0))


        params = [
            be_start_dt,
            be_end_dt,
            prev_year,
            prev_month,
            selected_kibetu,
            selected_kibetu,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(MATCHING_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows



class S_DriveBonusView(generic.ListView):
    template_name = "s_drive_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster

    def get_queryset(self):
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_drive_bonus_result
                ORDER BY kibetu DESC
            """)
            registered_kibetu_list = [row[0] for row in cursor.fetchall()]

        if not registered_kibetu_list:
            return PeriodMaster.objects.using("rds").none()

        return (
            PeriodMaster.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-st_date", "-kibetu")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])

            selected_kibetu_list = context.get("selected_kibetu_list", [])
            search_introducer_code = context.get("search_introducer_code", "")
            search_jwoa_code = context.get("search_jwoa_code", "")

            filename = "drive_bonus_result"
            if selected_kibetu_list:
                filename += "_" + "_".join(selected_kibetu_list)
            if search_introducer_code:
                filename += f"_intro_{search_introducer_code}"
            if search_jwoa_code:
                filename += f"_member_{search_jwoa_code}"
            filename += ".xlsx"

            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["drive_bonus"],
                "DriveBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu_list = self.request.GET.getlist("kibetu")
        search_introducer_code = self.request.GET.get("introducer_code", "").strip()
        search_jwoa_code = self.request.GET.get("jwoa_code", "").strip()

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "title_name": "title_name",
                "introducer_code": "introducer_code",
                "jwoa_code": "jwoa_code",
                "jwoa_name": "jwoa_name",
                "sum_bv": "sum_bv",
                "rate": "rate",
                "sum_bonus_amount": "sum_bonus_amount",
            },
            default_sort="kibetu",
        )
        ctx.update(sort_ctx)

        ctx["selected_kibetu_list"] = selected_kibetu_list
        ctx["search_introducer_code"] = search_introducer_code
        ctx["search_jwoa_code"] = search_jwoa_code
        ctx["rows"] = []

        if not selected_kibetu_list and not search_introducer_code and not search_jwoa_code:
            return ctx

        sql = """
            SELECT
                id,
                kibetu,
                title_name,
                introducer_code,
                jwoa_code,
                jwoa_name,
                sum_bv,
                rate,
                sum_bonus_amount,
                created_at
            FROM bonus_db.B_drive_bonus_result
            WHERE 1 = 1
        """

        params = []

        if selected_kibetu_list:
            placeholders = ", ".join(["%s"] * len(selected_kibetu_list))
            sql += f"""
                AND kibetu IN ({placeholders})
            """
            params.extend(selected_kibetu_list)

        if search_introducer_code:
            sql += """
                AND introducer_code LIKE %s
            """
            params.append(f"%{search_introducer_code}%")

        if search_jwoa_code:
            sql += """
                AND jwoa_code LIKE %s
            """
            params.append(f"%{search_jwoa_code}%")

        sql += f"""
            ORDER BY {sort_ctx['order_sql']}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx


class S_BasicBonusView(generic.ListView):
    template_name = "s_basic_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster

    def get_queryset(self):
        # B_basic_bonus_result に登録済みの期別だけ取得
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_basic_bonus_result
                ORDER BY kibetu DESC
            """)
            registered_kibetu_list = [row[0] for row in cursor.fetchall()]

        if not registered_kibetu_list:
            return PeriodMaster.objects.using("rds").none()

        return (
            PeriodMaster.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-st_date", "-kibetu")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            selected_kibetu_list = context.get("selected_kibetu_list", [])
            kibetu = "_".join(selected_kibetu_list) if selected_kibetu_list else ""
            filename = build_bonus_export_filename("basic_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["basic_bonus"],
                "BasicBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu_list = self.request.GET.getlist("kibetu")
        ctx["selected_kibetu_list"] = selected_kibetu_list
        ctx["rows"] = []

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "placement_code": "placement_code",
                "placement_name": "placement_name",
                "placement_rank": "placement_rank",
                "line_code": "line_code",
                "purchaser_code": "purchaser_code",
                "purchaser_name": "purchaser_name",
                "sum_bv": "sum_bv",
                "bonus_rate": "bonus_rate",
                "bonus_amount": "bonus_amount",
                "blue_daiya_flg": "blue_daiya_flg",
                "created_at": "created_at",
            },
            default_sort="placement_code",
        )
        ctx.update(sort_ctx)

        placement_code = self.request.GET.get("placement_code", "").strip()
        purchaser_code = self.request.GET.get("purchaser_code", "").strip()
        purchaser_name = self.request.GET.get("purchaser_name", "").strip()
        line_code = self.request.GET.get("line_code", "").strip()

        if (
            not selected_kibetu_list
            and not placement_code
            and not purchaser_code
            and not purchaser_name
            and not line_code
        ):
            ctx.update({
                "placement_code": placement_code,
                "purchaser_code": purchaser_code,
                "purchaser_name": purchaser_name,
                "line_code": line_code,
            })
            return ctx

        sql = """
            SELECT
                id,
                kibetu,
                placement_code,
                placement_name,
                placement_rank,
                line_code,
                purchaser_code,
                purchaser_name,
                sum_bv,
                bonus_rate,
                bonus_amount,
                blue_daiya_flg,
                created_at
            FROM bonus_db.B_basic_bonus_result
            WHERE 1 = 1
        """

        params = []

        if selected_kibetu_list:
            placeholders = ", ".join(["%s"] * len(selected_kibetu_list))
            sql += f"""
                AND kibetu IN ({placeholders})
            """
            params.extend(selected_kibetu_list)

        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "placement_code": "placement_code",
                "purchaser_code": "purchaser_code",
                "purchaser_name": "purchaser_name",
                "line_code": "line_code",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx




class S_MatchingBonusView(generic.ListView):
    template_name = "s_matching_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster

    def get_queryset(self):
        # B_drive_bonus_result に登録済みの期別だけ取得
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_matching_bonus_result
                ORDER BY kibetu DESC
            """)
            registered_kibetu_list = [row[0] for row in cursor.fetchall()]

        if not registered_kibetu_list:
            return PeriodMaster.objects.using("rds").none()

        return (
            PeriodMaster.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-st_date", "-kibetu")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            selected_kibetu_list = context.get("selected_kibetu_list", [])
            kibetu = "_".join(selected_kibetu_list) if selected_kibetu_list else ""
            filename = build_bonus_export_filename("matching_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["matching_bonus"],
                "MatchingBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu_list = self.request.GET.getlist("kibetu")
        ctx["selected_kibetu_list"] = selected_kibetu_list
        ctx["rows"] = []

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "introducer_code": "introducer_code",
                "introducer_name": "introducer_name",
                "line_code": "line_code",
                "basic_code": "basic_code",
                "basic_name": "basic_name",
                "level": "level",
                "tree_level": "tree_level",
                "active_count": "active_count",
                "basic_bonus": "basic_bonus",
                "matching_bonus": "matching_bonus",
                "created_at": "created_at",
            },
            default_sort="introducer_code",
        )
        ctx.update(sort_ctx)

        introducer_code = self.request.GET.get("introducer_code", "").strip()
        introducer_name = self.request.GET.get("introducer_name", "").strip()
        matching_level = self.request.GET.get("matching_level", "").strip()

        if not selected_kibetu_list and not introducer_code and not introducer_name and not matching_level:
            ctx.update({
                "introducer_code": introducer_code,
                "introducer_name": introducer_name,
                "matching_level": matching_level,
            })
            return ctx

        sql = """
            SELECT
                id,
                kibetu,
                introducer_code,
                introducer_name,
                line_code,
                basic_code,
                basic_name,
                level,
                tree_level,
                active_count,
                basic_bonus,
                matching_bonus,
                created_at
            FROM bonus_db.B_matching_bonus_result
            WHERE 1 = 1
        """

        params = []

        if selected_kibetu_list:
            placeholders = ", ".join(["%s"] * len(selected_kibetu_list))
            sql += f"""
                AND kibetu IN ({placeholders})
            """
            params.extend(selected_kibetu_list)

        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "introducer_code": "introducer_code",
                "introducer_name": "introducer_name",
            },
        )
        ctx.update(filter_values)

        if matching_level:
            sql += """
                AND level = %s
            """
            params.append(matching_level)
        ctx["matching_level"] = matching_level

        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx


class TitleBonusView(generic.ListView):
    template_name = "title_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        return MonthlyPeriod.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):

        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "title_bonus")

        if action != "register_title_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:title_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:title_bonus")

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:title_bonus")

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(request, "登録")
            return redirect(f"/title_bonus/?kibetu={selected_kibetu}")

        try:
            title_bonus_rows = self._get_title_bonus_rows(
                selected_kibetu,
                period
            )

            if not title_bonus_rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "title_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/title_bonus/?kibetu={selected_kibetu}")

            insert_sql, insert_params = (
                register_sql.get_title_bonus_insert_data(
                    selected_kibetu,
                    title_bonus_rows
                )
            )

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.executemany(insert_sql, insert_params)

                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "title_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(title_bonus_rows)}件登録"
                        ]
                    )

            messages.success(
                request,
                f"{len(title_bonus_rows)}件をタイトルボーナス結果に登録しました。"
            )

        except Exception as e:
            logger.exception("タイトルボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/title_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:title_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period
        if not ensure_kibetu_purchase_info(self.request, selected_kibetu, period):
            insert_empty_bonus_history_on_display(
                self.request,
                "title_bonus",
                selected_kibetu,
            )
            return ctx

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(self.request)
            return ctx

        ctx["rows"] = self._get_title_bonus_rows(
            selected_kibetu,
            period
        )
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "title_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_title_bonus_rows(self, selected_kibetu, period):

        kibetu_year = period.year
        kibetu_month = period.month

        current_month_first = datetime(
            kibetu_year,
            kibetu_month,
            1
        )

        prev_month_last = current_month_first - timedelta(days=1)

        prev_year = prev_month_last.year
        prev_month = prev_month_last.month

        params = [
            kibetu_year,
            kibetu_month,
            selected_kibetu,
            kibetu_year,
            kibetu_month,
            prev_year,
            prev_month,
            kibetu_year,
            kibetu_month,
            prev_year,
            prev_month,
            kibetu_year,
            kibetu_month,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(TITLE_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows


class S_TitleBonusView(generic.ListView):
    template_name = "s_title_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):

        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_title_bonus_result
                ORDER BY kibetu DESC
            """)

            registered_kibetu_list = [
                row[0]
                for row in cursor.fetchall()
            ]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):

        self.object_list = self.get_queryset()

        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            filename = build_bonus_export_filename("title_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["title_bonus"],
                "TitleBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "root_jwoa_code": "root_jwoa_code",
                "root_name": "root_name",
                "up_jwoa_code": "up_jwoa_code",
                "down_jwoa_code": "down_jwoa_code",
                "down_name": "down_name",
                "tree_level": "tree_level",
                "match_level": "match_level",
                "title_name": "title_name",
                "sum_bv": "sum_bv",
                "rate": "rate",
                "bonus_amount": "bonus_amount",
                "created_at": "created_at",
            },
            default_sort="root_jwoa_code",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                tbr.id,
                tbr.kibetu,
                tbr.root_jwoa_code,
                tbr.root_name,
                tbr.up_jwoa_code,
                tbr.down_jwoa_code,
                tbr.down_name,
                tbr.tree_level,
                tbr.match_level,
                COALESCE(tm.title_name, 'タイトルなし') AS title_name,
                tbr.sum_bv,
                tbr.rate,
                tbr.bonus_amount,
                tbr.created_at
            FROM bonus_db.B_title_bonus_result AS tbr
            LEFT JOIN bonus_db.title_master AS tm
              ON tbr.title_id = tm.title_id
            WHERE tbr.kibetu = %s
        """

        params = [selected_kibetu]
        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "root_jwoa_code": "tbr.root_jwoa_code",
                "up_jwoa_code": "tbr.up_jwoa_code",
                "down_jwoa_code": "tbr.down_jwoa_code",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:

            cursor.execute(sql, params)

            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]

            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        ctx["rows"] = rows

        return ctx


class TitleDiffBonusView(generic.ListView):
    template_name = "title_diff_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        return MonthlyPeriod.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "title_diff_bonus")

        if action != "register_title_diff_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:title_diff_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:title_diff_bonus")

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:title_diff_bonus")

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(request, "登録")
            return redirect(f"/title_diff_bonus/?kibetu={selected_kibetu}")

        try:
            title_diff_bonus_rows = self._get_title_diff_bonus_rows(
                selected_kibetu,
                period
            )

            if not title_diff_bonus_rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "title_diff_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/title_diff_bonus/?kibetu={selected_kibetu}")

            insert_sql, insert_params = (
                register_sql.get_title_diff_bonus_insert_data(
                    selected_kibetu,
                    title_diff_bonus_rows
                )
            )

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.executemany(insert_sql, insert_params)

                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "title_diff_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(title_diff_bonus_rows)}件登録"
                        ]
                    )

            messages.success(
                request,
                f"{len(title_diff_bonus_rows)}件をタイトル差額ボーナス結果に登録しました。"
            )

        except Exception as e:
            logger.exception("タイトル差額ボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/title_diff_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:title_diff_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period
        if not ensure_kibetu_purchase_info(self.request, selected_kibetu, period):
            insert_empty_bonus_history_on_display(
                self.request,
                "title_diff_bonus",
                selected_kibetu,
            )
            return ctx

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(self.request)
            return ctx

        ctx["rows"] = self._get_title_diff_bonus_rows(
            selected_kibetu,
            period
        )
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "title_diff_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_title_diff_bonus_rows(self, selected_kibetu, period):

        kibetu_year = period.year
        kibetu_month = period.month

        params = [
            kibetu_month,
            kibetu_year,
            selected_kibetu,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(TITLE_DIFF_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows


class S_TitleDiffBonusView(generic.ListView):
    template_name = "s_title_diff_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):

        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_title_diff_bonus_result
                ORDER BY kibetu DESC
            """)

            registered_kibetu_list = [
                row[0]
                for row in cursor.fetchall()
            ]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            filename = build_bonus_export_filename("title_diff_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["title_diff_bonus"],
                "TitleDiffBonusResult",
                filename,
            )

        return self.render_to_response(context)


    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "tdbr.kibetu",
                "root_title_name": "root_title_name",
                "root_bonus_rate": "tdbr.root_bonus_rate",
                "root_jwoa_code": "tdbr.root_jwoa_code",
                "root_name": "tdbr.root_name",
                "down_title_name": "down_title_name",
                "down_bonus_rate": "tdbr.down_bonus_rate",
                "down_jwoa_code": "tdbr.down_jwoa_code",
                "down_name": "tdbr.down_name",
                "pay_bonus_rate": "tdbr.pay_bonus_rate",
                "sum_bv": "tdbr.sum_bv",
                "title_diff_bonus": "tdbr.title_diff_bonus",
                "created_at": "tdbr.created_at",
                "updated_at": "tdbr.updated_at",
            },
            default_sort="root_jwoa_code",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                tdbr.kibetu,
                tdbr.root_title_id,
                COALESCE(root_tm.title_name, 'タイトルなし') AS root_title_name,
                tdbr.root_bonus_rate,
                tdbr.root_jwoa_code,
                tdbr.root_name,
                tdbr.down_title_id,
                COALESCE(down_tm.title_name, 'タイトルなし') AS down_title_name,
                tdbr.down_bonus_rate,
                tdbr.down_jwoa_code,
                tdbr.down_name,
                tdbr.pay_bonus_rate,
                tdbr.sum_bv,
                tdbr.title_diff_bonus,
                tdbr.created_at,
                tdbr.updated_at
            FROM bonus_db.B_title_diff_bonus_result AS tdbr
            LEFT JOIN bonus_db.title_master AS root_tm
              ON tdbr.root_title_id = root_tm.title_id
            LEFT JOIN bonus_db.title_master AS down_tm
              ON tdbr.down_title_id = down_tm.title_id
            WHERE tdbr.kibetu = %s
        """

        params = [selected_kibetu]
        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "root_jwoa_code": "tdbr.root_jwoa_code",
                "down_jwoa_code": "tdbr.down_jwoa_code",
                "root_name": "tdbr.root_name",
                "down_name": "tdbr.down_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx


class RepurchaseOverBonusView(generic.ListView):
    template_name = "repurchase_over_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        return (
            MonthlyPeriod.objects.using("rds")
            .all()
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "repurchase_over_bonus")

        if action != "repurchase_over_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:repurchase_over_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:repurchase_over_bonus")

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:repurchase_over_bonus")

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(request, "登録")
            return redirect(f"/repurchase_over_bonus/?kibetu={selected_kibetu}")

        try:
            repurchase_over_bonus_rows = self._get_repurchase_over_bonus_rows(
                selected_kibetu=selected_kibetu,
                period=period,
            )

            if not repurchase_over_bonus_rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "repurchase_over_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/repurchase_over_bonus/?kibetu={selected_kibetu}")

            insert_sql, insert_params = (
                register_sql.get_repurchase_over_bonus_insert_data(
                    selected_kibetu,
                    repurchase_over_bonus_rows,
                )
            )

            if not insert_params:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "repurchase_over_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（登録対象なし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/repurchase_over_bonus/?kibetu={selected_kibetu}")

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.executemany(insert_sql, insert_params)

                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "repurchase_over_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(insert_params)}件登録",
                        ],
                    )

            messages.success(
                request,
                f"{len(insert_params)}件を再購入オーバーボーナス結果に登録しました。"
            )

        except Exception as e:
            logger.exception("再購入オーバーボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/repurchase_over_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = self.request.GET.get("kibetu", "").strip()
        root_code = self.request.GET.get("root_code", "").strip()
        down_code = self.request.GET.get("down_code", "").strip()

        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:repurchase_over_bonus"
        ctx["root_code"] = root_code
        ctx["down_code"] = down_code
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period
        if not ensure_kibetu_purchase_info(self.request, selected_kibetu, period):
            insert_empty_bonus_history_on_display(
                self.request,
                "repurchase_over_bonus",
                selected_kibetu,
            )
            return ctx

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(self.request)
            return ctx

        ctx["rows"] = self._get_repurchase_over_bonus_rows(
            selected_kibetu=selected_kibetu,
            period=period,
            root_code=root_code,
            down_code=down_code,
        )
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "repurchase_over_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_repurchase_over_bonus_rows(
        self,
        selected_kibetu,
        period,
        root_code="",
        down_code="",
    ):
        kibetu_year = period.year
        kibetu_month = period.month

        params = [
            kibetu_year,
            kibetu_month,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(REPURCHASE_OVER_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        if root_code:
            root_code = root_code.upper()
            rows = [
                r for r in rows
                if root_code in str(r.get("root_code", "")).upper()
            ]

        if down_code:
            down_code = down_code.upper()
            rows = [
                r for r in rows
                if down_code in str(r.get("down_code", "")).upper()
            ]

        return rows


class S_RepurchaseOverBonusView(generic.ListView):
    template_name = "s_repurchase_over_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):

        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_repurchase_over_bonus_result
                ORDER BY kibetu DESC
            """)

            registered_kibetu_list = [
                row[0]
                for row in cursor.fetchall()
            ]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            filename = build_bonus_export_filename("repurchase_over_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["repurchase_over_bonus"],
                "RepurchaseOverBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        selected_kibetu = self.request.GET.get("kibetu", "").strip()

        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "root_code": "root_code",
                "root_name": "root_name",
                "down_code": "down_code",
                "down_name": "down_name",
                "tree_level": "tree_level",
                "match_count": "match_count",
                "rate": "rate",
                "sum_bv": "sum_bv",
                "over_bonus": "over_bonus",
                "created_at": "created_at",
                "updated_at": "updated_at",
            },
            default_sort="root_code",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                *
            FROM bonus_db.B_repurchase_over_bonus_result
            WHERE kibetu = %s
        """

        params = [selected_kibetu]
        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "root_code": "root_code",
                "down_code": "down_code",
                "root_name": "root_name",
                "down_name": "down_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        ctx["rows"] = rows

        return ctx



class BusinessPersonalPerformanceView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "business_personal_performance.html"
    period_label = "月別"
    active_menu = "business_personal_performance"
    reset_url_name = "connect:business_personal_performance"

    def _build_where(self, q_jmoa_code="", q_name=""):
        where = []
        params = []

        if q_jmoa_code:
            where.append("u.jmoa_code LIKE %s")
            params.append(f"%{q_jmoa_code}%")

        if q_name:
            where.append("(u.send_bv_name LIKE %s OR u.name LIKE %s)")
            params.append(f"%{q_name}%")
            params.append(f"%{q_name}%")

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return where_sql, params

    def _fetch_total_count(self, q_jmoa_code="", q_name=""):
        where_sql, params = self._build_where(q_jmoa_code=q_jmoa_code, q_name=q_name)
        sql = f"""
            SELECT COUNT(*)
            FROM nexus_production.users u
            {where_sql}
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
        return int(row[0]) if row else 0

    def _fetch_rows(self, q_jmoa_code="", q_name="", limit=200, offset=0):
        where_sql, params = self._build_where(q_jmoa_code=q_jmoa_code, q_name=q_name)
        sql = f"""
            SELECT
                u.jmoa_code,
                u.send_bv_name,
                u.rank,
                u.status_code,
                COUNT(o.id) AS order_count,
                COALESCE(SUM(o.total_bv), 0) AS total_bv,
                COALESCE(SUM(o.total_price), 0) AS total_price,
                MAX(o.order_at) AS last_order_at
            FROM nexus_production.users u
            LEFT JOIN nexus_production.orders o
                ON o.jwoa_code = u.jmoa_code
            {where_sql}
            GROUP BY
                u.jmoa_code,
                u.send_bv_name,
                u.rank,
                u.status_code
            ORDER BY u.jmoa_code
            LIMIT %s OFFSET %s
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q_jmoa_code = (self.request.GET.get("q_jmoa_code") or "").strip()
        q_name = (self.request.GET.get("q_name") or "").strip()

        per_page = self.get_per_page()
        total_count = self._fetch_total_count(q_jmoa_code=q_jmoa_code, q_name=q_name)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            q_jmoa_code=q_jmoa_code,
            q_name=q_name,
            limit=per_page,
            offset=offset,
        )

        ctx["q_jmoa_code"] = q_jmoa_code
        ctx["q_name"] = q_name
        ctx["period_label"] = self.period_label
        ctx["active_menu"] = self.active_menu
        ctx["reset_url_name"] = self.reset_url_name
        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params={
                "q_jmoa_code": q_jmoa_code,
                "q_name": q_name,
                "per_page": per_page,
            },
        )


class BusinessPersonalMonthPerformanceView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "business_personal_month_performance.html"

    def _build_where(self, q_kibetu_list=(), q_jwoa_code=""):
        where = []
        params = []

        kibetu_sql, kibetu_params = build_kibetu_condition(q_kibetu_list)
        if kibetu_sql:
            where.append(kibetu_sql)
            params.extend(kibetu_params)

        if q_jwoa_code:
            where.append("jwoa_code LIKE %s")
            params.append(f"%{q_jwoa_code}%")

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return where_sql, params

    def _fetch_total_count(self, q_kibetu_list=(), q_jwoa_code=""):
        where_sql, params = self._build_where(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
        )
        sql = f"""
            SELECT COUNT(*)
            FROM bonus_db.B_month_bonus_result
            {where_sql}
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
        return int(row[0]) if row else 0

    def _fetch_rows(self, q_kibetu_list=(), q_jwoa_code="", limit=200, offset=0):
        where_sql, params = self._build_where(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
        )
        sql = f"""
            SELECT
                id,
                kibetu,
                jwoa_code,
                jwoa_name,
                title_bonus,
                repurchase_over_bonus,
                title_diff_bonus,
                three_star_diamond_global_bonus,
                crown_three_star_diamond_global_bonus,
                month_bonus,
                created_at,
                updated_at
            FROM bonus_db.B_month_bonus_result
            {where_sql}
            ORDER BY kibetu DESC, jwoa_code
            LIMIT %s OFFSET %s
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q_kibetu_list = parse_kibetu_list(self.request.GET.get("q_kibetu"))
        q_kibetu = join_kibetu_list(q_kibetu_list)
        q_jwoa_code = (self.request.GET.get("q_jwoa_code") or "").strip()

        per_page = self.get_per_page()
        total_count = self._fetch_total_count(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
        )
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
            limit=per_page,
            offset=offset,
        )

        ctx["q_kibetu"] = q_kibetu
        ctx["q_kibetu_list"] = q_kibetu_list
        ctx["q_jwoa_code"] = q_jwoa_code
        ctx["active_menu"] = "business_personal_performance"
        ctx["registration_history_rows"] = fetch_registration_history_rows(
            MONTH_PERSONAL_RESULT_TABLE,
            kibetu_not_like="%W%",
        )
        ctx["registration_history_modal_id"] = "monthPersonalRegistrationModal"
        ctx["registration_modal_title"] = "登録履歴（月別 個人業績）"
        ctx["registration_target_url_name"] = "connect:business_personal_performance"
        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params={
                "q_kibetu": q_kibetu,
                "q_jwoa_code": q_jwoa_code,
                "per_page": per_page,
            },
        )


class BusinessPersonalWeekPerformanceView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "business_personal_week_performance.html"

    # 会員コードは前方一致なら idx_jwoa_code が使える。中間一致は54万件の全件走査に
    # なるため、まず前方一致で引き、0件のときだけ中間一致に切り替える。
    # コードは「JP/HK + 数字」の形式なので、前方一致で見つかる限り中間一致との
    # 結果差は生じない。数字部分だけを入力した場合は前方一致が0件になり、
    # 従来どおり中間一致で探せる。
    CODE_MATCH_PREFIX = "prefix"
    CODE_MATCH_CONTAINS = "contains"

    def _build_where(self, q_kibetu_list=(), q_jwoa_code="", code_match=CODE_MATCH_PREFIX):
        where = []
        params = []

        kibetu_sql, kibetu_params = build_kibetu_condition(q_kibetu_list)
        if kibetu_sql:
            where.append(kibetu_sql)
            params.extend(kibetu_params)

        if q_jwoa_code:
            where.append("jwoa_code LIKE %s")
            if code_match == self.CODE_MATCH_CONTAINS:
                params.append(f"%{q_jwoa_code}%")
            else:
                params.append(f"{q_jwoa_code}%")

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return where_sql, params

    def _resolve_code_match(self, q_kibetu_list=(), q_jwoa_code=""):
        """会員コードの照合方法と、その条件での件数を返す。"""
        count = self._fetch_total_count(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
            code_match=self.CODE_MATCH_PREFIX,
        )
        if not q_jwoa_code or count:
            return self.CODE_MATCH_PREFIX, count

        return self.CODE_MATCH_CONTAINS, self._fetch_total_count(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
            code_match=self.CODE_MATCH_CONTAINS,
        )

    def _fetch_total_count(self, q_kibetu_list=(), q_jwoa_code="", code_match=CODE_MATCH_PREFIX):
        where_sql, params = self._build_where(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
            code_match=code_match,
        )
        sql = f"""
            SELECT COUNT(*)
            FROM bonus_db.B_week_bonus_result
            {where_sql}
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
        return int(row[0]) if row else 0

    def _fetch_rows(self, q_kibetu_list=(), q_jwoa_code="", limit=200, offset=0,
                    code_match=CODE_MATCH_PREFIX):
        where_sql, params = self._build_where(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
            code_match=code_match,
        )
        sql = f"""
            SELECT
                id,
                kibetu,
                jwoa_code,
                jwoa_name,
                drive_bonus,
                basic_bonus,
                matching_bonus,
                week_bonus,
                created_at,
                updated_at
            FROM bonus_db.B_week_bonus_result
            {where_sql}
            ORDER BY kibetu DESC, jwoa_code
            LIMIT %s OFFSET %s
        """
        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q_kibetu_list = parse_kibetu_list(self.request.GET.get("q_kibetu"))
        q_jwoa_code = (self.request.GET.get("q_jwoa_code") or "").strip()
        kibetu_choice_mode = self.request.GET.get("kibetu_choice_mode") or "recent"

        q_kibetu_list, q_kibetu_defaulted = resolve_default_kibetu_list(
            WEEK_PERSONAL_RESULT_TABLE,
            kibetu_list=q_kibetu_list,
            kibetu_like="%W%",
            other_filters=(q_jwoa_code,),
        )
        q_kibetu = join_kibetu_list(q_kibetu_list)

        per_page = self.get_per_page()
        code_match, total_count = self._resolve_code_match(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
        )
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            q_kibetu_list=q_kibetu_list,
            q_jwoa_code=q_jwoa_code,
            limit=per_page,
            offset=offset,
            code_match=code_match,
        )

        ctx["q_kibetu"] = q_kibetu
        ctx["q_kibetu_list"] = q_kibetu_list
        ctx["q_kibetu_defaulted"] = q_kibetu_defaulted
        ctx["q_jwoa_code"] = q_jwoa_code
        ctx["active_menu"] = "business_personal_week_performance"
        ctx["registration_history_rows"] = fetch_registration_history_rows(
            WEEK_PERSONAL_RESULT_TABLE,
            kibetu_like="%W%",
        )
        ctx["registration_history_modal_id"] = "weekPersonalRegistrationModal"
        ctx["registration_modal_title"] = "登録履歴（週別 個人業績）"
        ctx["registration_target_url_name"] = "connect:business_personal_week_performance"
        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params={
                "q_kibetu": q_kibetu,
                "q_jwoa_code": q_jwoa_code,
                "per_page": per_page,
                "kibetu_choice_mode": kibetu_choice_mode,
            },
        )


_FULLWIDTH_KANA_TO_HALFWIDTH = {
    "ア": "ｱ", "イ": "ｲ", "ウ": "ｳ", "エ": "ｴ", "オ": "ｵ",
    "カ": "ｶ", "キ": "ｷ", "ク": "ｸ", "ケ": "ｹ", "コ": "ｺ",
    "サ": "ｻ", "シ": "ｼ", "ス": "ｽ", "セ": "ｾ", "ソ": "ｿ",
    "タ": "ﾀ", "チ": "ﾁ", "ツ": "ﾂ", "テ": "ﾃ", "ト": "ﾄ",
    "ナ": "ﾅ", "ニ": "ﾆ", "ヌ": "ﾇ", "ネ": "ﾈ", "ノ": "ﾉ",
    "ハ": "ﾊ", "ヒ": "ﾋ", "フ": "ﾌ", "ヘ": "ﾍ", "ホ": "ﾎ",
    "マ": "ﾏ", "ミ": "ﾐ", "ム": "ﾑ", "メ": "ﾒ", "モ": "ﾓ",
    "ヤ": "ﾔ", "ユ": "ﾕ", "ヨ": "ﾖ",
    "ラ": "ﾗ", "リ": "ﾘ", "ル": "ﾙ", "レ": "ﾚ", "ロ": "ﾛ",
    "ワ": "ﾜ", "ヲ": "ｦ", "ン": "ﾝ",
    "ァ": "ｧ", "ィ": "ｨ", "ゥ": "ｩ", "ェ": "ｪ", "ォ": "ｫ",
    "ャ": "ｬ", "ュ": "ｭ", "ョ": "ｮ", "ッ": "ｯ",
    "ー": "ｰ", "・": "･",
}

_FULLWIDTH_VOICED_KANA_TO_HALFWIDTH = {
    "ガ": "ｶﾞ", "ギ": "ｷﾞ", "グ": "ｸﾞ", "ゲ": "ｹﾞ", "ゴ": "ｺﾞ",
    "ザ": "ｻﾞ", "ジ": "ｼﾞ", "ズ": "ｽﾞ", "ゼ": "ｾﾞ", "ゾ": "ｿﾞ",
    "ダ": "ﾀﾞ", "ヂ": "ﾁﾞ", "ヅ": "ﾂﾞ", "デ": "ﾃﾞ", "ド": "ﾄﾞ",
    "バ": "ﾊﾞ", "ビ": "ﾋﾞ", "ブ": "ﾌﾞ", "ベ": "ﾍﾞ", "ボ": "ﾎﾞ",
    "ヴ": "ｳﾞ",
}

_FULLWIDTH_SEMIVOICED_KANA_TO_HALFWIDTH = {
    "パ": "ﾊﾟ", "ピ": "ﾋﾟ", "プ": "ﾌﾟ", "ペ": "ﾍﾟ", "ポ": "ﾎﾟ",
}


def _fullwidth_kana_to_halfwidth(value):
    converted = []
    for char in value:
        if char in _FULLWIDTH_VOICED_KANA_TO_HALFWIDTH:
            converted.append(_FULLWIDTH_VOICED_KANA_TO_HALFWIDTH[char])
        elif char in _FULLWIDTH_SEMIVOICED_KANA_TO_HALFWIDTH:
            converted.append(_FULLWIDTH_SEMIVOICED_KANA_TO_HALFWIDTH[char])
        else:
            converted.append(_FULLWIDTH_KANA_TO_HALFWIDTH.get(char, char))
    return "".join(converted)


def _kana_search_variants(value):
    raw = (value or "").strip()
    if not raw:
        return []

    fullwidth = unicodedata.normalize("NFKC", raw)
    candidates = [
        raw,
        fullwidth,
        _fullwidth_kana_to_halfwidth(raw),
        _fullwidth_kana_to_halfwidth(fullwidth),
    ]
    variants = []
    for candidate in candidates:
        if candidate and candidate not in variants:
            variants.append(candidate)
    return variants


class UsersView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "users.html"
    EXPORT_FETCH_SIZE = 5000
    SORT_COLUMNS = {
        "group_code": "u.group_code",
        "jmoa_code": "u.jmoa_code",
        "send_bv_name": "u.send_bv_name",
        "name_kana": "u.name_kana",
        "introducer_code": "u.introducer_code",
        "placement_code": "u.placement_code",
        "prev_month_active_status": "prev_month_active_status",
        "rank": "u.`rank`",
        "status_code": "u.status_code",
        "activated": "u.activated",
        "company": "u.company",
        "interim_at": "u.interim_at",
        "activated_at": "u.activated_at",
        "last_purchase_at": "u.last_purchase_at",
        "created_at": "u.created_at",
        "updated_at": "u.updated_at",
    }
    DEFAULT_SORT = "created_at"
    DEFAULT_SORT_DIRECTION = "desc"

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort=self.DEFAULT_SORT,
            default_direction=self.DEFAULT_SORT_DIRECTION,
        )

    def _build_where(
        self,
        q_jpid: str = "",
        q_name: str = "",
        q_name_kana: str = "",
        q_introducer: str = "",
        q_placement: str = "",
        q_status: str = "",
        q_rank: str = "",
        q_interim_from: str = "",
        q_interim_to: str = "",
        q_activated_from: str = "",
        q_activated_to: str = "",
    ):
        where = []
        params = []

        if q_jpid:
            where.append("u.jmoa_code LIKE %s")
            params.append(f"{q_jpid}%")

        if q_name:
            where.append("(u.send_bv_name LIKE %s OR u.name LIKE %s)")
            params.append(f"%{q_name}%")
            params.append(f"%{q_name}%")

        if q_name_kana:
            kana_variants = _kana_search_variants(q_name_kana)
            if kana_variants:
                where.append(
                    "("
                    + " OR ".join(["u.name_kana LIKE %s"] * len(kana_variants))
                    + ")"
                )
                params.extend([f"%{variant}%" for variant in kana_variants])

        if q_introducer:
            where.append("u.introducer_code LIKE %s")
            params.append(f"{q_introducer}%")

        if q_placement:
            where.append("u.placement_code LIKE %s")
            params.append(f"{q_placement}%")

        if q_status:
            where.append("u.status_code = %s")
            params.append(q_status)

        if q_rank:
            where.append("u.rank = %s")
            params.append(q_rank)

        if q_interim_from:
            where.append("u.interim_at >= %s")
            params.append(q_interim_from)

        if q_interim_to:
            where.append("u.interim_at < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_interim_to)

        if q_activated_from:
            where.append("u.activated_at >= %s")
            params.append(q_activated_from)

        if q_activated_to:
            where.append("u.activated_at < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_activated_to)

        where_sql = "WHERE " + " AND ".join(where) if where else ""

        return where_sql, params

    @staticmethod
    def _has_active_period(active_year, active_month, active_start_date, active_end_date):
        return bool(active_year and active_month and active_start_date and active_end_date)

    @staticmethod
    def _apply_active_result_filter(where_sql, q_active_result):
        if q_active_result == "active":
            condition = "(IFNULL(au.active_status, 0) = 1 OR IFNULL(pb.prev_month_bv, 0) >= 50)"
        elif q_active_result == "inactive":
            condition = "(IFNULL(au.active_status, 0) <> 1 AND IFNULL(pb.prev_month_bv, 0) < 50)"
        else:
            return where_sql

        if where_sql:
            return f"{where_sql} AND {condition}"
        return f"WHERE {condition}"

    @staticmethod
    def _build_active_join_sql():
        return """
            LEFT JOIN bonus_db.active_users au
              ON au.jwoa_code = u.jmoa_code
             AND au.year = %s
             AND au.month = %s
            LEFT JOIN (
                SELECT
                    jwoa_code,
                    SUM(IFNULL(bv, 0)) AS prev_month_bv
                FROM bonus_db.purchase_info_list
                WHERE bonus_payment_date >= %s
                  AND bonus_payment_date < %s
                GROUP BY jwoa_code
            ) pb
              ON pb.jwoa_code = u.jmoa_code
            """

    @staticmethod
    def _needs_inline_active_join(q_active_result="", order_sql="", force_active_join=False):
        # 判定結果で絞り込む／アクティブ確認列で並べ替える場合だけ、
        # 全会員に購入集計をJOINする。一覧表示は LIMIT 後に200件だけ判定する。
        if force_active_join:
            return True
        if q_active_result in ("active", "inactive"):
            return True
        return "prev_month_active_status" in (order_sql or "")

    def _fetch_total_count(
        self,
        q_jpid: str = "",
        q_name: str = "",
        q_name_kana: str = "",
        q_introducer: str = "",
        q_placement: str = "",
        q_status: str = "",
        q_rank: str = "",
        q_interim_from: str = "",
        q_interim_to: str = "",
        q_activated_from: str = "",
        q_activated_to: str = "",
        active_year=None,
        active_month=None,
        active_start_date=None,
        active_end_date=None,
        q_active_result: str = "",
    ) -> int:

        where_sql, params = self._build_where(
            q_jpid=q_jpid,
            q_name=q_name,
            q_name_kana=q_name_kana,
            q_introducer=q_introducer,
            q_placement=q_placement,
            q_status=q_status,
            q_rank=q_rank,
            q_interim_from=q_interim_from,
            q_interim_to=q_interim_to,
            q_activated_from=q_activated_from,
            q_activated_to=q_activated_to,
        )
        active_params = []
        active_join_sql = ""
        if (
            q_active_result in ("active", "inactive")
            and self._has_active_period(
                active_year,
                active_month,
                active_start_date,
                active_end_date,
            )
        ):
            active_join_sql = self._build_active_join_sql()
            active_params = [
                active_year,
                active_month,
                active_start_date,
                active_end_date,
            ]
            where_sql = self._apply_active_result_filter(where_sql, q_active_result)

        count_expr = "COUNT(DISTINCT u.id)" if active_join_sql else "COUNT(*)"
        sql = f"""
            SELECT {count_expr}
            FROM nexus_production.users u
            {active_join_sql}
            {where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, active_params + params)
            row = cursor.fetchone()

        return int(row[0]) if row else 0

    def _build_rows_sql(
        self,
        q_jpid: str = "",
        q_name: str = "",
        q_name_kana: str = "",
        q_introducer: str = "",
        q_placement: str = "",
        q_status: str = "",
        q_rank: str = "",
        q_interim_from: str = "",
        q_interim_to: str = "",
        q_activated_from: str = "",
        q_activated_to: str = "",
        active_year=None,
        active_month=None,
        active_start_date=None,
        active_end_date=None,
        q_active_result: str = "",
        order_sql: str = "",
        force_active_join: bool = False,
    ):
        """一覧のSQLとパラメータを返す。LIMIT/OFFSET は呼び出し側で付ける。"""

        where_sql, params = self._build_where(
            q_jpid=q_jpid,
            q_name=q_name,
            q_name_kana=q_name_kana,
            q_introducer=q_introducer,
            q_placement=q_placement,
            q_status=q_status,
            q_rank=q_rank,
            q_interim_from=q_interim_from,
            q_interim_to=q_interim_to,
            q_activated_from=q_activated_from,
            q_activated_to=q_activated_to,
        )

        active_params = []
        has_period = self._has_active_period(
            active_year,
            active_month,
            active_start_date,
            active_end_date,
        )
        use_inline_join = has_period and self._needs_inline_active_join(
            q_active_result=q_active_result,
            order_sql=order_sql,
            force_active_join=force_active_join,
        )
        if use_inline_join:
            active_select_sql = """
                CASE
                    WHEN au.active_status = 1 OR IFNULL(pb.prev_month_bv, 0) >= 50 THEN 1
                    ELSE 0
                END AS prev_month_active_status,
                IFNULL(pb.prev_month_bv, 0) AS prev_month_bv
            """
            active_join_sql = self._build_active_join_sql()
            active_params = [
                active_year,
                active_month,
                active_start_date,
                active_end_date,
            ]
            where_sql = self._apply_active_result_filter(where_sql, q_active_result)
        else:
            active_select_sql = """
                NULL AS prev_month_active_status,
                NULL AS prev_month_bv
            """
            active_join_sql = ""

        sql = f"""
            SELECT
                u.id,
                u.group_code,
                u.jmoa_code,
                u.send_bv_name,
                u.name_kana,
                u.introducer_code,
                u.placement_code,
                u.rank,
                u.status_code,
                u.activated,
                u.interim_at,
                u.activated_at,
                u.company,
                u.last_purchase_at,
                u.created_at,
                u.updated_at,
                {active_select_sql}
            FROM nexus_production.users u
            {active_join_sql}
            {where_sql}
            ORDER BY {order_sql or "u.created_at DESC"}, u.id DESC
        """

        return sql, active_params + params

    def _attach_prev_month_active_status(
        self,
        rows,
        active_year,
        active_month,
        active_start_date,
        active_end_date,
    ):
        if not rows or not self._has_active_period(
            active_year,
            active_month,
            active_start_date,
            active_end_date,
        ):
            return rows

        codes = [str(row.get("jmoa_code")) for row in rows if row.get("jmoa_code")]
        if not codes:
            return rows

        placeholders = ", ".join(["%s"] * len(codes))
        active_map = {}
        bv_map = {}

        with connections["rds"].cursor() as cursor:
            cursor.execute(
                f"""
SELECT jwoa_code, active_status
FROM bonus_db.active_users
WHERE year = %s
  AND month = %s
  AND jwoa_code IN ({placeholders})
                """,
                [active_year, active_month, *codes],
            )
            for code, status in cursor.fetchall():
                active_map[str(code)] = status

            cursor.execute(
                f"""
SELECT jwoa_code, SUM(IFNULL(bv, 0))
FROM bonus_db.purchase_info_list
WHERE bonus_payment_date >= %s
  AND bonus_payment_date < %s
  AND jwoa_code IN ({placeholders})
GROUP BY jwoa_code
                """,
                [active_start_date, active_end_date, *codes],
            )
            for code, bv in cursor.fetchall():
                bv_map[str(code)] = bv

        for row in rows:
            code = str(row.get("jmoa_code") or "")
            bv = bv_map.get(code) or 0
            row["prev_month_bv"] = bv
            row["prev_month_active_status"] = (
                1 if active_map.get(code) == 1 or bv >= 50 else 0
            )
        return rows

    def _fetch_rows(self, limit: int = 200, offset: int = 0, **filters):
        sql, params = self._build_rows_sql(**filters)

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql + " LIMIT %s OFFSET %s", params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        if (
            self._has_active_period(
                filters.get("active_year"),
                filters.get("active_month"),
                filters.get("active_start_date"),
                filters.get("active_end_date"),
            )
            and not self._needs_inline_active_join(
                q_active_result=filters.get("q_active_result") or "",
                order_sql=filters.get("order_sql") or "",
            )
        ):
            self._attach_prev_month_active_status(
                rows,
                filters.get("active_year"),
                filters.get("active_month"),
                filters.get("active_start_date"),
                filters.get("active_end_date"),
            )

        return rows

    def _resolve_active_period(self, q_active_kibetu):
        if not q_active_kibetu:
            return None, None, None, None, None

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=q_active_kibetu)
            .first()
        )
        if not period:
            return None, None, None, None, None

        selected_month_start = date(period.year, period.month, 1)
        selected_month_end = selected_month_start + relativedelta(months=1)
        return (
            period,
            selected_month_start.year,
            selected_month_start.month,
            selected_month_start,
            selected_month_end,
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        q_jpid = (self.request.GET.get("q_jpid") or "").strip()
        q_name = (self.request.GET.get("q_name") or "").strip()
        q_name_kana = (self.request.GET.get("q_name_kana") or "").strip()
        q_introducer = (self.request.GET.get("q_introducer") or "").strip()
        q_placement = (self.request.GET.get("q_placement") or "").strip()
        q_status = (self.request.GET.get("q_status") or "").strip()
        q_rank = (self.request.GET.get("q_rank") or "").strip()
        q_interim_from = (self.request.GET.get("q_interim_from") or "").strip()
        q_interim_to = (self.request.GET.get("q_interim_to") or "").strip()
        q_activated_from = (self.request.GET.get("q_activated_from") or "").strip()
        q_activated_to = (self.request.GET.get("q_activated_to") or "").strip()
        q_active_kibetu = (self.request.GET.get("q_active_kibetu") or "").strip()
        q_active_result = (self.request.GET.get("q_active_result") or "").strip()
        if not q_active_kibetu or q_active_result not in ("active", "inactive"):
            q_active_result = ""
        (
            selected_active_period,
            active_year,
            active_month,
            active_start_date,
            active_end_date,
        ) = self._resolve_active_period(q_active_kibetu)

        per_page = self.get_per_page()
        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        total_count = self._fetch_total_count(
            q_jpid=q_jpid,
            q_name=q_name,
            q_name_kana=q_name_kana,
            q_introducer=q_introducer,
            q_placement=q_placement,
            q_status=q_status,
            q_rank=q_rank,
            q_interim_from=q_interim_from,
            q_interim_to=q_interim_to,
            q_activated_from=q_activated_from,
            q_activated_to=q_activated_to,
            active_year=active_year,
            active_month=active_month,
            active_start_date=active_start_date,
            active_end_date=active_end_date,
            q_active_result=q_active_result,
        )

        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            q_jpid=q_jpid,
            q_name=q_name,
            q_name_kana=q_name_kana,
            q_introducer=q_introducer,
            q_placement=q_placement,
            q_status=q_status,
            q_rank=q_rank,
            q_interim_from=q_interim_from,
            q_interim_to=q_interim_to,
            q_activated_from=q_activated_from,
            q_activated_to=q_activated_to,
            active_year=active_year,
            active_month=active_month,
            active_start_date=active_start_date,
            active_end_date=active_end_date,
            q_active_result=q_active_result,
            order_sql=sort_ctx["order_sql"],
            limit=per_page,
            offset=offset,
        )

        ctx["q_jpid"] = q_jpid
        ctx["q_name"] = q_name
        ctx["q_name_kana"] = q_name_kana
        ctx["q_introducer"] = q_introducer
        ctx["q_placement"] = q_placement
        ctx["q_status"] = q_status
        ctx["q_rank"] = q_rank
        ctx["q_interim_from"] = q_interim_from
        ctx["q_interim_to"] = q_interim_to
        ctx["q_activated_from"] = q_activated_from
        ctx["q_activated_to"] = q_activated_to
        ctx["q_active_kibetu"] = q_active_kibetu
        ctx["q_active_result"] = q_active_result
        ctx["selected_active_period"] = selected_active_period
        ctx["active_year"] = active_year
        ctx["active_month"] = active_month
        ctx["monthly_period_choices"] = (
            MonthlyPeriod.objects.using("rds")
            .only("kibetu", "year", "month")
            .order_by("-kibetu")
        )
        # Excel出力は件数が多いと時間がかかるため、上限を超えたらCSVへ誘導する。
        ctx["excel_export_max_rows"] = UsersExportView.EXCEL_EXPORT_MAX_ROWS
        ctx["excel_export_allowed"] = (
            total_count <= UsersExportView.EXCEL_EXPORT_MAX_ROWS
        )

        base_params = {}

        if q_jpid:
            base_params["q_jpid"] = q_jpid

        if q_name:
            base_params["q_name"] = q_name

        if q_name_kana:
            base_params["q_name_kana"] = q_name_kana

        if q_introducer:
            base_params["q_introducer"] = q_introducer

        if q_placement:
            base_params["q_placement"] = q_placement

        if q_status:
            base_params["q_status"] = q_status

        if q_rank:
            base_params["q_rank"] = q_rank

        if q_interim_from:
            base_params["q_interim_from"] = q_interim_from

        if q_interim_to:
            base_params["q_interim_to"] = q_interim_to

        if q_activated_from:
            base_params["q_activated_from"] = q_activated_from

        if q_activated_to:
            base_params["q_activated_to"] = q_activated_to

        if q_active_kibetu:
            base_params["q_active_kibetu"] = q_active_kibetu

        if q_active_result:
            base_params["q_active_result"] = q_active_result

        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        add_sort_params(base_params, sort_ctx)

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )


class UsersExportView(UsersView):
    # Excelは1行あたりの書き出しが重く、上限なしだと7万件で1分近くかかる。
    EXCEL_EXPORT_MAX_ROWS = 10000
    EXPORT_HEADERS = [
        "グループID",
        "会員ID",
        "会員名",
        "氏名(カナ)",
        "上位者ID",
        "紹介者ID",
        "アクティブ確認",
        "ランク",
        "ステータス",
        "本登録FLG",
        "法人FLG",
        "仮登録日時",
        "本登録日時",
        "最終購入日",
        "作成日時",
        "更新日時",
    ]
    RANK_LABELS = {
        1: "シルバー",
        2: "ゴールド",
        3: "プラチナ",
        4: "ダイヤ",
        9: "一般会員",
    }
    STATUS_LABELS = {
        1: "アクティブ",
        2: "凍結",
        3: "退会",
        4: "中途解約",
        5: "非アクティブ",
    }

    def _rank_label(self, value):
        return self.RANK_LABELS.get(value, "-")

    def _status_label(self, value):
        return self.STATUS_LABELS.get(value, "-")

    def _activated_label(self, value):
        return "本登録" if value == 1 else "仮登録中"

    def _company_label(self, value):
        return "法人" if value == 1 else "-"

    def _active_status_label(self, row, q_active_kibetu):
        if not q_active_kibetu:
            return "-"
        if row.get("prev_month_active_status") == 1:
            return "アクティブ"
        return "非アクティブ"

    def _row_to_excel(self, row, q_active_kibetu):
        return [
            row.get("group_code"),
            row.get("jmoa_code"),
            row.get("send_bv_name"),
            row.get("name_kana") or "-",
            row.get("introducer_code"),
            row.get("placement_code"),
            self._active_status_label(row, q_active_kibetu),
            self._rank_label(row.get("rank")),
            self._status_label(row.get("status_code")),
            self._activated_label(row.get("activated")),
            self._company_label(row.get("company")),
            _excel_jst_datetime(row.get("interim_at")),
            _excel_jst_datetime(row.get("activated_at")),
            _excel_jst_datetime(row.get("last_purchase_at")),
            _excel_jst_datetime(row.get("created_at")),
            _excel_jst_datetime(row.get("updated_at")),
        ]

    def _build_row_mapper(self, cols, q_active_kibetu):
        """取得結果のタプルをExcelの1行に変換する関数を作る。

        7万件超を書き出すので、1行ごとに辞書を作らず列位置で直接読む。
        """
        pos = {name: i for i, name in enumerate(cols)}
        has_active = bool(q_active_kibetu)

        def to_excel_row(row):
            active_status = "-"
            if has_active:
                active_status = (
                    "アクティブ"
                    if row[pos["prev_month_active_status"]] == 1
                    else "非アクティブ"
                )
            return [
                row[pos["group_code"]],
                row[pos["jmoa_code"]],
                row[pos["send_bv_name"]],
                row[pos["name_kana"]] or "-",
                row[pos["introducer_code"]],
                row[pos["placement_code"]],
                active_status,
                self.RANK_LABELS.get(row[pos["rank"]], "-"),
                self.STATUS_LABELS.get(row[pos["status_code"]], "-"),
                "本登録" if row[pos["activated"]] == 1 else "仮登録中",
                "法人" if row[pos["company"]] == 1 else "-",
                _excel_jst_datetime(row[pos["interim_at"]]),
                _excel_jst_datetime(row[pos["activated_at"]]),
                _excel_jst_datetime(row[pos["last_purchase_at"]]),
                _excel_jst_datetime(row[pos["created_at"]]),
                _excel_jst_datetime(row[pos["updated_at"]]),
            ]

        return to_excel_row

    def _resolve_export_filters(self, request):
        """検索条件を一覧と同じ形で取り出す。"""
        q_active_kibetu = (request.GET.get("q_active_kibetu") or "").strip()
        q_active_result = (request.GET.get("q_active_result") or "").strip()
        if not q_active_kibetu or q_active_result not in ("active", "inactive"):
            q_active_result = ""
        (
            _selected_active_period,
            active_year,
            active_month,
            active_start_date,
            active_end_date,
        ) = self._resolve_active_period(q_active_kibetu)

        filters = {
            "q_jpid": (request.GET.get("q_jpid") or "").strip(),
            "q_name": (request.GET.get("q_name") or "").strip(),
            "q_name_kana": (request.GET.get("q_name_kana") or "").strip(),
            "q_introducer": (request.GET.get("q_introducer") or "").strip(),
            "q_placement": (request.GET.get("q_placement") or "").strip(),
            "q_status": (request.GET.get("q_status") or "").strip(),
            "q_rank": (request.GET.get("q_rank") or "").strip(),
            "q_interim_from": (request.GET.get("q_interim_from") or "").strip(),
            "q_interim_to": (request.GET.get("q_interim_to") or "").strip(),
            "q_activated_from": (request.GET.get("q_activated_from") or "").strip(),
            "q_activated_to": (request.GET.get("q_activated_to") or "").strip(),
            "active_year": active_year,
            "active_month": active_month,
            "active_start_date": active_start_date,
            "active_end_date": active_end_date,
            "q_active_result": q_active_result,
        }
        return filters, q_active_kibetu

    def _iter_export_rows(self, filters, q_active_kibetu):
        """検索結果をエクスポート用の1行ずつに変換して返す。

        LIMIT/OFFSET で分割すると毎回7万件超を並べ替え直すことになるため、
        クエリは1回だけ実行して結果を少しずつ読み出す。
        """
        sql, params = self._build_rows_sql(
            order_sql=self._get_sort_context()["order_sql"],
            force_active_join=True,
            **filters,
        )

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            to_export_row = self._build_row_mapper(cols, q_active_kibetu)
            while True:
                rows = cursor.fetchmany(self.EXPORT_FETCH_SIZE)
                if not rows:
                    break
                for row in rows:
                    yield to_export_row(row)

    def _list_url(self, request):
        params = request.GET.copy()
        params.pop("format", None)
        query = params.urlencode()
        url = reverse("connect:users")
        return f"{url}?{query}" if query else url

    def get(self, request, *args, **kwargs):
        filters, q_active_kibetu = self._resolve_export_filters(request)
        export_format = (request.GET.get("format") or "excel").strip().lower()

        if export_format == "csv":
            return self._export_csv(filters, q_active_kibetu)

        # 画面側でもボタンを止めているが、URLを直接叩かれても通さないようにする。
        total_count = self._fetch_total_count(**filters)
        if total_count > self.EXCEL_EXPORT_MAX_ROWS:
            messages.error(
                request,
                f"該当が{total_count:,}件あり、Excel出力の上限"
                f"（{self.EXCEL_EXPORT_MAX_ROWS:,}件）を超えています。"
                "条件を絞り込むか、CSV出力をご利用ください。",
            )
            return redirect(self._list_url(request))

        return self._export_excel(filters, q_active_kibetu)

    def _export_excel(self, filters, q_active_kibetu):
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("会員一覧")
        ws.append(self.EXPORT_HEADERS)

        for row in self._iter_export_rows(filters, q_active_kibetu):
            ws.append(row)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="users.xlsx"'
        wb.save(response)
        return response

    def _export_csv(self, filters, q_active_kibetu):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="users.csv"'
        # BOM が無いと Excel で開いたときに文字化けする。
        response.write("\ufeff")

        writer = csv.writer(response, lineterminator="\r\n")
        writer.writerow(self.EXPORT_HEADERS)
        for row in self._iter_export_rows(filters, q_active_kibetu):
            writer.writerow(["" if v is None else v for v in row])

        return response



class ThreeStarGlobalBonusView(generic.ListView):
    template_name = "three_star_global_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        return (
            MonthlyPeriod.objects.using("rds")
            .all()
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "three_star_global_bonus")

        if action != "three_star_global_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:three_star_global_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:three_star_global_bonus")

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:three_star_global_bonus")

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(request, "登録")
            return redirect(f"/three_star_global_bonus/?kibetu={selected_kibetu}")

        try:
            three_star_global_bonus_rows = self._get_three_star_global_bonus_rows(
                selected_kibetu=selected_kibetu,
                period=period,
            )

            if not three_star_global_bonus_rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "three_star_global_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(
                    f"/three_star_global_bonus_rows/?kibetu={selected_kibetu}"
                )

            insert_sql, insert_params = (
                register_sql.get_three_star_global_bonus_insert_data(
                    selected_kibetu,
                    three_star_global_bonus_rows,
                )
            )

            if not insert_params:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "three_star_global_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（登録対象なし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(
                    f"/three_star_global_bonus/?kibetu={selected_kibetu}"
                )

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    # 再購入オーバーボーナス登録
                    cursor.executemany(insert_sql, insert_params)

                    # 登録履歴
                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "three_star_global_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(insert_params)}件登録",
                        ],
                    )

            messages.success(
                request,
                f"{len(insert_params)}件を3スターダイヤグローバル配当結果に登録しました。"
            )

        except Exception as e:
            logger.exception("再購入オーバーボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(
            f"/three_star_global_bonus/?kibetu={selected_kibetu}"
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:three_star_global_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period
        if not ensure_kibetu_purchase_info(self.request, selected_kibetu, period):
            insert_empty_bonus_history_on_display(
                self.request,
                "three_star_global_bonus",
                selected_kibetu,
            )
            return ctx

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(self.request)
            return ctx

        ctx["rows"] = self._get_three_star_global_bonus_rows(
            selected_kibetu=selected_kibetu,
            period=period,
        )
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "three_star_global_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_three_star_global_bonus_rows(self, selected_kibetu, period):

        # 今月
        kibetu_year = period.year
        kibetu_month = period.month

        # 当月1日を作成
        current_date = date(kibetu_year, kibetu_month, 1)

        # 先月
        prev_month_period = current_date - relativedelta(months=1)

        prev_year = prev_month_period.year
        prev_month = prev_month_period.month

        params = [
            prev_year,
            prev_month,
            prev_year,
            prev_month,
            selected_kibetu,
            kibetu_year,
            kibetu_month,
            
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(
                THREE_STAR_DIAMOND_GLOBAL_BONUS_Q_SQL,
                params
            )

            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]

            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        return rows


class S_ThreeStarGlobalBonusView(generic.ListView):
    template_name = "s_three_star_global_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):

        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_three_star_global_bonus_result
                ORDER BY kibetu DESC
            """)

            registered_kibetu_list = [
                row[0]
                for row in cursor.fetchall()
            ]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            filename = build_bonus_export_filename("three_star_global_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["three_star_global_bonus"],
                "ThreeStarGlobalBonus",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "tsgbr.kibetu",
                "jwoa_code": "tsgbr.jwoa_code",
                "jwoa_name": "tsgbr.jwoa_name",
                "title_name": "title_name",
                "score": "tsgbr.score",
                "total_score": "tsgbr.total_score",
                "total_over_bv": "tsgbr.total_over_bv",
                "bonus_amount": "tsgbr.bonus_amount",
                "created_at": "tsgbr.created_at",
                "updated_at": "tsgbr.updated_at",
            },
            default_sort="bonus_amount",
            default_direction="desc",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                tsgbr.id,
                tsgbr.kibetu,
                tsgbr.jwoa_code,
                tsgbr.jwoa_name,
                tsgbr.title_id,
                COALESCE(tm.title_name, 'タイトルなし') AS title_name,
                tsgbr.score,
                tsgbr.total_score,
                tsgbr.total_over_bv,
                tsgbr.bonus_amount,
                tsgbr.created_at,
                tsgbr.updated_at
            FROM bonus_db.B_three_star_global_bonus_result AS tsgbr
            LEFT JOIN bonus_db.title_master AS tm
              ON tsgbr.title_id = tm.title_id
            WHERE tsgbr.kibetu = %s
        """

        params = [selected_kibetu]
        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "jwoa_code": "tsgbr.jwoa_code",
                "jwoa_name": "tsgbr.jwoa_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        ctx["rows"] = rows

        return ctx


class GlobalBonusView(generic.ListView):
    template_name = "global_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        return (
            MonthlyPeriod.objects.using("rds")
            .all()
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "global_bonus")

        if action != "global_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:global_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:global_bonus")

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:global_bonus")

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(request, "登録")
            return redirect(f"/global_bonus/?kibetu={selected_kibetu}")

        try:
            global_bonus_rows = self._get_global_bonus_rows(
                selected_kibetu=selected_kibetu,
                period=period,
            )

            if not global_bonus_rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "global_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(
                    f"/global_bonus_rows/?kibetu={selected_kibetu}"
                )

            insert_sql, insert_params = (
                register_sql.get_global_bonus_insert_data(
                    selected_kibetu,
                    global_bonus_rows,
                )
            )

            if not insert_params:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "global_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（登録対象なし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(
                    f"/global_bonus/?kibetu={selected_kibetu}"
                )

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    # グローバル配当登録
                    cursor.executemany(insert_sql, insert_params)

                    # 登録履歴
                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "global_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(insert_params)}件登録",
                        ],
                    )

            messages.success(
                request,
                f"{len(insert_params)}件をグローバル配当結果に登録しました。"
            )

        except Exception as e:
            logger.exception("グローバル配当結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(
            f"/global_bonus/?kibetu={selected_kibetu}"
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:global_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period

        if not has_month_title_rows(selected_kibetu):
            warn_month_title_required(self.request)
            return ctx

        ctx["rows"] = self._get_global_bonus_rows(
            selected_kibetu=selected_kibetu,
            period=period,
        )
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "global_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_global_bonus_rows(self, selected_kibetu, period):

        # 今月
        kibetu_year = period.year
        kibetu_month = period.month

        # 当月1日を作成
        current_date = date(kibetu_year, kibetu_month, 1)

        # 先月
        prev_month_period = current_date - relativedelta(months=1)

        prev_year = prev_month_period.year
        prev_month = prev_month_period.month

        params = [
            prev_year,
            prev_month,
            prev_year,
            prev_month,
            selected_kibetu,
            kibetu_year,
            kibetu_month,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(
                CROWN_DIAMOND_GLOBAL_BONUS_Y_SQL,
                params
            )

            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]

            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        return rows


class S_GlobalBonusView(generic.ListView):
    template_name = "s_global_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):

        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_global_bonus_result
                ORDER BY kibetu DESC
            """)

            registered_kibetu_list = [
                row[0]
                for row in cursor.fetchall()
            ]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            filename = build_bonus_export_filename("global_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["global_bonus"],
                "GlobalBonus",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()

        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu=selected_kibetu)
            .first()
        )

        if not period:
            return ctx

        ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "gbr.kibetu",
                "jwoa_code": "gbr.jwoa_code",
                "jwoa_name": "gbr.jwoa_name",
                "title_name": "title_name",
                "score": "gbr.score",
                "total_score": "gbr.total_score",
                "total_bv": "gbr.total_bv",
                "bonus_amount": "gbr.bonus_amount",
                "created_at": "gbr.created_at",
                "updated_at": "gbr.updated_at",
            },
            default_sort="bonus_amount",
            default_direction="desc",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                gbr.id,
                gbr.kibetu,
                gbr.jwoa_code,
                gbr.jwoa_name,
                gbr.title_id,
                COALESCE(tm.title_name, 'タイトルなし') AS title_name,
                gbr.score,
                gbr.total_score,
                gbr.total_bv,
                gbr.bonus_amount,
                gbr.created_at,
                gbr.updated_at
            FROM bonus_db.B_global_bonus_result AS gbr
            LEFT JOIN bonus_db.title_master AS tm
              ON gbr.title_id = tm.title_id
            WHERE gbr.kibetu = %s
        """

        params = [selected_kibetu]
        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "jwoa_code": "gbr.jwoa_code",
                "jwoa_name": "gbr.jwoa_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")

            cols = [c[0] for c in cursor.description]
            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        ctx["rows"] = rows

        return ctx



class OrdersView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "orders.html"

    # nexus_production.orders のインデックスは id / order_code / order_status /
    # order_type / payment_option だけで created_at には無い。created_at で並べると
    # 10 万行のフルスキャン + filesort になるが、id は採番順で created_at と並びが
    # 完全に一致する（全件で逆行 0 件、NULL なし）ため主キー走査に置き換えている。
    SORT_COLUMNS = {
        "order_code": "o.order_code",
        "order_status": "o.order_status",
        "order_type": "o.order_type",
        "order_year": "o.order_year",
        "order_month": "o.order_month",
        "jwoa_code": "o.jwoa_code",
        "order_name": "o.order_name",
        "total_price": "o.total_price",
        "total_bv": "o.total_bv",
        "bv_actived_flg": "o.bv_actived_flg",
        "distribution_count": "distribution_count",
        "deposit_at": "o.deposit_at",
        "created_at": "o.id",
    }

    DISTRIBUTION_COUNT_SORT = "distribution_count"

    # 注文番号は "MF" + 数字（末尾に _FI / _FR が付くものもある）で、"MF" は先頭以外に
    # 現れない。そのため "MF" で始まる入力なら部分一致と前方一致の結果は同じになる。
    # 前方一致にすると order_code のインデックスが使えて、部分一致の約 6 秒が 0.1 秒になる。
    ORDER_CODE_PREFIX = "MF"

    # 会員ID部分一致で拾う振分先 order_code の上限。これを超えたら IN 句が
    # 肥大するので従来の相関 EXISTS に戻す。
    MAX_DISTRIBUTION_MATCH_CODES = 10000

    TOTAL_COUNT_CACHE_TIMEOUT = 30

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        # 一覧・件数・Excel出力で同じ結果を使い回すためのリクエスト内キャッシュ
        self._distribution_code_cache = {}

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="created_at",
            default_direction="desc",
        )

    def _build_order_by(self, sort_ctx):
        """並び替え句を組み立てる。

        同順位を決める id は並び替え列と同じ向きに揃える。向きが混ざると MySQL は
        二次インデックス（末尾に主キーを含む）を並び替えに使えず、フルスキャン +
        filesort に落ちる。
        """
        column = self.SORT_COLUMNS.get(sort_ctx["sort"], "o.id")
        direction = "DESC" if sort_ctx["direction"] == "desc" else "ASC"

        if column == "o.id":
            return f"o.id {direction}"

        return f"{column} {direction}, o.id {direction}"

    def _distribution_order_codes(self, q_jwoa_code):
        """会員IDの部分一致に該当する BV 振分先の order_code を返す。

        相関 EXISTS のままだと orders の 10 万行ごとに振分テーブルを引くうえ、
        一覧と件数で 2 回走る。先に振分テーブルだけを引いて order_code の集合に
        しておき、一覧・件数・Excel出力で使い回す。

        該当が上限を超えたときは None を返し、呼び出し側で EXISTS に戻す。
        """
        if q_jwoa_code in self._distribution_code_cache:
            return self._distribution_code_cache[q_jwoa_code]

        sql = """
            SELECT d.order_code
            FROM nexus_production.orders_distribution_bv d
            WHERE d.jwoa_code LIKE %s
            LIMIT %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(
                sql,
                [f"%{q_jwoa_code}%", self.MAX_DISTRIBUTION_MATCH_CODES + 1],
            )
            fetched = [row[0] for row in cursor.fetchall() if row[0]]

        if len(fetched) > self.MAX_DISTRIBUTION_MATCH_CODES:
            codes = None
        else:
            # SQL 側の DISTINCT は一時テーブルを作って数倍遅くなるので Python で重複を除く
            codes = list(dict.fromkeys(fetched))

        self._distribution_code_cache[q_jwoa_code] = codes
        return codes

    def _build_where(
        self,
        q_order_code="",
        q_jwoa_code="",
        q_name="",
        q_order_statuses=None,
        q_order_types=None,
        q_deposit_from="",
        q_deposit_to="",
        q_year="",
        q_month="",
        q_bv_actived_flg="",
    ):
        if q_order_statuses is None:
            q_order_statuses = []
        if q_order_types is None:
            q_order_types = []

        where = []
        params = []

        if q_order_code:
            where.append("o.order_code LIKE %s")
            if q_order_code.startswith(self.ORDER_CODE_PREFIX):
                params.append(f"{q_order_code}%")
            else:
                params.append(f"%{q_order_code}%")

        if q_jwoa_code:
            like = f"%{q_jwoa_code}%"
            matched_codes = self._distribution_order_codes(q_jwoa_code)

            if matched_codes is None:
                where.append(
                    """
                    (
                        o.jwoa_code LIKE %s
                        OR EXISTS (
                            SELECT 1
                            FROM nexus_production.orders_distribution_bv d
                            WHERE d.order_code = o.order_code
                              AND d.jwoa_code LIKE %s
                        )
                    )
                    """
                )
                params.append(like)
                params.append(like)

            elif matched_codes:
                placeholders = ", ".join(["%s"] * len(matched_codes))
                where.append(
                    f"(o.jwoa_code LIKE %s OR o.order_code IN ({placeholders}))"
                )
                params.append(like)
                params.extend(matched_codes)

            else:
                where.append("o.jwoa_code LIKE %s")
                params.append(like)

        if q_name:
            where.append("o.order_name LIKE %s")
            params.append(f"%{q_name}%")

        if q_order_statuses:
            placeholders = ", ".join(["%s"] * len(q_order_statuses))
            where.append(f"o.order_status IN ({placeholders})")
            params.extend(q_order_statuses)

        if q_order_types:
            placeholders = ", ".join(["%s"] * len(q_order_types))
            where.append(f"o.order_type IN ({placeholders})")
            params.extend(q_order_types)

        if q_deposit_from:
            where.append("o.deposit_at >= %s")
            params.append(q_deposit_from)

        if q_deposit_to:
            where.append("o.deposit_at < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_deposit_to)

        if q_year:
            where.append("o.order_year = %s")
            params.append(q_year)

        if q_month:
            where.append("o.order_month = %s")
            params.append(q_month)

        if q_bv_actived_flg in ("0", "1", "3"):
            where.append("o.bv_actived_flg = %s")
            params.append(int(q_bv_actived_flg))

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return where_sql, params

    def _total_count_cache_key(self, **filters):
        signature = json.dumps(filters, sort_keys=True, default=str)
        digest = hashlib.md5(signature.encode("utf-8")).hexdigest()
        return f"orders:total_count:{digest}"

    def _fetch_total_count(self, **filters):
        """一覧の総件数を返す。

        orders は絞り込み対象の列にインデックスがほとんど無く、COUNT(*) だけで
        1 秒前後（会員ID絞り込みでは数秒）かかる。ページを送るたびに同じ数を
        数え直さないよう、条件ごとに短時間だけキャッシュする。
        """
        cache_key = self._total_count_cache_key(**filters)
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        where_sql, params = self._build_where(**filters)

        sql = f"""
            SELECT COUNT(*)
            FROM nexus_production.orders o
            {where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()

        total_count = int(row[0]) if row else 0
        cache.set(cache_key, total_count, self.TOTAL_COUNT_CACHE_TIMEOUT)
        return total_count

    def _fetch_distribution_counts(self, order_codes):
        """表示するページ分の BV 振分件数だけをまとめて 1 クエリで取得する。"""
        codes = list(dict.fromkeys(code for code in order_codes if code))
        if not codes:
            return {}

        placeholders = ", ".join(["%s"] * len(codes))
        sql = f"""
            SELECT d.order_code, COUNT(*)
            FROM nexus_production.orders_distribution_bv d
            WHERE d.order_code IN ({placeholders})
            GROUP BY d.order_code
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, codes)
            return {row[0]: int(row[1]) for row in cursor.fetchall()}

    def _fetch_rows(self, limit=200, offset=0, order_by="", sort="", **filters):
        where_sql, params = self._build_where(**filters)
        q_jwoa_code = (filters.get("q_jwoa_code") or "").strip()
        select_params = []

        if q_jwoa_code:
            member_match_select = (
                "CASE WHEN o.jwoa_code LIKE %s THEN 1 ELSE 0 END AS order_member_matched"
            )
            select_params.append(f"%{q_jwoa_code}%")
        else:
            member_match_select = "0 AS order_member_matched"

        # 振分件数を相関サブクエリで持つと 1 行ごとに振分テーブルを引くことになるので、
        # 通常は SELECT に含めず取得後にまとめて引く。振分件数で並び替えるときだけは
        # SQL 側に値が必要なのでサブクエリを使う。
        sort_by_distribution = sort == self.DISTRIBUTION_COUNT_SORT

        if sort_by_distribution:
            distribution_select = """,
                (
                    SELECT COUNT(*)
                    FROM nexus_production.orders_distribution_bv d_count
                    WHERE d_count.order_code = o.order_code
                ) AS distribution_count"""
        else:
            distribution_select = ""

        sql = f"""
            SELECT
                o.id,
                o.order_code,
                o.order_status,
                o.order_type,
                o.order_year,
                o.order_month,
                o.jwoa_code,
                o.order_name,
                o.total_price,
                o.total_bv,
                o.bv_actived_flg,
                o.deposit_at,
                o.created_at,
                {member_match_select}{distribution_select}
            FROM nexus_production.orders o
            {where_sql}
            ORDER BY {order_by or "o.id DESC"}
            LIMIT %s OFFSET %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, select_params + params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        if not sort_by_distribution:
            counts = self._fetch_distribution_counts(
                row["order_code"] for row in rows
            )
            for row in rows:
                row["distribution_count"] = counts.get(row["order_code"], 0)

        return rows

    def _get_filters(self):
        """一覧と Excel 出力で同じ検索条件を使うため、GET から絞り込み条件を取り出す。"""
        q_order_statuses = [
            x for x in self.request.GET.getlist("q_order_status") if x
        ]
        q_order_types = [
            x for x in self.request.GET.getlist("q_order_type") if x
        ]

        return {
            "q_order_code": (self.request.GET.get("q_order_code") or "").strip(),
            "q_jwoa_code": (self.request.GET.get("q_jwoa_code") or "").strip(),
            "q_name": (self.request.GET.get("q_name") or "").strip(),
            "q_order_statuses": q_order_statuses,
            "q_order_types": q_order_types,
            "q_deposit_from": (self.request.GET.get("q_deposit_from") or "").strip(),
            "q_deposit_to": (self.request.GET.get("q_deposit_to") or "").strip(),
            "q_year": (self.request.GET.get("q_year") or "").strip(),
            "q_month": (self.request.GET.get("q_month") or "").strip(),
            "q_bv_actived_flg": (self.request.GET.get("q_bv_actived_flg") or "").strip(),
        }

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        filters = self._get_filters()
        q_order_statuses = filters["q_order_statuses"]
        q_order_types = filters["q_order_types"]

        per_page = self.get_per_page()
        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        total_count = self._fetch_total_count(**filters)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            limit=per_page,
            offset=offset,
            order_by=self._build_order_by(sort_ctx),
            sort=sort_ctx["sort"],
            **filters,
        )

        ctx["q_order_code"] = filters["q_order_code"]
        ctx["q_jwoa_code"] = filters["q_jwoa_code"]
        ctx["q_name"] = filters["q_name"]
        ctx["q_order_statuses"] = q_order_statuses
        ctx["q_order_types"] = q_order_types
        ctx["q_deposit_from"] = filters["q_deposit_from"]
        ctx["q_deposit_to"] = filters["q_deposit_to"]
        ctx["q_year"] = filters["q_year"]
        ctx["q_month"] = filters["q_month"]
        ctx["q_bv_actived_flg"] = filters["q_bv_actived_flg"]

        base_params = {}
        for key in (
            "q_order_code",
            "q_jwoa_code",
            "q_name",
            "q_deposit_from",
            "q_deposit_to",
            "q_year",
            "q_month",
            "q_bv_actived_flg",
        ):
            value = filters[key]
            if value:
                base_params[key] = value

        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        add_sort_params(base_params, sort_ctx)

        ctx = self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )

        base_qs = ctx["base_qs"]
        for status in q_order_statuses:
            if base_qs:
                base_qs += "&"
            base_qs += urlencode({"q_order_status": status})
        for order_type in q_order_types:
            if base_qs:
                base_qs += "&"
            base_qs += urlencode({"q_order_type": order_type})
        ctx["base_qs"] = base_qs

        return ctx


class OrdersExportView(OrdersView):
    """注文一覧の絞り込み条件そのままで Excel 出力する。

    注文は 10 万件規模になるため、次の 3 点で落ちないようにしている。
    - 取得は主キーのキーセット方式で分割する（OFFSET を深くするとスキャン量が
      跳ね上がり、タイムアウトの原因になるため）
    - openpyxl は write_only で使い、全行をメモリ上のシートに保持しない
    - Excel の行数上限（1,048,576 行）に届く前に件数で打ち切り、
      検索条件の絞り込みを促す

    並び順は一覧画面の created_at 降順ではなく id 降順にしている。
    created_at は NULL を取り得てキーセットの継続条件が途切れるおそれがあるためで、
    id は採番順なので実質的に新しい順のまま出力される。
    """

    EXPORT_FETCH_SIZE = 5000
    MAX_EXPORT_ROWS = 300000

    ORDER_STATUS_LABELS = {
        201: "入金待ち",
        202: "入金確認済",
        203: "決済完了",
        204: "出荷依頼済",
        205: "出荷完了",
        206: "キャンセル",
        207: "返品処理中",
        208: "返品処理完了",
        209: "商品交換処理中",
        210: "再出荷依頼中",
        211: "再出荷完了",
    }

    ORDER_TYPE_LABELS = {
        101: "再購入品",
        102: "初回購入品",
        103: "ランクアップ購入品",
        105: "特別対応購入品",
        200: "クーリングオフ",
    }

    BV_ACTIVED_FLG_LABELS = {
        0: "未反映",
        1: "反映済",
        3: "反映無効",
    }

    EXPORT_HEADER = [
        "ID",
        "注文番号",
        "注文状況",
        "注文区分",
        "注文年",
        "注文月",
        "注文者_会員ID",
        "注文者_氏名",
        "購入合計金額",
        "合計BV",
        "BV反映FLG",
        "BV振分件数",
        "入金日",
        "作成日時",
    ]

    def _fetch_export_rows(self, last_id=None, limit=None, **filters):
        """id の降順で 1 チャンク分だけ取得する。last_id より小さい id が次のチャンク。"""
        where_sql, params = self._build_where(**filters)

        if last_id is not None:
            keyset_sql = "o.id < %s"
            if where_sql:
                where_sql = where_sql + " AND " + keyset_sql
            else:
                where_sql = "WHERE " + keyset_sql
            params = params + [last_id]

        sql = f"""
            SELECT
                o.id,
                o.order_code,
                o.order_status,
                o.order_type,
                o.order_year,
                o.order_month,
                o.jwoa_code,
                o.order_name,
                o.total_price,
                o.total_bv,
                o.bv_actived_flg,
                (
                    SELECT COUNT(*)
                    FROM nexus_production.orders_distribution_bv d_count
                    WHERE d_count.order_code = o.order_code
                ) AS distribution_count,
                o.deposit_at,
                o.created_at
            FROM nexus_production.orders o
            {where_sql}
            ORDER BY o.id DESC
            LIMIT %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit or self.EXPORT_FETCH_SIZE])
            return cursor.fetchall()

    @staticmethod
    def _excel_value(value):
        """Excel はタイムゾーン付き日時を扱えないので、naive に落としてから渡す。"""
        return _excel_jst_datetime(value)

    def _row_to_excel(self, row):
        (
            order_id,
            order_code,
            order_status,
            order_type,
            order_year,
            order_month,
            jwoa_code,
            order_name,
            total_price,
            total_bv,
            bv_actived_flg,
            distribution_count,
            deposit_at,
            created_at,
        ) = row

        return [
            order_id,
            order_code,
            self.ORDER_STATUS_LABELS.get(order_status, order_status),
            self.ORDER_TYPE_LABELS.get(order_type, order_type),
            order_year,
            order_month,
            jwoa_code,
            order_name,
            total_price,
            total_bv,
            self.BV_ACTIVED_FLG_LABELS.get(bv_actived_flg, bv_actived_flg),
            distribution_count,
            self._excel_value(deposit_at),
            self._excel_value(created_at),
        ]

    def get(self, request, *args, **kwargs):
        filters = self._get_filters()

        total_count = self._fetch_total_count(**filters)

        if total_count == 0:
            messages.error(request, "出力対象のデータがありません。")
            return redirect(self._back_url(request))

        if total_count > self.MAX_EXPORT_ROWS:
            messages.error(
                request,
                f"対象が{total_count:,}件あります。"
                f"Excel出力は{self.MAX_EXPORT_ROWS:,}件までです。"
                "検索条件を絞り込んでください。",
            )
            return redirect(self._back_url(request))

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("注文一覧")
        ws.append(self.EXPORT_HEADER)

        last_id = None
        while True:
            rows = self._fetch_export_rows(
                last_id=last_id,
                limit=self.EXPORT_FETCH_SIZE,
                **filters,
            )
            if not rows:
                break

            for row in rows:
                ws.append(self._row_to_excel(row))

            last_id = rows[-1][0]

            if len(rows) < self.EXPORT_FETCH_SIZE:
                break

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
        wb.save(response)
        return response

    @staticmethod
    def _back_url(request):
        """絞り込み条件を保ったまま注文一覧へ戻す。"""
        query = request.GET.urlencode()
        url = reverse("connect:orders")
        return f"{url}?{query}" if query else url


class OrderDetailView(generic.TemplateView):
    template_name = "order_detail.html"

    @staticmethod
    def _format_order_value(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return jst_datetime(value).strftime("%Y-%m-%d %H:%M:%S")
        return value

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        from connect.order_field_labels import get_order_field_label

        order_id = self.kwargs.get("pk")

        sql = """
            SELECT
                *
            FROM nexus_production.orders
            WHERE id = %s
            LIMIT 1
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [order_id])
            cols = [c[0] for c in cursor.description]
            row = cursor.fetchone()

        if row:
            order = dict(zip(cols, row))
            ctx["order"] = order
            # 注文状況だけバッジで色分けするため、列名も一緒に渡す。
            ctx["order_rows"] = [
                (
                    col,
                    get_order_field_label(col),
                    self._format_order_value(order.get(col)),
                )
                for col in cols
            ]
        else:
            ctx["order"] = None
            ctx["order_rows"] = []

        return ctx



class OrdersDistributionBvView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "orders_distribution_bv.html"

    # created_at にはインデックスが無く、並べると 24 万行の filesort になる。
    # id は採番順で created_at と並びが完全に一致する（全件で逆行 0 件、NULL なし）
    # ため、作成日時ソートは主キー走査に置き換える。
    SORT_COLUMNS = {
        "order_code": "a.order_code",
        "purchaser_jwoa_code": "b.jwoa_code",
        "jwoa_code": "a.jwoa_code",
        "distribution_bv": "a.distribution_bv",
        "usage_fee": "a.usage_fee",
        "bv_actived_flg": "b.bv_actived_flg",
        "created_at": "a.id",
        "updated_at": "a.updated_at",
    }
    ORDERS_JOIN_SORTS = frozenset({"purchaser_jwoa_code", "bv_actived_flg"})
    ORDER_CODE_PREFIX = "MF"
    JWOA_CODE_PREFIXES = ("JP", "JN", "FREE")
    TOTAL_COUNT_CACHE_TIMEOUT = 30

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="created_at",
            default_direction="desc",
        )

    def _build_order_by(self, sort_ctx):
        """並び替え句を組み立てる。

        同順位を決める id は並び替え列と同じ向きに揃える。向きが混ざると MySQL は
        二次インデックス（末尾に主キーを含む）を並び替えに使えず、フルスキャン +
        filesort に落ちる。
        """
        column = self.SORT_COLUMNS.get(sort_ctx["sort"], "a.id")
        direction = "DESC" if sort_ctx["direction"] == "desc" else "ASC"

        if column == "a.id":
            return f"a.id {direction}"

        return f"{column} {direction}, a.id {direction}"

    @staticmethod
    def _like_param(value, prefix=None, prefixes=()):
        if prefix and value.startswith(prefix):
            return f"{value}%"
        if prefixes and any(value.startswith(p) for p in prefixes):
            return f"{value}%"
        return f"%{value}%"

    def _needs_orders_join(self, sort="", **filters):
        """購入者側の列で絞り込む／並べ替えるときだけ orders を JOIN する。"""
        if (filters.get("q_purchaser_jwoa_code") or "").strip():
            return True
        if (filters.get("q_bv_actived_flg") or "").strip() in ("0", "1", "3"):
            return True
        return sort in self.ORDERS_JOIN_SORTS

    def _build_where(
        self,
        q_order_code="",
        q_user_id="",
        q_jwoa_code="",
        q_purchaser_jwoa_code="",
        q_bv_actived_flg="",
        q_created_from="",
        q_created_to="",
    ):
        where = []
        params = []

        if q_order_code:
            where.append("a.order_code LIKE %s")
            params.append(
                self._like_param(q_order_code, prefix=self.ORDER_CODE_PREFIX)
            )

        if q_user_id:
            where.append("a.user_id = %s")
            params.append(q_user_id)

        if q_jwoa_code:
            where.append("a.jwoa_code LIKE %s")
            params.append(
                self._like_param(q_jwoa_code, prefixes=self.JWOA_CODE_PREFIXES)
            )

        if q_purchaser_jwoa_code:
            where.append("b.jwoa_code LIKE %s")
            params.append(
                self._like_param(
                    q_purchaser_jwoa_code, prefixes=self.JWOA_CODE_PREFIXES
                )
            )

        if q_bv_actived_flg in ("0", "1", "3"):
            where.append("b.bv_actived_flg = %s")
            params.append(int(q_bv_actived_flg))

        if q_created_from:
            where.append("a.created_at >= %s")
            params.append(q_created_from)

        if q_created_to:
            where.append("a.created_at < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_created_to)

        where_sql = "WHERE " + " AND ".join(where) if where else ""

        return where_sql, params

    def _total_count_cache_key(self, **filters):
        signature = json.dumps(filters, sort_keys=True, default=str)
        digest = hashlib.md5(signature.encode("utf-8")).hexdigest()
        return f"orders_distribution_bv:total_count:{digest}"

    def _orders_join_sql(self):
        return """
            LEFT JOIN nexus_production.orders AS b
                ON a.order_code = b.order_code
        """

    def _fetch_total_count(self, **filters):
        """一覧の総件数を返す。

        購入者側の列を使わない件数は JOIN 不要。orders とつなぐと 24 万行の
        突合せになり、COUNT だけで 10 秒超になる。ページを送るたびに同じ数を
        数え直さないよう、条件ごとに短時間だけキャッシュする。
        """
        cache_key = self._total_count_cache_key(**filters)
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        where_sql, params = self._build_where(**filters)
        join_sql = self._orders_join_sql() if self._needs_orders_join(**filters) else ""

        sql = f"""
            SELECT COUNT(*)
            FROM nexus_production.orders_distribution_bv AS a
            {join_sql}
            {where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()

        total_count = int(row[0]) if row else 0
        cache.set(cache_key, total_count, self.TOTAL_COUNT_CACHE_TIMEOUT)
        return total_count

    def _fetch_order_info(self, order_codes):
        """表示するページ分の購入者情報だけをまとめて 1 クエリで取得する。"""
        codes = list(dict.fromkeys(code for code in order_codes if code))
        if not codes:
            return {}

        placeholders = ", ".join(["%s"] * len(codes))
        sql = f"""
            SELECT order_code, jwoa_code, bv_actived_flg
            FROM nexus_production.orders
            WHERE order_code IN ({placeholders})
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, codes)
            return {
                row[0]: {"jwoa_code": row[1], "bv_actived_flg": row[2]}
                for row in cursor.fetchall()
            }

    def _fetch_rows(self, limit=200, offset=0, order_sql="", sort="", **filters):
        where_sql, params = self._build_where(**filters)
        needs_join = self._needs_orders_join(sort=sort, **filters)

        if needs_join:
            extra_select = """,
                b.jwoa_code AS purchaser_jwoa_code,
                b.bv_actived_flg"""
            join_sql = self._orders_join_sql()
        else:
            extra_select = ""
            join_sql = ""

        sql = f"""
            SELECT
                a.id,
                a.order_code,
                a.user_id,
                a.jwoa_code,
                a.distribution_bv,
                a.usage_fee,
                a.created_at,
                a.updated_at
                {extra_select}
            FROM nexus_production.orders_distribution_bv AS a
            {join_sql}
            {where_sql}
            ORDER BY {order_sql or "a.id DESC"}
            LIMIT %s OFFSET %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        if not needs_join:
            order_info = self._fetch_order_info(row["order_code"] for row in rows)
            for row in rows:
                info = order_info.get(row["order_code"], {})
                row["purchaser_jwoa_code"] = info.get("jwoa_code")
                row["bv_actived_flg"] = info.get("bv_actived_flg")

        return rows

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        filters = {
            "q_order_code": (self.request.GET.get("q_order_code") or "").strip(),
            "q_user_id": (self.request.GET.get("q_user_id") or "").strip(),
            "q_jwoa_code": (self.request.GET.get("q_jwoa_code") or "").strip(),
            "q_purchaser_jwoa_code": (
                self.request.GET.get("q_purchaser_jwoa_code") or ""
            ).strip(),
            "q_bv_actived_flg": (
                self.request.GET.get("q_bv_actived_flg") or ""
            ).strip(),
            "q_created_from": (self.request.GET.get("q_created_from") or "").strip(),
            "q_created_to": (self.request.GET.get("q_created_to") or "").strip(),
        }
        return_to = (self.request.GET.get("return_to") or "").strip()
        if not return_to.startswith("/") or return_to.startswith("//"):
            return_to = ""

        per_page = self.get_per_page()
        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        total_count = self._fetch_total_count(**filters)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            limit=per_page,
            offset=offset,
            order_sql=self._build_order_by(sort_ctx),
            sort=sort_ctx["sort"],
            **filters,
        )

        ctx.update(filters)
        ctx["return_to"] = return_to

        base_params = {
            k: v for k, v in filters.items()
            if v
        }
        if return_to:
            base_params["return_to"] = return_to

        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        add_sort_params(base_params, sort_ctx)

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )


class OrdersDistributionBvUpdateView(OrdersDistributionBvView):
    template_name = "orders_distribution_bv_update.html"

    def _build_where(
        self,
        q_order_code="",
        q_user_id="",
        q_jwoa_code="",
        q_created_from="",
        q_created_to="",
    ):
        where = []
        params = []

        if q_order_code:
            where.append("a.order_code LIKE %s")
            params.append(f"%{q_order_code}%")

        if q_jwoa_code:
            where.append("a.jwoa_code LIKE %s")
            params.append(f"%{q_jwoa_code}%")

        if q_created_from:
            where.append("a.created_at >= %s")
            params.append(q_created_from)

        if q_created_to:
            where.append("a.created_at < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_created_to)

        where_sql = "WHERE " + " AND ".join(where) if where else ""

        return where_sql, params

    def _fetch_total_count(self, **filters):
        where_sql, params = self._build_where(**filters)

        sql = f"""
            SELECT COUNT(*)
            FROM bonus_db.purchase_info_list AS a
            {where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()

        return int(row[0]) if row else 0

    def _fetch_rows(self, limit=200, offset=0, **filters):
        where_sql, params = self._build_where(**filters)

        sql = f"""
            SELECT
                a.id,
                a.order_code,
                a.jwoa_code,
                a.send_bv_name,
                a.total_bv,
                a.bv,
                a.created_at,
                a.updated_at
            FROM bonus_db.purchase_info_list AS a
            {where_sql}
            ORDER BY a.id DESC
            LIMIT %s OFFSET %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            rows = [
                dict(zip(cols, r))
                for r in cursor.fetchall()
            ]

        return rows

    def get_context_data(self, **kwargs):
        ctx = super(OrdersDistributionBvView, self).get_context_data(**kwargs)

        filters = {
            "q_order_code": (self.request.GET.get("q_order_code") or "").strip(),
            "q_jwoa_code": (self.request.GET.get("q_jwoa_code") or "").strip(),
            "q_created_from": (self.request.GET.get("q_created_from") or "").strip(),
            "q_created_to": (self.request.GET.get("q_created_to") or "").strip(),
        }

        per_page = self.get_per_page()
        total_count = self._fetch_total_count(**filters)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            limit=per_page,
            offset=offset,
            **filters,
        )

        ctx.update(filters)
        base_params = {
            k: v for k, v in filters.items()
            if v
        }

        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )

    def post(self, request, *args, **kwargs):
        user_access = get_user_access(request.user)
        if not user_access.can_menu("orders_distribution_bv_update") or not user_access.can_update:
            return HttpResponse("権限がありません。", status=403)

        row_id = (request.POST.get("id") or "").strip()
        bv = (request.POST.get("bv") or "").strip()
        next_query = (request.POST.get("next_query") or "").strip()
        redirect_url = reverse("connect:orders_distribution_bv_update")
        if next_query:
            redirect_url = f"{redirect_url}?{next_query}"

        try:
            row_id_int = int(row_id)
        except ValueError:
            messages.error(request, "更新対象IDが不正です。")
            return redirect(redirect_url)

        try:
            bv_int = int(bv)
        except ValueError:
            messages.error(request, "BVは整数で入力してください。")
            return redirect(redirect_url)

        if bv_int < 0:
            messages.error(request, "BVは0以上で入力してください。")
            return redirect(redirect_url)

        before_row = fetch_one_dict(
            "rds",
            """
                SELECT
                    id,
                    order_code,
                    jwoa_code,
                    send_bv_name,
                    total_bv,
                    bv,
                    created_at,
                    updated_at
                FROM bonus_db.purchase_info_list
                WHERE id = %s
            """,
            [row_id_int],
        )
        if not before_row:
            messages.error(request, "更新対象データが見つかりませんでした。")
            return redirect(redirect_url)

        sql = """
            UPDATE bonus_db.purchase_info_list
            SET
                bv = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """

        try:
            with connections["rds"].cursor() as cursor:
                logger.info(
                    "BV振分変更画面からpurchase_info_listのBV更新SQLを実行します。id=%s bv=%s",
                    row_id_int,
                    bv_int,
                )
                cursor.execute(sql, [bv_int, row_id_int])
                updated_count = cursor.rowcount

            if updated_count:
                after_row = dict(before_row)
                after_row["bv"] = bv_int
                record_change_audit(
                    request,
                    screen_name="BV振分変更",
                    action_type="update",
                    target_table="purchase_info_list",
                    target_pk=row_id_int,
                    summary=(
                        f"注文番号 {before_row.get('order_code')} / "
                        f"JWOA会員ID {before_row.get('jwoa_code')} のBVを更新"
                    ),
                    before_values=before_row,
                    after_values=after_row,
                )
                messages.success(request, "BVを更新しました。")
            else:
                messages.info(request, "BVは変更されていません。")
        except Exception as e:
            logger.exception("BV振分変更画面のBV更新エラー")
            messages.error(request, f"BVの更新中にエラーが発生しました: {e}")

        return redirect(redirect_url)


class ApiUsersBvView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "api_users_bv.html"
    SORT_COLUMNS = {
        "id": "a.id",
        "doc_no": "a.doc_no",
        "member_no": "a.member_no",
        "firstname": "a.firstname",
        "order_type": "a.order_type",
        "order_year": "a.order_year",
        "order_month": "a.order_month",
        "price": "a.price",
        "total_bv": "a.total_bv",
        "payment_date": "a.payment_date",
        "is_posted": "a.is_posted",
        "choice_type": "a.choice_type",
        "desc": "a.`desc`",
        "created_by": "a.created_by",
        "post_by": "a.post_by",
    }

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="id",
            default_direction="desc",
        )

    def _build_where(
        self,
        q_doc_no="",
        q_member_no="",
        q_name="",
        q_order_type="",
        q_order_year="",
        q_order_month="",
        q_is_posted="",
        q_payment_from="",
        q_payment_to="",
    ):
        where = []
        params = []

        if q_doc_no:
            where.append("a.doc_no LIKE %s")
            params.append(f"%{q_doc_no}%")

        if q_member_no:
            where.append("a.member_no LIKE %s")
            params.append(f"%{q_member_no}%")

        if q_name:
            where.append("a.firstname LIKE %s")
            params.append(f"%{q_name}%")

        if q_order_type:
            where.append("a.order_type = %s")
            params.append(q_order_type)

        if q_order_year:
            where.append("a.order_year = %s")
            params.append(q_order_year)

        if q_order_month:
            where.append("a.order_month = %s")
            params.append(q_order_month)

        if q_is_posted:
            where.append("a.is_posted = %s")
            params.append(q_is_posted)

        if q_payment_from:
            where.append("a.payment_date >= %s")
            params.append(q_payment_from)

        if q_payment_to:
            where.append("a.payment_date < DATE_ADD(%s, INTERVAL 1 DAY)")
            params.append(q_payment_to)

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return where_sql, params

    def _fetch_total_count(self, **filters):
        where_sql, params = self._build_where(**filters)

        sql = f"""
            SELECT COUNT(*)
            FROM nexus_production.api_users_bv AS a
            {where_sql}
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()

        return int(row[0]) if row else 0

    def _fetch_rows(self, limit=200, offset=0, order_sql="", **filters):
        where_sql, params = self._build_where(**filters)

        sql = f"""
            SELECT
                a.id,
                a.price,
                a.total_bv,
                a.order_type,
                a.doc_no,
                a.firstname,
                a.member_no,
                a.order_year,
                a.order_month,
                a.payment_date,
                a.is_posted,
                a.`desc`,
                a.choice_type,
                a.created_by,
                a.post_by
            FROM nexus_production.api_users_bv AS a
            {where_sql}
            ORDER BY {order_sql or "a.id DESC"}, a.id DESC
            LIMIT %s OFFSET %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params + [limit, offset])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        filters = {
            "q_doc_no": (self.request.GET.get("q_doc_no") or "").strip(),
            "q_member_no": (self.request.GET.get("q_member_no") or "").strip(),
            "q_name": (self.request.GET.get("q_name") or "").strip(),
            "q_order_type": (self.request.GET.get("q_order_type") or "").strip(),
            "q_order_year": (self.request.GET.get("q_order_year") or "").strip(),
            "q_order_month": (self.request.GET.get("q_order_month") or "").strip(),
            "q_is_posted": (self.request.GET.get("q_is_posted") or "").strip(),
            "q_payment_from": (self.request.GET.get("q_payment_from") or "").strip(),
            "q_payment_to": (self.request.GET.get("q_payment_to") or "").strip(),
        }

        per_page = self.get_per_page()
        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        total_count = self._fetch_total_count(**filters)
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            limit=per_page,
            offset=offset,
            order_sql=sort_ctx["order_sql"],
            **filters,
        )

        ctx.update(filters)

        base_params = {k: v for k, v in filters.items() if v}

        if per_page != self.DEFAULT_PER_PAGE:
            base_params["per_page"] = per_page

        add_sort_params(base_params, sort_ctx)

        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params=base_params,
        )


class WeekBonusView(generic.ListView):
    template_name = "week_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster
    PRE_REGISTER_OPTIONS = (
        ("drive", "ドライブボーナス"),
        ("basic", "ベーシックボーナス"),
        ("matching", "マッチングボーナス"),
    )

    def get_queryset(self):
        return PeriodMaster.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "week_bonus")

        if action != "register_week_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:week_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:week_bonus")

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:week_bonus")

        try:
            rows = self._get_week_bonus_rows(selected_kibetu, period)

            if not rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "week_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/week_bonus/?kibetu={selected_kibetu}")

            insert_sql = """
                INSERT INTO bonus_db.B_week_bonus_result (
                    kibetu,
                    jwoa_code,
                    jwoa_name,
                    drive_bonus,
                    basic_bonus,
                    matching_bonus,
                    week_bonus,
                    created_at,
                    updated_at
                ) VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
                )
                ON DUPLICATE KEY UPDATE
                    jwoa_name = VALUES(jwoa_name),
                    drive_bonus = VALUES(drive_bonus),
                    basic_bonus = VALUES(basic_bonus),
                    matching_bonus = VALUES(matching_bonus),
                    week_bonus = VALUES(week_bonus),
                    updated_at = CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
            """

            insert_params = []
            for r in rows:
                insert_params.append([
                    r.get("期別"),
                    r.get("会員番号") or "",
                    r.get("会員名") or "",
                    r.get("ドライブボーナス") or 0,
                    r.get("ベーシックボーナス") or 0,
                    r.get("マッチングボーナス") or 0,
                    r.get("週間ボーナス") or 0,
                ])

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.executemany(insert_sql, insert_params)

                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "week_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(rows)}件登録"
                        ]
                    )

            messages.success(request, f"{len(rows)}件を週ボーナス結果に登録しました。")

        except Exception as e:
            logger.exception("週ボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/week_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        pre_register_targets = self._get_pre_register_targets()
        ctx["selected_kibetu"] = selected_kibetu
        ctx["pre_register_targets"] = pre_register_targets
        ctx["pre_register_options"] = self.PRE_REGISTER_OPTIONS
        ctx["history_rows"] = get_week_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:week_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            return ctx

        ctx["selected_period"] = period
        if not self._pre_register_selected_bonuses(
            selected_kibetu,
            period,
            pre_register_targets,
        ):
            return ctx

        ctx["rows"] = self._get_week_bonus_rows(selected_kibetu, period)
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "week_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_pre_register_targets(self):
        if "kibetu" not in self.request.GET:
            return [key for key, _label in self.PRE_REGISTER_OPTIONS]
        return [
            value
            for value in self.request.GET.getlist("pre_register")
            if value in {key for key, _label in self.PRE_REGISTER_OPTIONS}
        ]

    def _pre_register_selected_bonuses(self, selected_kibetu, period, targets):
        if not targets:
            return True

        registered_labels = []
        warning_labels = []
        target_bonus_names = []
        current_label = "事前登録"
        current_key = ""
        started_at = get_rds_jst_now()

        try:
            if "drive" in targets:
                current_label = "ドライブボーナス"
                current_key = "drive_bonus"
                target_bonus_names.append(current_key)
                count = self._register_drive_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("ドライブ")
                else:
                    warning_labels.append("ドライブ")

            if "basic" in targets:
                current_label = "ベーシックボーナス"
                current_key = "basic_bonus"
                target_bonus_names.append(current_key)
                count = self._register_basic_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("ベーシック")
                else:
                    warning_labels.append("ベーシック")

            if "matching" in targets:
                current_label = "マッチングボーナス"
                current_key = "matching_bonus"
                target_bonus_names.append(current_key)
                count = self._register_matching_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("マッチング")
                else:
                    warning_labels.append("マッチング")

        except Exception as e:
            history_deleted = delete_auto_register_history_after(
                selected_kibetu,
                target_bonus_names,
                self.request.user.username,
                started_at,
            )
            logger.exception(
                "週ボーナス表示前の事前登録エラー: kibetu=%s bonus_key=%s bonus_label=%s",
                selected_kibetu,
                current_key,
                current_label,
            )
            cleanup_message = (
                "登録履歴を削除しました"
                if history_deleted
                else "登録履歴の削除に失敗しました。管理者に確認してください"
            )
            messages.error(
                self.request,
                f"{current_label}の事前登録中にエラーが発生しました。{cleanup_message}: {e}",
            )
            return False

        if registered_labels or warning_labels:
            message_parts = []
            if registered_labels:
                message_parts.append(
                    f"登録済み: {'・'.join(registered_labels)}"
                )
            if warning_labels:
                message_parts.append(
                    f"対象データなし: {'・'.join(warning_labels)}"
                )
            message_text = "事前登録結果 - " + " / ".join(message_parts)
            if warning_labels:
                messages.warning(self.request, message_text)
            else:
                messages.success(self.request, message_text)

        return True

    def _insert_auto_register_history(self, cursor, bonus_name, selected_kibetu, count):
        cursor.execute(
            """
                INSERT INTO bonus_db.bonus_register_history (
                    bonus_name,
                    kibetu,
                    registered_at,
                    registered_by,
                    comment_text
                )
                VALUES (
                    %s,
                    %s,
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                    %s,
                    %s
                )
            """,
            [
                bonus_name,
                selected_kibetu,
                self.request.user.username,
                f"週ボーナス表示前の自動登録: {count}件登録",
            ],
        )

    def _register_drive_bonus(self, selected_kibetu, period):
        if not ensure_week_purchase_info(self.request, selected_kibetu, period):
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "drive_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        if not ensure_user_target_rank_for_kibetu(self.request, selected_kibetu):
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "drive_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        rows = DriveBonusView()._get_drive_bonus_rows(selected_kibetu, period)
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "drive_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        insert_sql = """
            INSERT INTO bonus_db.B_drive_bonus_result (
                kibetu,
                title_name,
                introducer_code,
                jwoa_code,
                jwoa_name,
                rate,
                sum_bv,
                sum_bonus_amount,
                created_at
            ) VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                NOW()
            )
            ON DUPLICATE KEY UPDATE
                title_name = VALUES(title_name),
                introducer_code = VALUES(introducer_code),
                jwoa_name = VALUES(jwoa_name),
                rate = VALUES(rate),
                sum_bv = VALUES(sum_bv),
                sum_bonus_amount = VALUES(sum_bonus_amount),
                created_at = NOW()
        """
        insert_params = [
            [
                selected_kibetu,
                r.get("title_name") or "",
                r.get("introducer_code") or "",
                r.get("jwoa_code") or "",
                r.get("jwoa_name") or "",
                r.get("rate") or 0,
                r.get("sum_bv") or 0,
                r.get("sum_bonus_amount") or 0,
            ]
            for r in rows
        ]

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "drive_bonus",
                    selected_kibetu,
                    len(rows),
                )

        return len(rows)

    def _register_basic_bonus(self, selected_kibetu, period):
        basic_view = BasicBonusView()
        basic_bonus_rows = basic_view._get_basic_bonus_rows(selected_kibetu, period)
        basic_bv_line_rows = basic_view._get_basic_bv_line_rows(selected_kibetu, period)
        if not basic_bonus_rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "basic_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        (
            delete_bonus_result_sql,
            delete_bv_line_sql,
            delete_params,
            insert_sql,
            insert_params,
            basic_bv_line_insert_sql,
            basic_bv_line_insert_params,
        ) = register_sql.get_basic_bonus_delete_insert_data(
            selected_kibetu,
            basic_bonus_rows,
            basic_bv_line_rows,
        )

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                logger.info(
                    "ベーシックボーナス登録前削除SQLを実行します。kibetu=%s",
                    selected_kibetu,
                )
                cursor.execute(delete_bonus_result_sql, delete_params)
                cursor.execute(delete_bv_line_sql, delete_params)

                logger.info(
                    "ベーシックボーナス登録INSERT SQLを実行します。kibetu=%s",
                    selected_kibetu,
                )
                cursor.executemany(insert_sql, insert_params)
                if basic_bv_line_insert_params:
                    cursor.executemany(
                        basic_bv_line_insert_sql,
                        basic_bv_line_insert_params,
                    )
                self._insert_auto_register_history(
                    cursor,
                    "basic_bonus",
                    selected_kibetu,
                    len(basic_bonus_rows),
                )

        return len(basic_bonus_rows)

    def _register_matching_bonus(self, selected_kibetu, period):
        rows = MatchingBonusView()._get_basic_bonus_rows(selected_kibetu, period)
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "matching_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        insert_sql = """
            INSERT INTO bonus_db.B_matching_bonus_result (
                kibetu,
                introducer_code,
                introducer_name,
                line_code,
                basic_code,
                basic_name,
                level,
                tree_level,
                active_count,
                basic_bonus,
                matching_bonus,
                created_at
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
            )
            ON DUPLICATE KEY UPDATE
                introducer_name = VALUES(introducer_name),
                basic_name      = VALUES(basic_name),
                level           = VALUES(level),
                tree_level      = VALUES(tree_level),
                active_count    = VALUES(active_count),
                basic_bonus     = VALUES(basic_bonus),
                matching_bonus  = VALUES(matching_bonus),
                created_at      = NOW()
        """
        insert_params = [
            [
                selected_kibetu,
                r.get("introducer_code") or "",
                r.get("introducer_name") or "",
                r.get("line_code") or "",
                r.get("basic_code") or "",
                r.get("basic_name") or "",
                r.get("level") or 0,
                r.get("tree_level") or 0,
                r.get("active_count") or 0,
                r.get("basic_bonus") or 0,
                r.get("matching_bonus") or 0,
            ]
            for r in rows
        ]

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "matching_bonus",
                    selected_kibetu,
                    len(rows),
                )

        return len(rows)

    def _get_week_bonus_rows(self, selected_kibetu, period):
        params = [
            selected_kibetu,
            selected_kibetu,
            selected_kibetu,
            selected_kibetu,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(WEEK_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows


class S_WeekBonusView(generic.ListView):
    template_name = "s_week_bonus.html"
    context_object_name = "object_list"
    model = PeriodMaster
    ALL_KIBETU_VALUE = "__all__"

    def get_queryset(self):
        # B_week_bonus_result に登録済みの期別だけ取得
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_week_bonus_result
                ORDER BY kibetu DESC
            """)
            registered_kibetu_list = [row[0] for row in cursor.fetchall()]

        if not registered_kibetu_list:
            return PeriodMaster.objects.using("rds").none()

        return (
            PeriodMaster.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-st_date", "-kibetu")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            if kibetu == self.ALL_KIBETU_VALUE:
                kibetu = "all"
            filename = build_bonus_export_filename("week_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["week_bonus"],
                "WeekBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        jwoa_code = (self.request.GET.get("jwoa_code") or "").strip()
        jwoa_name = (self.request.GET.get("jwoa_name") or "").strip()

        # 期別未選択なら、登録済み期別の先頭を自動選択
        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["all_kibetu_value"] = self.ALL_KIBETU_VALUE
        ctx["is_all_kibetu"] = selected_kibetu == self.ALL_KIBETU_VALUE
        ctx["jwoa_code"] = jwoa_code
        ctx["jwoa_name"] = jwoa_name
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        if ctx["is_all_kibetu"]:
            if not jwoa_code and not jwoa_name:
                ctx["all_kibetu_requires_member_filter"] = True
                return ctx
        else:
            period = PeriodMaster.objects.using("rds").filter(kibetu=selected_kibetu).first()
            if not period:
                return ctx
            ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "jwoa_code": "jwoa_code",
                "jwoa_name": "jwoa_name",
                "drive_bonus": "drive_bonus",
                "basic_bonus": "basic_bonus",
                "matching_bonus": "matching_bonus",
                "week_bonus": "week_bonus",
                "created_at": "created_at",
                "updated_at": "updated_at",
            },
            default_sort="week_bonus",
            default_direction="desc",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                id,
                kibetu,
                jwoa_code,
                jwoa_name,
                drive_bonus,
                basic_bonus,
                matching_bonus,
                week_bonus,
                created_at,
                updated_at
            FROM bonus_db.B_week_bonus_result
            WHERE 1 = 1
        """

        params = []
        if not ctx["is_all_kibetu"]:
            sql += "\n            AND kibetu = %s"
            params.append(selected_kibetu)

        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "jwoa_code": "jwoa_code",
                "jwoa_name": "jwoa_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx


class S_MonthTitleView(generic.ListView):
    template_name = "s_month_title.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.month_title
                ORDER BY kibetu DESC
            """)
            registered_kibetu_list = [row[0] for row in cursor.fetchall()]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get_title_options(self):
        title_options = []
        seen_ids = set()
        for title in TitleMaster.objects.using("rds").order_by("title_id"):
            title_id = str(title.title_id)
            if title_id in seen_ids:
                continue
            seen_ids.add(title_id)
            title_options.append(
                {
                    "title_id": title_id,
                    "title_name": title.title_name or "タイトルなし",
                }
            )
        if "0" not in seen_ids:
            title_options.insert(0, {"title_id": "0", "title_name": "タイトルなし"})
        return title_options

    def _parse_selected_title_ids(self):
        raw_ids = [
            (value or "").strip()
            for value in self.request.GET.getlist("title_id")
        ]
        selected = []
        seen = set()
        for value in raw_ids:
            if not value or not value.isdigit() or value in seen:
                continue
            seen.add(value)
            selected.append(value)
        return selected

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            filename = build_bonus_export_filename("month_title_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["month_title"],
                "MonthTitleResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        selected_title_ids = self._parse_selected_title_ids()
        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["selected_title_ids"] = selected_title_ids
        # 後方互換（テンプレート・旧URL）
        ctx["selected_title_id"] = (
            selected_title_ids[0] if len(selected_title_ids) == 1 else ""
        )
        ctx["title_options"] = self.get_title_options()
        ctx["rows"] = []
        ctx["selected_period"] = None
        ctx["selected_title_labels"] = [
            option["title_name"]
            for option in ctx["title_options"]
            if option["title_id"] in selected_title_ids
        ]
        ctx["selected_title_label"] = "、".join(ctx["selected_title_labels"])

        if not selected_kibetu:
            return ctx

        period = MonthlyPeriod.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            return ctx

        ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "mt.kibetu",
                "jwoa_code": "mt.jwoa_code",
                "jwoa_name": "mt.jwoa_name",
                "income_line_bv": "mt.income_line_bv",
                "basic_line_bv": "mt.basic_line_bv",
                "title_id": "mt.title_id",
                "title_name": "title_name",
                "updated_at": "mt.updated_at",
            },
            default_sort="jwoa_code",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                mt.id,
                mt.kibetu,
                mt.jwoa_code,
                mt.jwoa_name,
                mt.income_line_bv,
                mt.basic_line_bv,
                mt.title_id,
                COALESCE(tm.title_name, 'タイトルなし') AS title_name,
                mt.created_at,
                mt.updated_at
            FROM bonus_db.month_title AS mt
            LEFT JOIN bonus_db.title_master AS tm
              ON mt.title_id = tm.title_id
            WHERE mt.kibetu = %s
        """

        params = [selected_kibetu]
        if selected_title_ids:
            placeholders = ", ".join(["%s"] * len(selected_title_ids))
            sql += f"\n            AND mt.title_id IN ({placeholders})"
            params.extend(selected_title_ids)

        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "jwoa_code": "mt.jwoa_code",
                "jwoa_name": "mt.jwoa_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx



class MonthTitleView(generic.ListView):
    template_name = "month_title.html"
    context_object_name = "object_list"
    model = MonthlyPeriod

    def get_queryset(self):
        return MonthlyPeriod.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "month_title")

        if action != "register_month_title":
            messages.error(request, "不正な操作です。")
            return redirect("connect:month_title")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:month_title")

        period = MonthlyPeriod.objects.using("rds").filter(kibetu=selected_kibetu).first()
        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:month_title")

        try:
            rows = self._get_month_title_rows(period)

            if not rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "month_title",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/month_title/?kibetu={selected_kibetu}")

            delete_sql, delete_params, insert_sql, insert_params = (
                register_sql.get_month_title_delete_insert_data(
                    selected_kibetu,
                    rows,
                )
            )

            title_registration_status = ""

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    logger.info("月タイトル登録前削除SQLを実行します。kibetu=%s", selected_kibetu)
                    cursor.execute(delete_sql, delete_params)

                    logger.info("月タイトル登録INSERT SQLを実行します。kibetu=%s", selected_kibetu)
                    cursor.executemany(insert_sql, insert_params)

                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "month_title",
                            selected_kibetu,
                            request.user.username,
                            f"既存データ削除後、{len(rows)}件登録",
                        ],
                    )

                title_registration = TitleRegistrationView()
                if title_registration._exists_data(period.year, period.month):
                    title_registration_status = "skipped"
                    logger.info(
                        "タイトルユーザー登録は登録済みのためスキップします。year=%s month=%s",
                        period.year,
                        period.month,
                    )
                    insert_bonus_register_history(
                        "title_registration",
                        selected_kibetu,
                        request.user.username,
                        "月タイトル登録後の自動登録: 登録済みのため更新スキップ",
                    )
                else:
                    logger.info(
                        "月タイトル登録後にタイトルユーザー登録を実行します。kibetu=%s year=%s month=%s",
                        selected_kibetu,
                        period.year,
                        period.month,
                    )
                    title_registration_rows = title_registration._fetch_rows(
                        period.year,
                        period.month,
                        selected_kibetu,
                    )
                    title_registration._insert_rows(period.year, period.month, selected_kibetu)
                    title_registration._update_title(period.year, period.month, selected_kibetu)
                    title_registration._update_setting(period.year, period.month)
                    insert_bonus_register_history(
                        "title_registration",
                        selected_kibetu,
                        request.user.username,
                        f"月タイトル登録後の自動登録: {len(title_registration_rows)}件登録",
                    )
                    title_registration_status = "registered"

            if title_registration_status == "registered":
                messages.success(
                    request,
                    f"{len(rows)}件を月タイトル結果に登録し、タイトルユーザー登録も実行しました。",
                )
            elif title_registration_status == "skipped":
                messages.success(
                    request,
                    f"{len(rows)}件を月タイトル結果に登録しました。最高ピンタイトルの登録は登録済みのためスキップしました。",
                )
            else:
                messages.success(
                    request,
                    f"{len(rows)}件を月タイトル結果に登録しました。",
                )
            return redirect(f"/month_title/?kibetu={selected_kibetu}")

        except Exception as e:
            logger.exception("月タイトル登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")
            return redirect(f"/month_title/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = self.request.GET.get("kibetu", "").strip()
        ctx["selected_kibetu"] = selected_kibetu
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:month_title"
        ctx["selected_period"] = None
        ctx["rows"] = []

        if not selected_kibetu:
            return ctx

        period = MonthlyPeriod.objects.using("rds").filter(kibetu=selected_kibetu).first()
        ctx["selected_period"] = period

        if not period:
            return ctx

        ctx["rows"] = self._get_month_title_rows(period)
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "month_title",
                selected_kibetu,
            )

        return ctx

    def _get_month_title_rows(self, period):
        params = [
            period.year,
            period.month,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(MONTH_TITLE_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows


class MonthBonusView(generic.ListView):
    template_name = "month_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod
    PRE_REGISTER_OPTIONS = (
        ("month_title", "月タイトル"),
        ("title", "タイトルボーナス"),
        ("title_diff", "差額ボーナス"),
        ("repurchase_over", "再購入オーバーボーナス"),
        ("three_star", "3スター配当"),
    )

    def get_queryset(self):
        return MonthlyPeriod.objects.using("rds").all()

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        selected_kibetu = request.POST.get("kibetu", "").strip()

        if action == "delete":
            return delete_bonus_result_for_kibetu(request, "month_bonus")

        if action != "register_month_bonus":
            messages.error(request, "不正な操作です。")
            return redirect("connect:month_bonus")

        if not selected_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:month_bonus")

        period = MonthlyPeriod.objects.using("rds").filter(kibetu=selected_kibetu).first()

        if not period:
            messages.error(request, "選択された期別が存在しません。")
            return redirect("connect:month_bonus")

        if not has_month_title_rows(selected_kibetu):
            messages.warning(
                request,
                "月ボーナスを登録するには、同じ期別の月タイトルを先に計算・登録してください。",
            )
            return redirect(f"/month_bonus/?kibetu={selected_kibetu}")

        try:
            rows = self._get_month_bonus_rows(selected_kibetu)

            if not rows:
                with transaction.atomic(using="rds"):
                    insert_bonus_register_history(
                        "month_bonus",
                        selected_kibetu,
                        request.user.username,
                        "0件登録（対象データなし）",
                    )
                messages.warning(request, "登録対象データはありませんが、登録履歴を残しました。")
                return redirect(f"/month_bonus/?kibetu={selected_kibetu}")

            insert_sql = """
                INSERT INTO bonus_db.B_month_bonus_result (
                    kibetu,
                    jwoa_code,
                    jwoa_name,
                    title_bonus,
                    repurchase_over_bonus,
                    title_diff_bonus,
                    three_star_diamond_global_bonus,
                    crown_three_star_diamond_global_bonus,
                    month_bonus,
                    created_at,
                    updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
                )
                ON DUPLICATE KEY UPDATE
                    jwoa_name = VALUES(jwoa_name),
                    title_bonus = VALUES(title_bonus),
                    repurchase_over_bonus = VALUES(repurchase_over_bonus),
                    title_diff_bonus = VALUES(title_diff_bonus),
                    three_star_diamond_global_bonus = VALUES(three_star_diamond_global_bonus),
                    crown_three_star_diamond_global_bonus = VALUES(crown_three_star_diamond_global_bonus),
                    month_bonus = VALUES(month_bonus),
                    updated_at = CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
            """

            insert_params = []

            for r in rows:
                insert_params.append([
                    r.get("kibetu"),
                    r.get("jwoa_code") or "",
                    r.get("jwoa_name") or "",
                    r.get("title_bonus") or 0,
                    r.get("repurchase_over_bonus") or 0,
                    r.get("title_diff_bonus") or 0,
                    r.get("three_star_diamond_global_bonus") or 0,
                    r.get("crown_three_star_diamond_global_bonus") or 0,
                    r.get("month_bonus") or 0,
                ])

            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    cursor.executemany(insert_sql, insert_params)

                    history_sql = """
                        INSERT INTO bonus_db.bonus_register_history (
                            bonus_name,
                            kibetu,
                            registered_at,
                            registered_by,
                            comment_text
                        )
                        VALUES (
                            %s,
                            %s,
                            CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                            %s,
                            %s
                        )
                    """

                    cursor.execute(
                        history_sql,
                        [
                            "month_bonus",
                            selected_kibetu,
                            request.user.username,
                            f"{len(rows)}件登録"
                        ]
                    )

            messages.success(request, f"{len(rows)}件を月ボーナス結果に登録しました。")

        except Exception as e:
            logger.exception("月ボーナス結果登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {e}")

        return redirect(f"/month_bonus/?kibetu={selected_kibetu}")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        pre_register_targets = self._get_pre_register_targets()
        ctx["selected_kibetu"] = selected_kibetu
        ctx["pre_register_targets"] = pre_register_targets
        ctx["pre_register_options"] = self.PRE_REGISTER_OPTIONS
        ctx["history_rows"] = get_month_bonus_history_rows()
        ctx["history_target_url_name"] = "connect:month_bonus"
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        period = MonthlyPeriod.objects.using("rds").filter(kibetu=selected_kibetu).first()

        if not period:
            return ctx

        ctx["selected_period"] = period
        if not ensure_kibetu_purchase_info(self.request, selected_kibetu, period):
            insert_empty_bonus_history_on_display(
                self.request,
                "month_bonus",
                selected_kibetu,
            )
            return ctx

        if (
            "month_title" not in pre_register_targets
            and not has_month_title_rows(selected_kibetu)
        ):
            messages.warning(
                self.request,
                "月ボーナスを計算するには、同じ期別の月タイトルを先に計算・登録してください。",
            )
            return ctx

        if not self._pre_register_selected_bonuses(
            selected_kibetu,
            period,
            pre_register_targets,
        ):
            return ctx

        if not has_month_title_rows(selected_kibetu):
            return ctx

        ctx["rows"] = self._get_month_bonus_rows(selected_kibetu)
        if not ctx["rows"]:
            insert_empty_bonus_history_on_display(
                self.request,
                "month_bonus",
                selected_kibetu,
            )

        return ctx

    def _get_pre_register_targets(self):
        if "kibetu" not in self.request.GET:
            return [key for key, _label in self.PRE_REGISTER_OPTIONS]
        return [
            value
            for value in self.request.GET.getlist("pre_register")
            if value in {key for key, _label in self.PRE_REGISTER_OPTIONS}
        ]

    def _pre_register_selected_bonuses(self, selected_kibetu, period, targets):
        if not targets:
            return True

        registered_labels = []
        warning_labels = []
        target_bonus_names = []
        current_label = "事前登録"
        current_key = ""
        started_at = get_rds_jst_now()

        try:
            if "month_title" in targets:
                current_label = "月タイトル"
                current_key = "month_title"
                target_bonus_names.append(current_key)
                count = self._register_month_title(selected_kibetu, period)
                if count:
                    registered_labels.append("月タイトル")
                else:
                    warning_labels.append("月タイトル")

            if "title" in targets:
                current_label = "タイトルボーナス"
                current_key = "title_bonus"
                target_bonus_names.append(current_key)
                count = self._register_title_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("タイトル")
                else:
                    warning_labels.append("タイトル")

            if "title_diff" in targets:
                current_label = "差額ボーナス"
                current_key = "title_diff_bonus"
                target_bonus_names.append(current_key)
                count = self._register_title_diff_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("差額")
                else:
                    warning_labels.append("差額")

            if "repurchase_over" in targets:
                current_label = "再購入オーバーボーナス"
                current_key = "repurchase_over_bonus"
                target_bonus_names.append(current_key)
                count = self._register_repurchase_over_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("再購入オーバー")
                else:
                    warning_labels.append("再購入オーバー")

            if "three_star" in targets:
                current_label = "3スター配当"
                current_key = "three_star_global_bonus"
                target_bonus_names.append(current_key)
                count = self._register_three_star_global_bonus(selected_kibetu, period)
                if count:
                    registered_labels.append("3スター")
                else:
                    warning_labels.append("3スター")

        except Exception as e:
            history_deleted = delete_auto_register_history_after(
                selected_kibetu,
                target_bonus_names,
                self.request.user.username,
                started_at,
            )
            logger.exception(
                "月ボーナス表示前の事前登録エラー: kibetu=%s bonus_key=%s bonus_label=%s",
                selected_kibetu,
                current_key,
                current_label,
            )
            cleanup_message = (
                "登録履歴を削除しました"
                if history_deleted
                else "登録履歴の削除に失敗しました。管理者に確認してください"
            )
            messages.error(
                self.request,
                f"{current_label}の事前登録中にエラーが発生しました。{cleanup_message}: {e}",
            )
            return False

        if registered_labels or warning_labels:
            message_parts = []
            if registered_labels:
                message_parts.append(
                    f"登録済み: {'・'.join(registered_labels)}"
                )
            if warning_labels:
                message_parts.append(
                    f"対象データなし: {'・'.join(warning_labels)}"
                )
            message_text = "事前登録結果 - " + " / ".join(message_parts)
            if warning_labels:
                messages.warning(self.request, message_text)
            else:
                messages.success(self.request, message_text)

        return True

    def _insert_auto_register_history(self, cursor, bonus_name, selected_kibetu, count):
        cursor.execute(
            """
                INSERT INTO bonus_db.bonus_register_history (
                    bonus_name,
                    kibetu,
                    registered_at,
                    registered_by,
                    comment_text
                )
                VALUES (
                    %s,
                    %s,
                    CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo'),
                    %s,
                    %s
                )
            """,
            [
                bonus_name,
                selected_kibetu,
                self.request.user.username,
                f"月ボーナス表示前の自動登録: {count}件登録",
            ],
        )

    def _register_month_title(self, selected_kibetu, period):
        rows = MonthTitleView()._get_month_title_rows(period)
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "month_title",
                        selected_kibetu,
                        0,
                    )
            return 0

        delete_sql, delete_params, insert_sql, insert_params = (
            register_sql.get_month_title_delete_insert_data(
                selected_kibetu,
                rows,
            )
        )

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                logger.info("月ボーナス事前登録: 月タイトル削除SQLを実行します。kibetu=%s", selected_kibetu)
                cursor.execute(delete_sql, delete_params)
                logger.info("月ボーナス事前登録: 月タイトルINSERT SQLを実行します。kibetu=%s", selected_kibetu)
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "month_title",
                    selected_kibetu,
                    len(rows),
                )

        return len(rows)

    def _register_title_bonus(self, selected_kibetu, period):
        rows = TitleBonusView()._get_title_bonus_rows(selected_kibetu, period)
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "title_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        insert_sql, insert_params = register_sql.get_title_bonus_insert_data(
            selected_kibetu,
            rows,
        )

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "title_bonus",
                    selected_kibetu,
                    len(rows),
                )

        return len(rows)

    def _register_title_diff_bonus(self, selected_kibetu, period):
        rows = TitleDiffBonusView()._get_title_diff_bonus_rows(selected_kibetu, period)
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "title_diff_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        insert_sql, insert_params = register_sql.get_title_diff_bonus_insert_data(
            selected_kibetu,
            rows,
        )

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "title_diff_bonus",
                    selected_kibetu,
                    len(rows),
                )

        return len(rows)

    def _register_repurchase_over_bonus(self, selected_kibetu, period):
        rows = RepurchaseOverBonusView()._get_repurchase_over_bonus_rows(
            selected_kibetu=selected_kibetu,
            period=period,
        )
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "repurchase_over_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        insert_sql, insert_params = register_sql.get_repurchase_over_bonus_insert_data(
            selected_kibetu,
            rows,
        )
        if not insert_params:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "repurchase_over_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "repurchase_over_bonus",
                    selected_kibetu,
                    len(insert_params),
                )

        return len(insert_params)

    def _register_three_star_global_bonus(self, selected_kibetu, period):
        rows = ThreeStarGlobalBonusView()._get_three_star_global_bonus_rows(
            selected_kibetu=selected_kibetu,
            period=period,
        )
        if not rows:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "three_star_global_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        insert_sql, insert_params = register_sql.get_three_star_global_bonus_insert_data(
            selected_kibetu,
            rows,
        )
        if not insert_params:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    self._insert_auto_register_history(
                        cursor,
                        "three_star_global_bonus",
                        selected_kibetu,
                        0,
                    )
            return 0

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.executemany(insert_sql, insert_params)
                self._insert_auto_register_history(
                    cursor,
                    "three_star_global_bonus",
                    selected_kibetu,
                    len(insert_params),
                )

        return len(insert_params)

    def _get_month_bonus_rows(self, selected_kibetu):
        params = [
            selected_kibetu,
            selected_kibetu,
            selected_kibetu,
            selected_kibetu,
            selected_kibetu,
        ]

        with connections["rds"].cursor() as cursor:
            cursor.execute(MONTH_BONUS_SQL, params)
            logger.info(f"Executed SQL: {cursor._executed}")
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return rows



class S_MonthBonusView(generic.ListView):
    template_name = "s_month_bonus.html"
    context_object_name = "object_list"
    model = MonthlyPeriod
    ALL_KIBETU_VALUE = "__all__"

    def get_queryset(self):
        with connections["rds"].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT kibetu
                FROM bonus_db.B_month_bonus_result
                ORDER BY kibetu DESC
            """)
            registered_kibetu_list = [row[0] for row in cursor.fetchall()]

        if not registered_kibetu_list:
            return MonthlyPeriod.objects.using("rds").none()

        return (
            MonthlyPeriod.objects.using("rds")
            .filter(kibetu__in=registered_kibetu_list)
            .order_by("-year", "-month")
        )

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()

        if request.GET.get("export") == "excel":
            rows = context.get("rows", [])
            kibetu = context.get("selected_kibetu", "")
            if kibetu == self.ALL_KIBETU_VALUE:
                kibetu = "all"
            filename = build_bonus_export_filename("month_bonus_result", kibetu=kibetu)
            return export_search_rows_to_excel(
                rows,
                SEARCH_EXPORT_COLUMNS["month_bonus"],
                "MonthBonusResult",
                filename,
            )

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_kibetu = (self.request.GET.get("kibetu") or "").strip()
        jwoa_code = (self.request.GET.get("jwoa_code") or "").strip()
        jwoa_name = (self.request.GET.get("jwoa_name") or "").strip()

        if not selected_kibetu and self.object_list:
            selected_kibetu = self.object_list[0].kibetu

        ctx["selected_kibetu"] = selected_kibetu
        ctx["all_kibetu_value"] = self.ALL_KIBETU_VALUE
        ctx["is_all_kibetu"] = selected_kibetu == self.ALL_KIBETU_VALUE
        ctx["jwoa_code"] = jwoa_code
        ctx["jwoa_name"] = jwoa_name
        ctx["rows"] = []
        ctx["selected_period"] = None

        if not selected_kibetu:
            return ctx

        if ctx["is_all_kibetu"]:
            if not jwoa_code and not jwoa_name:
                ctx["all_kibetu_requires_member_filter"] = True
                return ctx
        else:
            period = (
                MonthlyPeriod.objects.using("rds")
                .filter(kibetu=selected_kibetu)
                .first()
            )
            if not period:
                return ctx
            ctx["selected_period"] = period

        sort_ctx = get_bonus_sort_context(
            self.request,
            {
                "kibetu": "kibetu",
                "jwoa_code": "jwoa_code",
                "jwoa_name": "jwoa_name",
                "title_bonus": "title_bonus",
                "repurchase_over_bonus": "repurchase_over_bonus",
                "title_diff_bonus": "title_diff_bonus",
                "three_star_diamond_global_bonus": "three_star_diamond_global_bonus",
                "crown_three_star_diamond_global_bonus": "crown_three_star_diamond_global_bonus",
                "month_bonus": "month_bonus",
                "created_at": "created_at",
                "updated_at": "updated_at",
            },
            default_sort="month_bonus",
            default_direction="desc",
        )
        ctx.update(sort_ctx)

        sql = """
            SELECT
                id,
                kibetu,
                jwoa_code,
                jwoa_name,
                title_bonus,
                repurchase_over_bonus,
                title_diff_bonus,
                three_star_diamond_global_bonus,
                crown_three_star_diamond_global_bonus,
                month_bonus,
                created_at,
                updated_at
            FROM bonus_db.B_month_bonus_result
            WHERE 1 = 1
        """

        params = []
        if not ctx["is_all_kibetu"]:
            sql += "\n            AND kibetu = %s"
            params.append(selected_kibetu)

        sql, filter_values = apply_like_filters(
            sql,
            params,
            self.request,
            {
                "jwoa_code": "jwoa_code",
                "jwoa_name": "jwoa_name",
            },
        )
        ctx.update(filter_values)
        sql += "\n            ORDER BY " + sort_ctx["order_sql"]

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        ctx["rows"] = rows

        return ctx



class BonusHistryView(generic.TemplateView):
    template_name = "bonus_histry.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = self._get_history_rows()
        return ctx

    def _get_history_rows(self):
        return get_week_bonus_history_rows()


def get_week_bonus_history_rows():
    from connect.sql.bonus_histry_sql import WEEK_BONUS_HISTORY_SQL

    logger.info("登録履歴（週）SQL実行")

    with connections["rds"].cursor() as cursor:
        cursor.execute(WEEK_BONUS_HISTORY_SQL)

        cols = [c[0] for c in cursor.description]
        rows = [
            dict(zip(cols, row))
            for row in cursor.fetchall()
        ]

    today = date.today()
    next_completion_date = min(
        (
            row["completion_date"]
            for row in rows
            if row["completion_date"] and row["completion_date"] >= today
        ),
        default=None,
    )
    for row in rows:
        row["is_next_completion_date"] = (
            row["completion_date"] == next_completion_date
        )

    return normalize_bonus_histry_rows(rows)


def get_month_bonus_history_rows():
    from connect.sql.bonus_histry_sql import MONTH_BONUS_HISTORY_SQL

    logger.info("登録履歴（月）SQL実行")

    with connections["rds"].cursor() as cursor:
        cursor.execute(MONTH_BONUS_HISTORY_SQL)

        cols = [c[0] for c in cursor.description]
        rows = [
            dict(zip(cols, row))
            for row in cursor.fetchall()
        ]

    previous_month = date.today() - relativedelta(months=1)
    for row in rows:
        row["is_previous_month"] = (
            row.get("year") == previous_month.year
            and row.get("month") == previous_month.month
        )

    return normalize_bonus_histry_rows(rows)


class BonusHistryMonthView(generic.TemplateView):
    template_name = "bonus_histry_month.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = self._get_history_rows()
        return ctx

    def _get_history_rows(self):
        return get_month_bonus_history_rows()


class CoolingOffView(generic.TemplateView):
    template_name = "cooling_off.html"
    SORT_COLUMNS = {
        "register_year": "p.register_year",
        "register_month": "p.register_month",
        "order_code": "c.order_code",
        "active_flag": "c.active_flag",
        "jwoa_code": "o.jwoa_code",
        "order_name": "o.order_name",
        "order_type": "o.order_type",
        "bv": "p.bv",
        "remarks": "c.remarks",
        "registered_by": "c.registered_by",
        "created_at": "c.created_at",
    }

    def _get_sort_context(self):
        return get_bonus_sort_context(
            self.request,
            self.SORT_COLUMNS,
            default_sort="created_at",
            default_direction="desc",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        detail_order_code = self.request.GET.get("detail_order_code")
        q_order_code = (self.request.GET.get("q_order_code") or "").strip()
        q_active_flag = (self.request.GET.get("q_active_flag") or "").strip()
        q_register_year = (self.request.GET.get("q_register_year") or "").strip()
        q_register_month = (self.request.GET.get("q_register_month") or "").strip()
        q_jwoa_code = (self.request.GET.get("q_jwoa_code") or "").strip()

        sort_ctx = self._get_sort_context()
        ctx.update(sort_ctx)

        ctx["rows"] = self._get_rows(
            q_order_code=q_order_code,
            q_active_flag=q_active_flag,
            q_register_year=q_register_year,
            q_register_month=q_register_month,
            q_jwoa_code=q_jwoa_code,
            order_sql=sort_ctx["order_sql"],
        )
        ctx["detail_order"] = None
        ctx["q_order_code"] = q_order_code
        ctx["q_active_flag"] = q_active_flag
        ctx["q_register_year"] = q_register_year
        ctx["q_register_month"] = q_register_month
        ctx["q_jwoa_code"] = q_jwoa_code

        if detail_order_code:
            ctx["detail_order"] = self._get_order_detail(detail_order_code)

        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        next_query = (request.POST.get("next_query") or "").strip()
        base_url = redirect("connect:cooling_off").url
        redirect_target = f"{base_url}?{next_query}" if next_query else base_url

        try:
            if action == "create":
                if self._create(request):
                    messages.success(request, "クーリングオフを登録しました。")
            elif action == "update":
                if self._update(request):
                    messages.success(request, "クーリングオフを更新しました。")
            elif action == "bulk_update":
                if self._bulk_update(request):
                    pass
            elif action == "delete":
                if self._delete(request):
                    messages.success(request, "クーリングオフを削除しました。")
            else:
                messages.error(request, "不正な操作です。")
        except IntegrityError:
            logger.exception("クーリングオフの登録エラー")
            messages.error(
                request,
                "クーリングオフの登録中にエラーが発生しました。注文番号を確認してください。",
            )
        except Exception as e:
            logger.exception("クーリングオフ操作エラー")
            messages.error(request, f"エラーが発生しました: {e}")

        return redirect(redirect_target)

    def _get_rows(
        self,
        q_order_code="",
        q_active_flag="",
        q_register_year="",
        q_register_month="",
        q_jwoa_code="",
        order_sql="",
    ):
        where = []
        params = []

        if q_order_code:
            where.append("c.order_code LIKE %s")
            params.append(f"%{q_order_code}%")

        if q_active_flag in ("0", "1"):
            where.append("c.active_flag = %s")
            params.append(int(q_active_flag))

        if q_register_year.isdigit():
            where.append("p.register_year = %s")
            params.append(int(q_register_year))

        if q_register_month.isdigit():
            where.append("p.register_month = %s")
            params.append(int(q_register_month))

        if q_jwoa_code:
            where.append("o.jwoa_code LIKE %s")
            params.append(f"%{q_jwoa_code}%")

        where_sql = f"WHERE {' AND '.join(where)}" if where else ""
        sql = """
            SELECT
                c.id,
                p.register_year,
                p.register_month,
                c.order_code,
                c.active_flag,
                c.remarks,
                c.registered_by,
                c.created_at,
                o.jwoa_code,
                o.order_name,
                o.order_type,
                p.bv
            FROM bonus_db.cooling_off c
            LEFT JOIN nexus_production.orders o
                ON c.order_code = o.order_code
            LEFT JOIN (
                SELECT
                    order_code,
                    MIN(register_year) AS register_year,
                    MIN(register_month) AS register_month,
                    SUM(IFNULL(bv, 0)) AS bv
                FROM bonus_db.purchase_info_list
                GROUP BY order_code
            ) p
                ON c.order_code = p.order_code
            {where_sql}
            ORDER BY {order_sql}, c.id DESC
        """.format(where_sql=where_sql, order_sql=order_sql or "c.created_at DESC")

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            cols = [c[0] for c in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]

    def _get_edit_row(self, edit_id):
        sql = """
            SELECT
                id,
                order_code,
                active_flag,
                remarks,
                registered_by,
                created_at
            FROM bonus_db.cooling_off
            WHERE id = %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [edit_id])
            row = cursor.fetchone()

            if not row:
                return None

            cols = [c[0] for c in cursor.description]
            return dict(zip(cols, row))

    def _get_order_detail(self, order_code):
        sql = """
            SELECT
                order_code,
                jwoa_code,
                order_name,
                order_type,
                order_status,
                total_bv,
                deposit_at,
                order_at
            FROM nexus_production.orders
            WHERE order_code = %s
        """

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, [order_code])
            row = cursor.fetchone()

            if not row:
                return None

            cols = [c[0] for c in cursor.description]
            return dict(zip(cols, row))

    def _set_purchase_info_cooling_off(self, cursor, order_code):
        sql = """
            UPDATE bonus_db.purchase_info_list
            SET
                order_type = 200,
                updated_at = CURRENT_TIMESTAMP
            WHERE order_code = %s
        """
        cursor.execute(sql, [order_code])

    def _restore_purchase_info_order_type(self, cursor, order_code):
        sql = """
            UPDATE bonus_db.purchase_info_list p
            JOIN nexus_production.orders o
              ON p.order_code = o.order_code
            SET
                p.order_type = o.order_type,
                p.updated_at = CURRENT_TIMESTAMP
            WHERE p.order_code = %s
        """
        cursor.execute(sql, [order_code])

    def _get_cooling_off_order_code(self, cursor, row_id):
        sql = """
            SELECT order_code
            FROM bonus_db.cooling_off
            WHERE id = %s
            FOR UPDATE
        """
        cursor.execute(sql, [row_id])
        row = cursor.fetchone()
        return row[0] if row else ""

    def _validate_cooling_off_order_code(self, order_code, exclude_id=None):
        if not order_code:
            return "注文番号を入力してください。"

        if not self._get_order_detail(order_code):
            return (
                f"注文番号 {order_code} は orders に存在しません。"
                "注文マスタへ登録されているか確認してください。"
            )

        sql = """
            SELECT id
            FROM bonus_db.cooling_off
            WHERE order_code = %s
        """
        params = [order_code]
        if exclude_id is not None:
            sql += " AND id <> %s"
            params.append(exclude_id)
        sql += " LIMIT 1"

        with connections["rds"].cursor() as cursor:
            cursor.execute(sql, params)
            if cursor.fetchone():
                return f"注文番号 {order_code} はすでにクーリングオフ登録済みです。"

        return None

    def _create(self, request):
        order_code = (request.POST.get("order_code") or "").strip()
        try:
            active_flag = int(request.POST.get("active_flag", 1))
        except (TypeError, ValueError):
            messages.error(request, "状態の値が不正です。")
            return False

        if active_flag not in (0, 1):
            messages.error(request, "状態の値が不正です。")
            return False

        error_message = self._validate_cooling_off_order_code(order_code)
        if error_message:
            messages.error(request, error_message)
            return False

        sql = """
            INSERT INTO bonus_db.cooling_off (
                order_code,
                active_flag,
                remarks,
                registered_by
            )
            VALUES (
                %s,
                %s,
                %s,
                %s
            )
        """

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                cursor.execute(
                    sql,
                    [
                        order_code,
                        active_flag,
                        request.POST.get("remarks"),
                        request.user.username,
                    ]
                )
                if active_flag == 1:
                    self._set_purchase_info_cooling_off(cursor, order_code)

        return True

    def _update(self, request):
        order_code = (request.POST.get("order_code") or "").strip()
        row_id = (request.POST.get("id") or "").strip()

        if not row_id:
            messages.error(request, "更新対象が不正です。")
            return False

        try:
            row_id = int(row_id)
            active_flag = int(request.POST.get("active_flag", 1))
        except (TypeError, ValueError):
            messages.error(request, "更新内容が不正です。")
            return False

        if active_flag not in (0, 1):
            messages.error(request, "状態の値が不正です。")
            return False

        error_message = self._validate_cooling_off_order_code(order_code, exclude_id=row_id)
        if error_message:
            messages.error(request, error_message)
            return False

        sql = """
            UPDATE bonus_db.cooling_off
            SET
                order_code = %s,
                active_flag = %s,
                remarks = %s,
                registered_by = %s
            WHERE id = %s
        """

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                old_order_code = self._get_cooling_off_order_code(cursor, row_id)
                if not old_order_code:
                    messages.error(request, "更新対象データがありません。")
                    return False

                cursor.execute(
                    sql,
                    [
                        order_code,
                        active_flag,
                        request.POST.get("remarks"),
                        request.user.username,
                        row_id,
                    ]
                )

                if old_order_code != order_code:
                    self._restore_purchase_info_order_type(cursor, old_order_code)

                if active_flag == 1:
                    self._set_purchase_info_cooling_off(cursor, order_code)
                else:
                    self._restore_purchase_info_order_type(cursor, order_code)

        return True

    def _bulk_update(self, request):
        raw_ids = request.POST.getlist("ids")
        try:
            active_flag = int(request.POST.get("active_flag", 1))
            row_ids = sorted({int(x) for x in raw_ids if str(x).strip().isdigit()})
        except (TypeError, ValueError):
            messages.error(request, "一括更新内容が不正です。")
            return False

        if active_flag not in (0, 1):
            messages.error(request, "状態の値が不正です。")
            return False

        if not row_ids:
            messages.error(request, "一括編集する行を選択してください。")
            return False

        remarks = request.POST.get("remarks")
        update_sql = """
            UPDATE bonus_db.cooling_off
            SET
                active_flag = %s,
                remarks = %s,
                registered_by = %s
            WHERE id = %s
        """

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                placeholders = ", ".join(["%s"] * len(row_ids))
                cursor.execute(
                    f"""
                    SELECT id, order_code
                    FROM bonus_db.cooling_off
                    WHERE id IN ({placeholders})
                    FOR UPDATE
                    """,
                    row_ids,
                )
                targets = cursor.fetchall()
                if not targets:
                    messages.error(request, "一括更新対象データがありません。")
                    return False

                for row_id, order_code in targets:
                    cursor.execute(
                        update_sql,
                        [
                            active_flag,
                            remarks,
                            request.user.username,
                            row_id,
                        ],
                    )
                    if active_flag == 1:
                        self._set_purchase_info_cooling_off(cursor, order_code)
                    else:
                        self._restore_purchase_info_order_type(cursor, order_code)

        messages.success(
            request,
            f"{len(targets)}件のクーリングオフを一括更新しました。",
        )
        return True

    def _delete(self, request):
        row_id = (request.POST.get("id") or "").strip()

        if not row_id:
            messages.error(request, "削除対象が不正です。")
            return False

        try:
            row_id = int(row_id)
        except (TypeError, ValueError):
            messages.error(request, "削除対象が不正です。")
            return False

        sql = """
            DELETE
            FROM bonus_db.cooling_off
            WHERE id = %s
        """

        with transaction.atomic(using="rds"):
            with connections["rds"].cursor() as cursor:
                order_code = self._get_cooling_off_order_code(cursor, row_id)
                if not order_code:
                    messages.warning(request, "削除対象データがありません。")
                    return False

                cursor.execute(sql, [row_id])
                self._restore_purchase_info_order_type(cursor, order_code)

        return True

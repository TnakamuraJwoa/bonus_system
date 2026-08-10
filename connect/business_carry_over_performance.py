import csv
import io
import logging
import math

import openpyxl
from django.contrib import messages
from django.core.cache import cache
from django.db import ProgrammingError, connections, transaction
from django.http import HttpResponse
from django.shortcuts import redirect
from django.views import generic

from connect.audit import record_change_audit

from .views import KeysetPaginationMixin


logger = logging.getLogger(__name__)

BASIC_BV_LINE_TABLE = "bonus_db.basic_bv_line"
BASIC_BV_LINE_IMPORT_COLUMNS = {
    "placement_code": "placement_code",
    "上位者コード": "placement_code",
    "上位者id": "placement_code",
    "jmoa_code": "jmoa_code",
    "会員コード": "jmoa_code",
    "会員id": "jmoa_code",
    "bv": "bv",
    "carry_over_bv": "carry_over_bv",
    "繰り越しbv": "carry_over_bv",
    "繰越しbv": "carry_over_bv",
    "繰越bv": "carry_over_bv",
    "kibetu": "kibetu",
    "期別": "kibetu",
}
REQUIRED_IMPORT_COLUMNS = ["placement_code", "jmoa_code", "bv", "carry_over_bv", "kibetu"]

# basic_bv_line は数百万件あり、期別ごとの MIN/MAX(created_at) 集計は全件走査になる。
# 登録履歴モーダル用の集計結果はキャッシュし、取込・削除のタイミングで破棄する。
CARRY_OVER_HISTORY_CACHE_KEY = "business_carry_over_performance:history_rows"
CARRY_OVER_HISTORY_CACHE_TIMEOUT = 600


def is_missing_table_error(exc):
    return bool(exc.args and exc.args[0] == 1146)


def normalize_import_header(value):
    return str(value or "").strip().lower().replace(" ", "").replace("　", "")


def parse_import_int(value, row_number, column_label, errors):
    if value is None or str(value).strip() == "":
        return 0

    text = str(value).strip().replace(",", "")
    try:
        return int(float(text))
    except ValueError:
        errors.append(f"{row_number}行目: {column_label} は数値で入力してください。")
        return 0


def iter_uploaded_rows(uploaded_file):
    filename = (uploaded_file.name or "").lower()

    if filename.endswith((".xlsx", ".xlsm")):
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))

    raw = uploaded_file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp932")
    return list(csv.reader(io.StringIO(text)))


def parse_basic_bv_line_import_rows(uploaded_file):
    source_rows = iter_uploaded_rows(uploaded_file)
    errors = []
    data_rows = []

    header_row = None
    header_row_number = 0
    for index, row in enumerate(source_rows, start=1):
        if any(str(cell or "").strip() for cell in row):
            header_row = row
            header_row_number = index
            break

    if not header_row:
        return [], ["取込ファイルにヘッダー行がありません。"]

    column_map = {}
    for index, header in enumerate(header_row):
        normalized_header = normalize_import_header(header)
        column_name = BASIC_BV_LINE_IMPORT_COLUMNS.get(normalized_header)
        if column_name:
            column_map[column_name] = index

    missing_columns = [
        column_name
        for column_name in REQUIRED_IMPORT_COLUMNS
        if column_name not in column_map
    ]
    if missing_columns:
        return [], [f"必須列が不足しています: {', '.join(missing_columns)}"]

    for row_number, row in enumerate(source_rows[header_row_number:], start=header_row_number + 1):
        if not any(str(cell or "").strip() for cell in row):
            continue

        def get_cell(column_name):
            index = column_map[column_name]
            return row[index] if index < len(row) else ""

        placement_code = str(get_cell("placement_code") or "").strip()
        jmoa_code = str(get_cell("jmoa_code") or "").strip()
        kibetu = str(get_cell("kibetu") or "").strip()

        if not placement_code:
            errors.append(f"{row_number}行目: 上位者コードは必須です。")
        if not jmoa_code:
            errors.append(f"{row_number}行目: 会員コードは必須です。")
        if not kibetu:
            errors.append(f"{row_number}行目: 期別は必須です。")

        bv = parse_import_int(get_cell("bv"), row_number, "BV", errors)
        carry_over_bv = parse_import_int(
            get_cell("carry_over_bv"),
            row_number,
            "繰り越しBV",
            errors,
        )

        if placement_code and jmoa_code and kibetu:
            data_rows.append([
                placement_code,
                jmoa_code,
                bv,
                carry_over_bv,
                kibetu,
            ])

    if not data_rows and not errors:
        errors.append("登録対象データがありません。")

    return data_rows, errors


def fetch_latest_registered_kibetu():
    """登録済みの最新期別を返す。idx_kibetu の MIN/MAX 最適化が効くため即時に返る。"""
    sql = f"SELECT MAX(kibetu) FROM {BASIC_BV_LINE_TABLE}"
    with connections["rds"].cursor() as cursor:
        try:
            cursor.execute(sql)
            row = cursor.fetchone()
        except ProgrammingError as exc:
            if is_missing_table_error(exc):
                logger.info("繰り越し業績テーブルが未作成です。table=%s", BASIC_BV_LINE_TABLE)
                return ""
            raise
    return (row[0] or "") if row else ""


def is_registered_kibetu(kibetu):
    """指定期別に明細があるかを判定する。全件集計せず idx_kibetu だけを見る。"""
    if not kibetu:
        return False

    sql = f"SELECT 1 FROM {BASIC_BV_LINE_TABLE} WHERE kibetu = %s LIMIT 1"
    with connections["rds"].cursor() as cursor:
        try:
            cursor.execute(sql, [kibetu])
            return cursor.fetchone() is not None
        except ProgrammingError as exc:
            if is_missing_table_error(exc):
                logger.info("繰り越し業績テーブルが未作成です。table=%s", BASIC_BV_LINE_TABLE)
                return False
            raise


def clear_carry_over_history_cache():
    cache.delete(CARRY_OVER_HISTORY_CACHE_KEY)


def fetch_carry_over_history_rows(use_cache=True):
    if use_cache:
        cached_rows = cache.get(CARRY_OVER_HISTORY_CACHE_KEY)
        if cached_rows is not None:
            return cached_rows

    rows = _query_carry_over_history_rows()
    cache.set(CARRY_OVER_HISTORY_CACHE_KEY, rows, CARRY_OVER_HISTORY_CACHE_TIMEOUT)
    return rows


def _query_carry_over_history_rows():
    # COUNT(*) と MIN/MAX(created_at) を1つの集計にまとめると
    # idx_kibetu_created のルースインデックススキャンが使えず全件走査になる。
    # 件数と日時を別々の派生表にすることで、
    # 件数は idx_kibetu のカバリングスキャン、日時は "Using index for group-by" になる。
    sql = f"""
        SELECT
            c.kibetu,
            c.row_count,
            t.created_at,
            t.updated_at,
            pm.st_date,
            pm.end_date,
            pm.completion_date,
            NULL AS year,
            NULL AS month,
            NULL AS payment_date
        FROM (
            SELECT
                kibetu,
                COUNT(*) AS row_count
            FROM {BASIC_BV_LINE_TABLE}
            GROUP BY kibetu
        ) AS c
        JOIN (
            SELECT
                kibetu,
                MIN(created_at) AS created_at,
                MAX(created_at) AS updated_at
            FROM {BASIC_BV_LINE_TABLE}
            GROUP BY kibetu
        ) AS t
          ON t.kibetu = c.kibetu
        LEFT JOIN bonus_db.period_master AS pm
          ON pm.kibetu = c.kibetu
        ORDER BY c.kibetu DESC
    """
    with connections["rds"].cursor() as cursor:
        logger.info("繰り越し業績履歴取得SQLを実行します。table=%s", BASIC_BV_LINE_TABLE)
        try:
            cursor.execute(sql)
            cols = [col[0] for col in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]
        except ProgrammingError as exc:
            if is_missing_table_error(exc):
                logger.info("繰り越し業績テーブルが未作成です。table=%s", BASIC_BV_LINE_TABLE)
                return []
            raise


class CarryOverPerformanceTemplateView(generic.View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "繰り越しBV"
        ws.append(["期別", "上位者コード", "会員コード", "BV", "繰り越しBV"])
        ws.append(["2025C03", "JP00000001", "JP05215357", 100, 50])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="carry_over_performance_template.xlsx"'
        )
        wb.save(response)
        return response


class CarryOverPerformanceView(KeysetPaginationMixin, generic.TemplateView):
    template_name = "business_carry_over_performance.html"

    def _delete_by_kibetu(self, request):
        q_kibetu = (request.POST.get("q_kibetu") or "").strip()
        if not q_kibetu:
            messages.error(request, "期別を選択してください。")
            return redirect("connect:business_carry_over_performance")
        if not is_registered_kibetu(q_kibetu):
            messages.error(request, "登録されている期別のみ削除できます。")
            return redirect("connect:business_carry_over_performance")

        delete_sql = f"DELETE FROM {BASIC_BV_LINE_TABLE} WHERE kibetu = %s"

        try:
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    logger.info(
                        "繰り越しBV期別削除SQLを実行します。kibetu=%s table=%s",
                        q_kibetu,
                        BASIC_BV_LINE_TABLE,
                    )
                    cursor.execute(delete_sql, [q_kibetu])
                    deleted_count = cursor.rowcount

            clear_carry_over_history_cache()

            record_change_audit(
                request,
                screen_name="繰り越し業績照会",
                action_type="bulk_delete",
                target_table="basic_bv_line",
                target_pk=q_kibetu,
                summary=f"期別 {q_kibetu} の繰り越しBVを削除: {deleted_count}件",
                before_values={"kibetu": q_kibetu, "count": deleted_count},
                after_values=None,
            )

            if deleted_count:
                messages.success(
                    request,
                    f"期別 {q_kibetu} の繰り越しBVを {deleted_count} 件削除しました。",
                )
            else:
                messages.warning(
                    request,
                    f"期別 {q_kibetu} に削除対象データはありません。",
                )
        except Exception as exc:
            logger.exception("繰り越しBV期別削除エラー")
            messages.error(request, f"削除中にエラーが発生しました: {exc}")

        return redirect("connect:business_carry_over_performance")

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip()
        if action == "delete":
            return self._delete_by_kibetu(request)

        uploaded_file = request.FILES.get("carry_over_file")
        if not uploaded_file:
            messages.error(request, "取込ファイルを選択してください。")
            return redirect("connect:business_carry_over_performance")

        try:
            rows, errors = parse_basic_bv_line_import_rows(uploaded_file)
            if errors:
                for error in errors[:10]:
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f"他 {len(errors) - 10} 件のエラーがあります。")
                return redirect("connect:business_carry_over_performance")

            insert_sql = f"""
                INSERT INTO {BASIC_BV_LINE_TABLE} (
                    placement_code,
                    jmoa_code,
                    bv,
                    carry_over_bv,
                    kibetu,
                    created_at
                ) VALUES (
                    %s, %s, %s, %s, %s, CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
                )
                ON DUPLICATE KEY UPDATE
                    bv = VALUES(bv),
                    carry_over_bv = VALUES(carry_over_bv),
                    created_at = CONVERT_TZ(NOW(), 'UTC', 'Asia/Tokyo')
            """
            with transaction.atomic(using="rds"):
                with connections["rds"].cursor() as cursor:
                    logger.info("繰り越しBV一括登録SQLを実行します。table=%s", BASIC_BV_LINE_TABLE)
                    cursor.executemany(insert_sql, rows)

            clear_carry_over_history_cache()

            messages.success(request, f"{len(rows)}件を繰り越しBV管理に一括登録しました。")
        except Exception as exc:
            logger.exception("繰り越しBV一括登録エラー")
            messages.error(request, f"登録中にエラーが発生しました: {exc}")

        return redirect("connect:business_carry_over_performance")

    @staticmethod
    def resolve_default_kibetu(q_kibetu="", q_placement_code="", q_jmoa_code=""):
        """条件無しのときに初期表示する期別を決める。

        basic_bv_line は数百万件あり、期別を絞らずに並べ替えると全件 filesort になって
        画面が開かない。会員コード等の条件も無い場合は最新の登録期別を初期値にする。
        """
        if q_kibetu or q_placement_code or q_jmoa_code:
            return q_kibetu, False

        latest_kibetu = fetch_latest_registered_kibetu()
        return latest_kibetu, bool(latest_kibetu)

    def _build_where(self, q_kibetu="", q_placement_code="", q_jmoa_code=""):
        where = []
        params = []

        if q_kibetu:
            where.append("kibetu = %s")
            params.append(q_kibetu)

        if q_placement_code:
            where.append("placement_code LIKE %s")
            params.append(f"%{q_placement_code}%")

        if q_jmoa_code:
            where.append("jmoa_code LIKE %s")
            params.append(f"%{q_jmoa_code}%")

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        return where_sql, params

    def _fetch_total_count(self, q_kibetu="", q_placement_code="", q_jmoa_code=""):
        where_sql, params = self._build_where(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
        )
        sql = f"""
            SELECT COUNT(*)
            FROM {BASIC_BV_LINE_TABLE}
            {where_sql}
        """
        with connections["rds"].cursor() as cursor:
            logger.info("繰り越し業績件数取得SQLを実行します。table=%s", BASIC_BV_LINE_TABLE)
            try:
                cursor.execute(sql, params)
                row = cursor.fetchone()
            except ProgrammingError as exc:
                if is_missing_table_error(exc):
                    logger.info("繰り越し業績テーブルが未作成です。table=%s", BASIC_BV_LINE_TABLE)
                    return 0
                raise
        return int(row[0]) if row else 0

    def _fetch_rows(self, q_kibetu="", q_placement_code="", q_jmoa_code="", limit=200, offset=0):
        where_sql, params = self._build_where(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
        )
        sql = f"""
            SELECT
                id,
                kibetu,
                placement_code,
                jmoa_code,
                bv,
                carry_over_bv,
                created_at
            FROM {BASIC_BV_LINE_TABLE}
            {where_sql}
            ORDER BY kibetu DESC, placement_code, jmoa_code
            LIMIT %s OFFSET %s
        """
        with connections["rds"].cursor() as cursor:
            logger.info("繰り越し業績一覧取得SQLを実行します。table=%s", BASIC_BV_LINE_TABLE)
            try:
                cursor.execute(sql, params + [limit, offset])
                cols = [col[0] for col in cursor.description]
                return [dict(zip(cols, row)) for row in cursor.fetchall()]
            except ProgrammingError as exc:
                if is_missing_table_error(exc):
                    logger.info("繰り越し業績テーブルが未作成です。table=%s", BASIC_BV_LINE_TABLE)
                    return []
                raise

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q_kibetu = (self.request.GET.get("q_kibetu") or "").strip()
        q_placement_code = (self.request.GET.get("q_placement_code") or "").strip()
        q_jmoa_code = (self.request.GET.get("q_jmoa_code") or "").strip()
        kibetu_choice_mode = self.request.GET.get("kibetu_choice_mode") or "recent"

        if q_kibetu and not is_registered_kibetu(q_kibetu):
            messages.warning(self.request, "登録されている期別のみ選択できます。")
            q_kibetu = ""

        q_kibetu, q_kibetu_defaulted = self.resolve_default_kibetu(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
        )
        registration_history_rows = fetch_carry_over_history_rows()

        per_page = self.get_per_page()
        total_count = self._fetch_total_count(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
        )
        total_pages = max(1, math.ceil(total_count / per_page))
        page = self.get_page_number(total_pages)
        offset = (page - 1) * per_page

        rows = self._fetch_rows(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
            limit=per_page,
            offset=offset,
        )

        ctx["q_kibetu"] = q_kibetu
        ctx["q_kibetu_defaulted"] = q_kibetu_defaulted
        ctx["q_placement_code"] = q_placement_code
        ctx["q_jmoa_code"] = q_jmoa_code
        ctx["active_menu"] = "business_carry_over_performance"
        ctx["registration_history_rows"] = registration_history_rows
        ctx["registration_history_modal_id"] = "carryOverPerformanceHistoryModal"
        ctx["registration_modal_title"] = "登録履歴（繰り越し業績照会）"
        ctx["registration_target_url_name"] = "connect:business_carry_over_performance"
        return self.set_page_context(
            ctx=ctx,
            rows=rows,
            per_page=per_page,
            total_count=total_count,
            total_pages=total_pages,
            page=page,
            base_params={
                "q_kibetu": q_kibetu,
                "q_placement_code": q_placement_code,
                "q_jmoa_code": q_jmoa_code,
                "per_page": per_page,
                "kibetu_choice_mode": kibetu_choice_mode,
            },
        )


class CarryOverPerformanceExportView(CarryOverPerformanceView):
    def get(self, request, *args, **kwargs):
        q_kibetu = (request.GET.get("q_kibetu") or "").strip()
        q_placement_code = (request.GET.get("q_placement_code") or "").strip()
        q_jmoa_code = (request.GET.get("q_jmoa_code") or "").strip()
        if q_kibetu and not is_registered_kibetu(q_kibetu):
            q_kibetu = ""

        q_kibetu, _ = self.resolve_default_kibetu(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
        )

        rows = self._fetch_rows(
            q_kibetu=q_kibetu,
            q_placement_code=q_placement_code,
            q_jmoa_code=q_jmoa_code,
            limit=1000000,
            offset=0,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "繰り越し業績照会"
        ws.append([
            "期別",
            "上位者コード",
            "会員コード",
            "BV",
            "繰り越しBV",
            "作成日時",
        ])

        for row in rows:
            created_at = row.get("created_at")
            if created_at is not None and hasattr(created_at, "strftime"):
                created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
            ws.append([
                row.get("kibetu"),
                row.get("placement_code"),
                row.get("jmoa_code"),
                row.get("bv") or 0,
                row.get("carry_over_bv") or 0,
                created_at,
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="business_carry_over_performance.xlsx"'
        )
        wb.save(response)
        return response
